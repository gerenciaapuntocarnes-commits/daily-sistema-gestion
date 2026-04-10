from fastapi import APIRouter, HTTPException, UploadFile, File
from pydantic import BaseModel
from typing import Optional, List
from datetime import date, datetime
import uuid
import re
import psycopg2
from database import get_conn

router = APIRouter()

# Cache de columnas para evitar queries a information_schema en cada request
_col_cache: dict = {}

def _has_col(cur, table: str, col: str) -> bool:
    key = f"{table}.{col}"
    if key not in _col_cache:
        cur.execute(
            "SELECT 1 FROM information_schema.columns WHERE table_name=%s AND column_name=%s LIMIT 1",
            (table, col)
        )
        _col_cache[key] = cur.fetchone() is not None
    return _col_cache[key]

# ═══════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════

def _gen_lote_produccion(new_id: int) -> str:
    return f"LOT-{datetime.now().strftime('%Y%m%d')}-{new_id:06d}"

def _gen_lote_item(new_id: int) -> str:
    return f"LMP-{datetime.now().strftime('%Y%m%d')}-{new_id:06d}"

def _gen_numero_remision(new_id: int) -> str:
    return f"REM-{datetime.now().strftime('%Y%m%d')}-{new_id:06d}"

# ═══════════════════════════════════════════════════════════════
# MATERIAS PRIMAS
# ═══════════════════════════════════════════════════════════════

class MPIn(BaseModel):
    nombre: str
    codigo: Optional[str] = None
    unidad: str = "kg"
    categoria: str = "General"

class CompraIn(BaseModel):
    mp_id: int
    fecha: date
    proveedor: Optional[str] = None
    cantidad: float
    precio_unit: float
    factura: Optional[str] = None
    notas: Optional[str] = None

@router.get("/materias-primas/check-codigo")
def check_codigo(codigo: str):
    """Verifica en tiempo real si un código ya existe."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "SELECT id FROM materias_primas WHERE codigo=%s AND activo=TRUE",
        (codigo.strip(),)
    )
    existe = cur.fetchone() is not None
    cur.close(); conn.close()
    return {"existe": existe}

@router.get("/materias-primas")
def listar_mp():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT mp.id, mp.codigo, mp.nombre, mp.unidad, mp.categoria, mp.activo,
               c.precio_unit AS precio_actual,
               c.fecha       AS ultima_compra,
               c.proveedor   AS ultimo_proveedor
        FROM materias_primas mp
        LEFT JOIN LATERAL (
            SELECT precio_unit, fecha, proveedor
            FROM compras_mp
            WHERE mp_id = mp.id
            ORDER BY fecha DESC, id DESC
            LIMIT 1
        ) c ON TRUE
        WHERE mp.activo = TRUE
        ORDER BY mp.categoria, mp.nombre
    """)
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id": r[0], "codigo": r[1], "nombre": r[2], "unidad": r[3],
             "categoria": r[4], "activo": r[5],
             "precio_actual": float(r[6]) if r[6] else 0,
             "ultima_compra": str(r[7]) if r[7] else None,
             "ultimo_proveedor": r[8]} for r in rows]

@router.post("/materias-primas")
def crear_mp(data: MPIn):
    conn = get_conn()
    cur = conn.cursor()
    codigo = data.codigo.strip() if data.codigo and data.codigo.strip() else None
    try:
        cur.execute(
            "INSERT INTO materias_primas (codigo, nombre, unidad, categoria) VALUES (%s,%s,%s,%s) RETURNING id",
            (codigo, data.nombre, data.unidad, data.categoria)
        )
        new_id = cur.fetchone()[0]
        conn.commit()
    except psycopg2.errors.UniqueViolation:
        conn.rollback()
        raise HTTPException(409, f"El código '{codigo}' ya está en uso por otra materia prima")
    finally:
        cur.close(); conn.close()
    return {"id": new_id, "ok": True}

@router.put("/materias-primas/{mp_id}")
def editar_mp(mp_id: int, data: MPIn):
    conn = get_conn()
    cur = conn.cursor()
    codigo = data.codigo.strip() if data.codigo and data.codigo.strip() else None
    try:
        cur.execute(
            "UPDATE materias_primas SET codigo=%s, nombre=%s, unidad=%s, categoria=%s WHERE id=%s",
            (codigo, data.nombre, data.unidad, data.categoria, mp_id)
        )
        conn.commit()
    except psycopg2.errors.UniqueViolation:
        conn.rollback()
        raise HTTPException(409, f"El código '{codigo}' ya está en uso por otra materia prima")
    finally:
        cur.close(); conn.close()
    return {"ok": True}

@router.delete("/materias-primas/{mp_id}")
def eliminar_mp(mp_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("UPDATE materias_primas SET activo=FALSE WHERE id=%s", (mp_id,))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

@router.post("/materias-primas/sync-siigo")
def sync_mp_desde_siigo():
    """Importa MP e Ingredientes de Siigo a la tabla materias_primas."""
    try:
        from siigo import fetch_products
        mp_siigo = fetch_products(tipo="mp")
    except Exception as e:
        raise HTTPException(500, f"Error Siigo: {str(e)}")
    conn = get_conn()
    cur = conn.cursor()
    creados = 0
    actualizados = 0
    for p in mp_siigo:
        code = p["code"]
        nombre = p["name"]
        unidad = p["unit"] or "unidad"
        categoria = p["group"] or "General"
        # Check if exists by codigo
        cur.execute("SELECT id FROM materias_primas WHERE codigo=%s", (code,))
        existing = cur.fetchone()
        if existing:
            cur.execute(
                "UPDATE materias_primas SET nombre=%s, categoria=%s, activo=TRUE WHERE id=%s",
                (nombre, categoria, existing[0])
            )
            actualizados += 1
        else:
            cur.execute(
                "INSERT INTO materias_primas (codigo, nombre, unidad, categoria) VALUES (%s,%s,%s,%s)",
                (code, nombre, unidad, categoria)
            )
            creados += 1
    conn.commit(); cur.close(); conn.close()
    return {"creados": creados, "actualizados": actualizados, "total": len(mp_siigo), "ok": True}

@router.get("/compras-mp")
def listar_compras(mp_id: Optional[int] = None, limit: int = 100):
    conn = get_conn()
    cur = conn.cursor()
    if mp_id:
        cur.execute("""
            SELECT c.id, c.fecha, mp.nombre, mp.unidad, c.proveedor,
                   c.cantidad, c.precio_unit, c.cantidad*c.precio_unit AS total,
                   c.factura, c.notas
            FROM compras_mp c JOIN materias_primas mp ON mp.id=c.mp_id
            WHERE c.mp_id=%s ORDER BY c.fecha DESC LIMIT %s
        """, (mp_id, limit))
    else:
        cur.execute("""
            SELECT c.id, c.fecha, mp.nombre, mp.unidad, c.proveedor,
                   c.cantidad, c.precio_unit, c.cantidad*c.precio_unit AS total,
                   c.factura, c.notas
            FROM compras_mp c JOIN materias_primas mp ON mp.id=c.mp_id
            ORDER BY c.fecha DESC LIMIT %s
        """, (limit,))
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id": r[0], "fecha": str(r[1]), "mp_nombre": r[2], "unidad": r[3],
             "proveedor": r[4], "cantidad": float(r[5]), "precio_unit": float(r[6]),
             "total": float(r[7]), "factura": r[8], "notas": r[9]} for r in rows]

@router.post("/compras-mp")
def registrar_compra(data: CompraIn):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO compras_mp (mp_id, fecha, proveedor, cantidad, precio_unit, factura, notas)
        VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id
    """, (data.mp_id, data.fecha, data.proveedor, data.cantidad,
          data.precio_unit, data.factura, data.notas))
    new_id = cur.fetchone()[0]
    conn.commit()
    # Auto-update inventory
    try:
        cur.execute("""
            INSERT INTO inventario (mp_id, cantidad_actual, updated_at)
            VALUES (%s, %s, NOW())
            ON CONFLICT (mp_id) DO UPDATE SET cantidad_actual = inventario.cantidad_actual + %s, updated_at = NOW()
        """, (data.mp_id, data.cantidad, data.cantidad))
        conn.commit()
    except:
        pass
    cur.close(); conn.close()
    return {"id": new_id, "ok": True}

@router.delete("/compras-mp/{compra_id}")
def eliminar_compra(compra_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM compras_mp WHERE id=%s", (compra_id,))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

# ═══════════════════════════════════════════════════════════════
# RECETAS
# ═══════════════════════════════════════════════════════════════

class RecetaIn(BaseModel):
    nombre: str
    categoria: str = "General"
    descripcion: Optional[str] = None
    porciones: float = 1
    precio_venta: float = 0
    batch_maximo: float = 0
    tiempo_batch_min: int = 0
    vida_util_dias: int = 30
    costo_mano_obra: float = 0
    costo_servicios: float = 0

class IngredienteIn(BaseModel):
    mp_id: int
    cantidad: float

@router.get("/recetas")
def listar_recetas():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT r.id, r.nombre, r.categoria, r.descripcion, r.porciones, r.precio_venta, r.activo,
               COALESCE(r.batch_maximo, 0), COALESCE(r.tiempo_batch_min, 0),
               COALESCE(r.vida_util_dias, 30), COALESCE(r.costo_mano_obra, 0),
               COALESCE(r.costo_servicios, 0)
        FROM recetas r WHERE r.activo=TRUE ORDER BY r.categoria, r.nombre
    """)
    recetas = cur.fetchall()
    if not recetas:
        cur.close(); conn.close()
        return []
    receta_ids = [r[0] for r in recetas]
    cur.execute("""
        SELECT ri.receta_id, COALESCE(SUM(ri.cantidad * COALESCE(ult.precio_unit, 0)), 0)
        FROM receta_ingredientes ri
        LEFT JOIN LATERAL (
            SELECT precio_unit FROM compras_mp
            WHERE mp_id = ri.mp_id ORDER BY fecha DESC, id DESC LIMIT 1
        ) ult ON TRUE
        WHERE ri.receta_id = ANY(%s)
        GROUP BY ri.receta_id
    """, (receta_ids,))
    costos = {row[0]: float(row[1]) for row in cur.fetchall()}
    result = []
    for r in recetas:
        costo = costos.get(r[0], 0.0)
        porciones = float(r[4]) if r[4] else 1
        precio_venta = float(r[5]) if r[5] else 0
        costo_mp = costo / porciones if porciones > 0 else 0
        batch_max = float(r[7])
        costo_mo = float(r[10])
        costo_sv = float(r[11])
        costo_fijo_und = (costo_mo + costo_sv) / batch_max if batch_max > 0 else 0
        costo_porcion = costo_mp + costo_fijo_und
        margen = precio_venta - costo_porcion
        margen_pct = (margen / precio_venta * 100) if precio_venta > 0 else 0
        result.append({
            "id": r[0], "nombre": r[1], "categoria": r[2], "descripcion": r[3],
            "porciones": porciones, "precio_venta": precio_venta, "activo": r[6],
            "batch_maximo": batch_max, "tiempo_batch_min": int(r[8]),
            "vida_util_dias": int(r[9]), "costo_mano_obra": costo_mo,
            "costo_servicios": costo_sv, "costo_fijo_batch": costo_mo + costo_sv,
            "costo_total": costo, "costo_porcion": costo_porcion,
            "margen": margen, "margen_pct": margen_pct
        })
    cur.close(); conn.close()
    return result

@router.get("/recetas/{receta_id}")
def detalle_receta(receta_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, nombre, categoria, descripcion, porciones, precio_venta,
               COALESCE(batch_maximo,0), COALESCE(tiempo_batch_min,0),
               COALESCE(vida_util_dias,30), COALESCE(costo_mano_obra,0),
               COALESCE(costo_servicios,0)
        FROM recetas WHERE id=%s
    """, (receta_id,))
    r = cur.fetchone()
    if not r:
        raise HTTPException(404, "Receta no encontrada")
    cur.execute("""
        SELECT ri.id, mp.id, mp.nombre, mp.unidad, ri.cantidad,
               COALESCE(ult.precio_unit,0) AS precio_unit,
               ri.cantidad * COALESCE(ult.precio_unit,0) AS costo_linea
        FROM receta_ingredientes ri
        JOIN materias_primas mp ON mp.id=ri.mp_id
        LEFT JOIN LATERAL (
            SELECT precio_unit FROM compras_mp
            WHERE mp_id=ri.mp_id ORDER BY fecha DESC, id DESC LIMIT 1
        ) ult ON TRUE
        WHERE ri.receta_id=%s
    """, (receta_id,))
    ingredientes = [{"id": i[0], "mp_id": i[1], "mp_nombre": i[2], "unidad": i[3],
                     "cantidad": float(i[4]), "precio_unit": float(i[5]),
                     "costo_linea": float(i[6])} for i in cur.fetchall()]
    cur.close(); conn.close()
    porciones = float(r[4]) if r[4] else 1
    precio_venta = float(r[5]) if r[5] else 0
    costo_mp_total = sum(i["costo_linea"] for i in ingredientes)
    batch_max = float(r[6])
    costo_mo = float(r[9])
    costo_sv = float(r[10])
    costo_fijo = costo_mo + costo_sv
    costo_fijo_und = costo_fijo / batch_max if batch_max > 0 else 0
    costo_mp_und = costo_mp_total / porciones if porciones > 0 else 0
    costo_porcion = costo_mp_und + costo_fijo_und
    margen = precio_venta - costo_porcion
    return {
        "id": r[0], "nombre": r[1], "categoria": r[2], "descripcion": r[3],
        "porciones": porciones, "precio_venta": precio_venta,
        "batch_maximo": batch_max, "tiempo_batch_min": int(r[7]),
        "vida_util_dias": int(r[8]), "costo_mano_obra": costo_mo,
        "costo_servicios": costo_sv, "costo_fijo_batch": costo_fijo,
        "ingredientes": ingredientes, "costo_total": costo_mp_total,
        "costo_mp_und": costo_mp_und, "costo_fijo_und": costo_fijo_und,
        "costo_porcion": costo_porcion,
        "margen": margen,
        "margen_pct": (margen / precio_venta * 100) if precio_venta > 0 else 0
    }

@router.post("/recetas")
def crear_receta(data: RecetaIn):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO recetas (nombre, categoria, descripcion, porciones, precio_venta,
                             batch_maximo, tiempo_batch_min, vida_util_dias, costo_mano_obra, costo_servicios)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id
    """, (data.nombre, data.categoria, data.descripcion, data.porciones, data.precio_venta,
          data.batch_maximo, data.tiempo_batch_min, data.vida_util_dias, data.costo_mano_obra, data.costo_servicios))
    new_id = cur.fetchone()[0]
    conn.commit(); cur.close(); conn.close()
    return {"id": new_id, "ok": True}

@router.put("/recetas/{receta_id}")
def editar_receta(receta_id: int, data: RecetaIn):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        UPDATE recetas SET nombre=%s, categoria=%s, descripcion=%s,
               porciones=%s, precio_venta=%s,
               batch_maximo=%s, tiempo_batch_min=%s, vida_util_dias=%s,
               costo_mano_obra=%s, costo_servicios=%s
        WHERE id=%s
    """, (data.nombre, data.categoria, data.descripcion,
          data.porciones, data.precio_venta,
          data.batch_maximo, data.tiempo_batch_min, data.vida_util_dias,
          data.costo_mano_obra, data.costo_servicios, receta_id))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

@router.delete("/recetas/{receta_id}")
def eliminar_receta(receta_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("UPDATE recetas SET activo=FALSE WHERE id=%s", (receta_id,))
    cur.execute("UPDATE producto_receta SET receta_id=NULL WHERE receta_id=%s", (receta_id,))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

@router.post("/recetas/{receta_id}/ingredientes")
def agregar_ingrediente(receta_id: int, data: IngredienteIn):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO receta_ingredientes (receta_id, mp_id, cantidad)
        VALUES (%s,%s,%s) RETURNING id
    """, (receta_id, data.mp_id, data.cantidad))
    new_id = cur.fetchone()[0]
    conn.commit(); cur.close(); conn.close()
    return {"id": new_id, "ok": True}

@router.delete("/receta-ingredientes/{ing_id}")
def eliminar_ingrediente(ing_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM receta_ingredientes WHERE id=%s", (ing_id,))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

# ═══════════════════════════════════════════════════════════════
# PRODUCCIÓN  — lote generado automáticamente por el servidor
# ═══════════════════════════════════════════════════════════════

class ProduccionIn(BaseModel):
    fecha: date
    receta_id: int
    porciones_planeadas: float = 0
    porciones: float
    operario: Optional[str] = None
    notas: Optional[str] = None

@router.get("/produccion")
def listar_produccion(limit: int = 200):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT p.id, p.fecha, r.nombre AS receta, r.categoria,
               p.porciones_planeadas, p.porciones, p.lote, p.operario, p.notas,
               COALESCE(costo.val,0) * p.porciones AS costo_total
        FROM produccion p
        LEFT JOIN recetas r ON r.id=p.receta_id
        LEFT JOIN LATERAL (
            SELECT SUM(ri.cantidad * COALESCE(ult.precio_unit,0)) / NULLIF(r.porciones,0) AS val
            FROM receta_ingredientes ri
            LEFT JOIN LATERAL (
                SELECT precio_unit FROM compras_mp
                WHERE mp_id=ri.mp_id ORDER BY fecha DESC, id DESC LIMIT 1
            ) ult ON TRUE
            WHERE ri.receta_id=p.receta_id
        ) costo ON TRUE
        ORDER BY p.fecha DESC, p.id DESC LIMIT %s
    """, (limit,))
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id": r[0], "fecha": str(r[1]), "receta": r[2], "categoria": r[3],
             "porciones_planeadas": float(r[4]) if r[4] else 0,
             "porciones": float(r[5]), "lote": r[6], "operario": r[7],
             "notas": r[8], "costo_total": float(r[9]) if r[9] else 0} for r in rows]

@router.post("/produccion")
def registrar_produccion(data: ProduccionIn):
    conn = get_conn()
    cur = conn.cursor()
    # Insertar con lote temporal único
    temp_lote = f"TEMP-{uuid.uuid4()}"
    cur.execute("""
        INSERT INTO produccion (fecha, receta_id, porciones_planeadas, porciones, lote, operario, notas)
        VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id
    """, (data.fecha, data.receta_id, data.porciones_planeadas, data.porciones,
          temp_lote, data.operario, data.notas))
    new_id = cur.fetchone()[0]
    # Generar lote definitivo con el ID real
    lote = _gen_lote_produccion(new_id)
    cur.execute("UPDATE produccion SET lote=%s WHERE id=%s", (lote, new_id))
    conn.commit(); cur.close(); conn.close()
    return {"id": new_id, "lote": lote, "ok": True}

@router.delete("/produccion/{prod_id}")
def eliminar_produccion(prod_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM produccion WHERE id=%s", (prod_id,))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

# ═══════════════════════════════════════════════════════════════
# GASTOS
# ═══════════════════════════════════════════════════════════════

class GastoIn(BaseModel):
    fecha: date
    tipo: str = "caja_menor"
    categoria: str = "General"
    descripcion: str
    monto: float
    responsable: Optional[str] = None
    comprobante: Optional[str] = None
    notas: Optional[str] = None

@router.get("/gastos")
def listar_gastos(tipo: Optional[str] = None, limit: int = 200):
    conn = get_conn()
    cur = conn.cursor()
    if tipo:
        cur.execute("""
            SELECT id, fecha, tipo, categoria, descripcion, monto, responsable, comprobante, notas
            FROM gastos WHERE tipo=%s ORDER BY fecha DESC, id DESC LIMIT %s
        """, (tipo, limit))
    else:
        cur.execute("""
            SELECT id, fecha, tipo, categoria, descripcion, monto, responsable, comprobante, notas
            FROM gastos ORDER BY fecha DESC, id DESC LIMIT %s
        """, (limit,))
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id": r[0], "fecha": str(r[1]), "tipo": r[2], "categoria": r[3],
             "descripcion": r[4], "monto": float(r[5]), "responsable": r[6],
             "comprobante": r[7], "notas": r[8]} for r in rows]

@router.post("/gastos")
def registrar_gasto(data: GastoIn):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO gastos (fecha, tipo, categoria, descripcion, monto, responsable, comprobante, notas)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id
    """, (data.fecha, data.tipo, data.categoria, data.descripcion,
          data.monto, data.responsable, data.comprobante, data.notas))
    new_id = cur.fetchone()[0]
    conn.commit(); cur.close(); conn.close()
    return {"id": new_id, "ok": True}

@router.delete("/gastos/{gasto_id}")
def eliminar_gasto(gasto_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM gastos WHERE id=%s", (gasto_id,))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

# ═══════════════════════════════════════════════════════════════
# PROCEDIMIENTOS
# ═══════════════════════════════════════════════════════════════

class ProcedimientoIn(BaseModel):
    nombre: str
    categoria: str = "General"
    descripcion: Optional[str] = None
    pasos: Optional[str] = None
    responsable: Optional[str] = None
    frecuencia: str = "Cada vez"

@router.get("/procedimientos")
def listar_procedimientos():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, nombre, categoria, descripcion, pasos, responsable, frecuencia, activo
        FROM procedimientos WHERE activo=TRUE ORDER BY categoria, nombre
    """)
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id": r[0], "nombre": r[1], "categoria": r[2], "descripcion": r[3],
             "pasos": r[4], "responsable": r[5], "frecuencia": r[6], "activo": r[7]} for r in rows]

@router.post("/procedimientos")
def crear_procedimiento(data: ProcedimientoIn):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO procedimientos (nombre, categoria, descripcion, pasos, responsable, frecuencia)
        VALUES (%s,%s,%s,%s,%s,%s) RETURNING id
    """, (data.nombre, data.categoria, data.descripcion,
          data.pasos, data.responsable, data.frecuencia))
    new_id = cur.fetchone()[0]
    conn.commit(); cur.close(); conn.close()
    return {"id": new_id, "ok": True}

@router.put("/procedimientos/{proc_id}")
def editar_procedimiento(proc_id: int, data: ProcedimientoIn):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        UPDATE procedimientos SET nombre=%s, categoria=%s, descripcion=%s,
               pasos=%s, responsable=%s, frecuencia=%s WHERE id=%s
    """, (data.nombre, data.categoria, data.descripcion,
          data.pasos, data.responsable, data.frecuencia, proc_id))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

@router.delete("/procedimientos/{proc_id}")
def eliminar_procedimiento(proc_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("UPDATE procedimientos SET activo=FALSE WHERE id=%s", (proc_id,))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

# ═══════════════════════════════════════════════════════════════
# REGISTROS SANITARIOS
# ═══════════════════════════════════════════════════════════════

class RegistroSanitarioIn(BaseModel):
    fecha: date
    tipo: str = "Control temperatura"
    descripcion: str
    resultado: str = "Aprobado"
    operario: Optional[str] = None
    observaciones: Optional[str] = None
    proxima_revision: Optional[date] = None

@router.get("/registros-sanitarios")
def listar_registros(limit: int = 200):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, fecha, tipo, descripcion, resultado, operario, observaciones, proxima_revision
        FROM registros_sanitarios ORDER BY fecha DESC, id DESC LIMIT %s
    """, (limit,))
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id": r[0], "fecha": str(r[1]), "tipo": r[2], "descripcion": r[3],
             "resultado": r[4], "operario": r[5], "observaciones": r[6],
             "proxima_revision": str(r[7]) if r[7] else None} for r in rows]

@router.post("/registros-sanitarios")
def crear_registro(data: RegistroSanitarioIn):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO registros_sanitarios
            (fecha, tipo, descripcion, resultado, operario, observaciones, proxima_revision)
        VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id
    """, (data.fecha, data.tipo, data.descripcion, data.resultado,
          data.operario, data.observaciones, data.proxima_revision))
    new_id = cur.fetchone()[0]
    conn.commit(); cur.close(); conn.close()
    return {"id": new_id, "ok": True}

@router.delete("/registros-sanitarios/{reg_id}")
def eliminar_registro(reg_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM registros_sanitarios WHERE id=%s", (reg_id,))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

# ═══════════════════════════════════════════════════════════════
# REMISIONES — flujo de aprobación
# ═══════════════════════════════════════════════════════════════

class RemisionItemIn(BaseModel):
    mp_id: int
    mp_nombre: str
    cantidad: float
    precio_unit: float

class RemisionIn(BaseModel):
    fecha: date
    proveedor: str
    operario: str
    notas: Optional[str] = None
    foto: str          # base64 obligatorio
    items: List[RemisionItemIn]

class AprobarIn(BaseModel):
    aprobado_por: str

class RechazarIn(BaseModel):
    rechazado_por: str
    motivo: str

@router.get("/remisiones")
def listar_remisiones(estado: Optional[str] = None, limit: int = 200):
    conn = get_conn()
    cur = conn.cursor()
    # Detect available columns (resultado cacheado tras el primer request)
    has_proveedor = _has_col(cur, 'remisiones', 'proveedor')
    has_operario  = _has_col(cur, 'remisiones', 'operario')

    prov_col = "proveedor" if has_proveedor else "NULL"
    oper_col = "operario" if has_operario else "NULL"
    base_q = f"""
        SELECT id, numero,
               COALESCE(fecha, creado_en::date) as fecha,
               {prov_col} as proveedor, {oper_col} as operario,
               notas, estado, aprobado_por, rechazo_motivo, creado_en
        FROM remisiones
    """
    if estado:
        cur.execute(base_q + " WHERE estado=%s ORDER BY creado_en DESC LIMIT %s", (estado, limit))
    else:
        cur.execute(base_q + " ORDER BY creado_en DESC LIMIT %s", (limit,))
    rows = cur.fetchall()

    # Detect remision_items columns (resultado cacheado)
    has_lote      = _has_col(cur, 'remision_items', 'lote')
    has_mp_nombre = _has_col(cur, 'remision_items', 'mp_nombre')
    has_precio    = _has_col(cur, 'remision_items', 'precio_unit')

    result = []
    if rows:
        rem_ids = [r[0] for r in rows]
        lote_col = "lote" if has_lote else "NULL"
        nombre_col = "mp_nombre" if has_mp_nombre else "NULL"
        precio_col = "precio_unit" if has_precio else "0"
        cur.execute(f"""
            SELECT remision_id, {nombre_col}, cantidad, {precio_col}, {lote_col}
            FROM remision_items WHERE remision_id = ANY(%s) ORDER BY remision_id, id
        """, (rem_ids,))
        items_by_rem = {}
        for i in cur.fetchall():
            items_by_rem.setdefault(i[0], []).append({
                "mp_nombre": i[1] or '', "cantidad": float(i[2]),
                "precio_unit": float(i[3]) if i[3] else 0, "lote": i[4]
            })
        for r in rows:
            result.append({
                "id": r[0], "numero": r[1], "fecha": str(r[2]) if r[2] else '',
                "proveedor": r[3] or '', "operario": r[4] or '', "notas": r[5],
                "estado": r[6], "aprobado_por": r[7], "rechazo_motivo": r[8],
                "creado_en": str(r[9]), "items": items_by_rem.get(r[0], [])
            })
    cur.close(); conn.close()
    return result

@router.get("/remisiones/{remision_id}/estado")
def estado_remision(remision_id: int):
    """Endpoint de polling para el operario."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT estado, aprobado_por, rechazo_motivo
        FROM remisiones WHERE id=%s
    """, (remision_id,))
    r = cur.fetchone()
    cur.close(); conn.close()
    if not r:
        raise HTTPException(404, "Remisión no encontrada")
    return {"estado": r[0], "aprobado_por": r[1], "rechazo_motivo": r[2]}

@router.get("/remisiones/{remision_id}/foto")
def foto_remision(remision_id: int):
    """Devuelve la foto de una remisión."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT foto FROM remisiones WHERE id=%s", (remision_id,))
    r = cur.fetchone()
    cur.close(); conn.close()
    if not r:
        raise HTTPException(404, "Remisión no encontrada")
    return {"foto": r[0]}

@router.post("/remisiones")
def crear_remision(data: RemisionIn):
    if not data.items:
        raise HTTPException(400, "Debe agregar al menos un ítem")
    if not data.foto:
        raise HTTPException(400, "La foto de la remisión es obligatoria")

    conn = get_conn()
    cur = conn.cursor()

    # 1. Insertar remisión con número temporal único
    temp_numero = f"TEMP-{uuid.uuid4()}"
    cur.execute("""
        INSERT INTO remisiones (numero, fecha, proveedor, operario, notas, foto, estado)
        VALUES (%s, %s, %s, %s, %s, %s, 'pendiente') RETURNING id
    """, (temp_numero, data.fecha, data.proveedor, data.operario,
          data.notas, data.foto))
    remision_id = cur.fetchone()[0]

    # 2. Generar número definitivo con el ID real
    numero = _gen_numero_remision(remision_id)
    cur.execute("UPDATE remisiones SET numero=%s WHERE id=%s", (numero, remision_id))

    # 3. Insertar ítems con lote generado automáticamente
    for item in data.items:
        temp_lote = f"TEMP-{uuid.uuid4()}"
        cur.execute("""
            INSERT INTO remision_items (remision_id, mp_id, mp_nombre, cantidad, precio_unit, lote)
            VALUES (%s, %s, %s, %s, %s, %s) RETURNING id
        """, (remision_id, item.mp_id, item.mp_nombre,
              item.cantidad, item.precio_unit, temp_lote))
        item_id = cur.fetchone()[0]
        lote_item = _gen_lote_item(item_id)
        cur.execute("UPDATE remision_items SET lote=%s WHERE id=%s", (lote_item, item_id))

    conn.commit(); cur.close(); conn.close()
    return {"id": remision_id, "numero": numero, "estado": "pendiente", "ok": True}

@router.post("/remisiones/{remision_id}/aprobar")
def aprobar_remision(remision_id: int, data: AprobarIn):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        UPDATE remisiones
        SET estado='aprobada', aprobado_por=%s, actualizado_en=NOW()
        WHERE id=%s AND estado='pendiente'
        RETURNING id
    """, (data.aprobado_por, remision_id))
    if not cur.fetchone():
        conn.rollback(); cur.close(); conn.close()
        raise HTTPException(400, "La remisión no existe o ya fue procesada")
    conn.commit(); cur.close(); conn.close()
    return {"ok": True, "estado": "aprobada"}

@router.post("/remisiones/{remision_id}/rechazar")
def rechazar_remision(remision_id: int, data: RechazarIn):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        UPDATE remisiones
        SET estado='rechazada', aprobado_por=%s, rechazo_motivo=%s, actualizado_en=NOW()
        WHERE id=%s AND estado='pendiente'
        RETURNING id
    """, (data.rechazado_por, data.motivo, remision_id))
    if not cur.fetchone():
        conn.rollback(); cur.close(); conn.close()
        raise HTTPException(400, "La remisión no existe o ya fue procesada")
    conn.commit(); cur.close(); conn.close()
    return {"ok": True, "estado": "rechazada"}

@router.delete("/remisiones/{remision_id}")
def eliminar_remision(remision_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM remisiones WHERE id=%s", (remision_id,))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

# ═══════════════════════════════════════════════════════════════
# DASHBOARD
# ═══════════════════════════════════════════════════════════════

@router.get("/dashboard")
def dashboard():
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) FROM materias_primas WHERE activo=TRUE")
    total_mp = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM recetas WHERE activo=TRUE")
    total_recetas = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*), COALESCE(SUM(porciones),0) FROM produccion WHERE fecha=CURRENT_DATE")
    r = cur.fetchone()
    prod_hoy_reg, prod_hoy_porciones = r[0], float(r[1])

    cur.execute("SELECT COALESCE(SUM(monto),0) FROM gastos WHERE DATE_TRUNC('month',fecha)=DATE_TRUNC('month',CURRENT_DATE)")
    gastos_mes = float(cur.fetchone()[0])

    cur.execute("SELECT COALESCE(SUM(monto),0) FROM gastos WHERE tipo='caja_menor' AND fecha=CURRENT_DATE")
    caja_hoy = float(cur.fetchone()[0])

    cur.execute("""
        SELECT COUNT(*) FROM registros_sanitarios
        WHERE proxima_revision <= CURRENT_DATE + INTERVAL '7 days'
          AND proxima_revision >= CURRENT_DATE
    """)
    alertas_sanitarias = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM remisiones WHERE estado='pendiente'")
    remisiones_pendientes = cur.fetchone()[0]

    # Stock bajo / critico / agotado count
    cur.execute("""
        SELECT COUNT(*) FROM inventario inv
        JOIN materias_primas mp ON mp.id = inv.mp_id AND mp.activo = TRUE
        WHERE inv.stock_minimo > 0
          AND inv.cantidad_actual <= inv.stock_minimo
    """)
    stock_bajo = cur.fetchone()[0]

    # Items con stock bajo (para mini-tabla dashboard)
    cur.execute("""
        SELECT mp.nombre, mp.unidad, inv.cantidad_actual, inv.stock_minimo,
               lc.proveedor AS ultimo_proveedor
        FROM inventario inv
        JOIN materias_primas mp ON mp.id = inv.mp_id AND mp.activo = TRUE
        LEFT JOIN LATERAL (
            SELECT proveedor FROM compras_mp WHERE mp_id = mp.id ORDER BY fecha DESC, id DESC LIMIT 1
        ) lc ON TRUE
        WHERE inv.stock_minimo > 0 AND inv.cantidad_actual <= inv.stock_minimo
        ORDER BY (inv.cantidad_actual / NULLIF(inv.stock_minimo, 0)) ASC
        LIMIT 8
    """)
    items_stock_bajo = [{"nombre": r[0], "unidad": r[1],
                         "cantidad_actual": float(r[2]), "stock_minimo": float(r[3]),
                         "proveedor": r[4]} for r in cur.fetchall()]

    cur.execute("""
        SELECT p.fecha, r.nombre, p.porciones_planeadas, p.porciones, p.operario
        FROM produccion p LEFT JOIN recetas r ON r.id=p.receta_id
        ORDER BY p.fecha DESC, p.id DESC LIMIT 5
    """)
    ultima_produccion = [{"fecha": str(r[0]), "receta": r[1],
                          "porciones_planeadas": float(r[2]) if r[2] else 0,
                          "porciones": float(r[3]), "operario": r[4]}
                         for r in cur.fetchall()]

    cur.close(); conn.close()
    return {
        "total_mp": total_mp,
        "total_recetas": total_recetas,
        "produccion_hoy": {"registros": prod_hoy_reg, "porciones": prod_hoy_porciones},
        "gastos_mes": gastos_mes,
        "caja_hoy": caja_hoy,
        "alertas_sanitarias": alertas_sanitarias,
        "remisiones_pendientes": remisiones_pendientes,
        "stock_bajo": stock_bajo,
        "items_stock_bajo": items_stock_bajo,
        "ultima_produccion": ultima_produccion
    }

@router.get("/dashboard/charts")
def dashboard_charts():
    """Data for dashboard charts: production trend, expense breakdown, monthly costs, recipe margins."""
    conn = get_conn()
    cur = conn.cursor()

    # Production daily last 30 days
    cur.execute("""
        SELECT fecha, COALESCE(SUM(porciones),0)
        FROM produccion
        WHERE fecha >= CURRENT_DATE - INTERVAL '30 days'
        GROUP BY fecha ORDER BY fecha
    """)
    produccion_30d = [{"fecha": str(r[0]), "porciones": float(r[1])} for r in cur.fetchall()]

    # Expenses by category current month
    cur.execute("""
        SELECT categoria, COALESCE(SUM(monto),0) as total
        FROM gastos
        WHERE DATE_TRUNC('month',fecha)=DATE_TRUNC('month',CURRENT_DATE)
        GROUP BY categoria ORDER BY total DESC
    """)
    gastos_categorias = [{"categoria": r[0], "total": float(r[1])} for r in cur.fetchall()]

    # Monthly expenses last 6 months
    cur.execute("""
        SELECT DATE_TRUNC('month', fecha) as mes, COALESCE(SUM(monto),0) as total
        FROM gastos
        WHERE fecha >= CURRENT_DATE - INTERVAL '6 months'
        GROUP BY mes ORDER BY mes
    """)
    gastos_mensuales = [{"mes": r[0].strftime('%Y-%m'), "total": float(r[1])} for r in cur.fetchall()]

    # Top 5 recipes by margin
    cur.execute("""
        SELECT r.nombre, r.precio_venta,
               COALESCE(SUM(ri.cantidad * COALESCE(lc.precio_unit, 0)), 0) / NULLIF(r.porciones, 0) as costo_porcion
        FROM recetas r
        LEFT JOIN receta_ingredientes ri ON ri.receta_id = r.id
        LEFT JOIN LATERAL (
            SELECT precio_unit FROM compras_mp WHERE mp_id = ri.mp_id ORDER BY fecha DESC, id DESC LIMIT 1
        ) lc ON TRUE
        WHERE r.activo = TRUE AND r.precio_venta > 0
        GROUP BY r.id, r.nombre, r.precio_venta, r.porciones
        ORDER BY (r.precio_venta - COALESCE(SUM(ri.cantidad * COALESCE(lc.precio_unit, 0)), 0) / NULLIF(r.porciones, 0)) DESC
        LIMIT 5
    """)
    top_recetas = []
    for r in cur.fetchall():
        costo = float(r[2]) if r[2] else 0
        venta = float(r[1])
        margen_pct = ((venta - costo) / venta * 100) if venta > 0 else 0
        top_recetas.append({"nombre": r[0], "precio_venta": venta, "costo_porcion": costo, "margen_pct": round(margen_pct, 1)})

    # KPI comparisons (current month vs previous month)
    cur.execute("""
        SELECT
            COALESCE(SUM(CASE WHEN DATE_TRUNC('month',fecha)=DATE_TRUNC('month',CURRENT_DATE) THEN porciones END),0) as prod_mes,
            COALESCE(SUM(CASE WHEN DATE_TRUNC('month',fecha)=DATE_TRUNC('month',CURRENT_DATE - INTERVAL '1 month') THEN porciones END),0) as prod_mes_ant
        FROM produccion
        WHERE fecha >= DATE_TRUNC('month', CURRENT_DATE - INTERVAL '1 month')
    """)
    r = cur.fetchone()
    prod_mes = float(r[0]); prod_mes_ant = float(r[1])

    cur.execute("""
        SELECT
            COALESCE(SUM(CASE WHEN DATE_TRUNC('month',fecha)=DATE_TRUNC('month',CURRENT_DATE) THEN monto END),0) as gastos_mes,
            COALESCE(SUM(CASE WHEN DATE_TRUNC('month',fecha)=DATE_TRUNC('month',CURRENT_DATE - INTERVAL '1 month') THEN monto END),0) as gastos_mes_ant
        FROM gastos
        WHERE fecha >= DATE_TRUNC('month', CURRENT_DATE - INTERVAL '1 month')
    """)
    r = cur.fetchone()
    gastos_mes = float(r[0]); gastos_mes_ant = float(r[1])

    # Production last 7 days for sparkline
    cur.execute("""
        SELECT d::date, COALESCE(SUM(p.porciones),0)
        FROM generate_series(CURRENT_DATE - INTERVAL '6 days', CURRENT_DATE, '1 day') d
        LEFT JOIN produccion p ON p.fecha = d::date
        GROUP BY d::date ORDER BY d::date
    """)
    sparkline_prod = [float(r[1]) for r in cur.fetchall()]

    # Expenses last 7 days for sparkline
    cur.execute("""
        SELECT d::date, COALESCE(SUM(g.monto),0)
        FROM generate_series(CURRENT_DATE - INTERVAL '6 days', CURRENT_DATE, '1 day') d
        LEFT JOIN gastos g ON g.fecha = d::date
        GROUP BY d::date ORDER BY d::date
    """)
    sparkline_gastos = [float(r[1]) for r in cur.fetchall()]

    cur.close(); conn.close()
    return {
        "produccion_30d": produccion_30d,
        "gastos_categorias": gastos_categorias,
        "gastos_mensuales": gastos_mensuales,
        "top_recetas": top_recetas,
        "comparaciones": {
            "prod_mes": prod_mes, "prod_mes_ant": prod_mes_ant,
            "gastos_mes": gastos_mes, "gastos_mes_ant": gastos_mes_ant
        },
        "sparklines": {
            "produccion": sparkline_prod,
            "gastos": sparkline_gastos
        }
    }

# ═══════════════════════════════════════════════════════════════
# INVENTARIO
# ═══════════════════════════════════════════════════════════════

class InventarioAjusteIn(BaseModel):
    mp_id: int
    stock_minimo: float = 0

@router.get("/inventario")
def listar_inventario():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT mp.id, mp.codigo, mp.nombre, mp.unidad, mp.categoria,
               COALESCE(inv.cantidad_actual, 0) as cantidad_actual,
               COALESCE(inv.stock_minimo, 0) as stock_minimo,
               inv.updated_at
        FROM materias_primas mp
        LEFT JOIN inventario inv ON inv.mp_id = mp.id
        WHERE mp.activo = TRUE
        ORDER BY mp.categoria, mp.nombre
    """)
    rows = cur.fetchall()
    cur.close(); conn.close()
    result = []
    for r in rows:
        cant = float(r[5])
        minimo = float(r[6])
        if minimo > 0:
            status = 'ok' if cant > minimo * 1.5 else ('bajo' if cant > minimo else ('critico' if cant > 0 else 'agotado'))
        else:
            status = 'ok' if cant > 0 else 'sin_stock'
        result.append({
            "id": r[0], "codigo": r[1], "nombre": r[2], "unidad": r[3], "categoria": r[4],
            "cantidad_actual": cant, "stock_minimo": minimo,
            "updated_at": str(r[7]) if r[7] else None,
            "status": status
        })
    return result

@router.post("/inventario/ajuste")
def ajustar_inventario(data: InventarioAjusteIn):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO inventario (mp_id, cantidad_actual, stock_minimo, updated_at)
        VALUES (%s, 0, %s, NOW())
        ON CONFLICT (mp_id) DO UPDATE SET stock_minimo=%s, updated_at=NOW()
    """, (data.mp_id, data.stock_minimo, data.stock_minimo))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

@router.post("/inventario/entrada")
def inventario_entrada(mp_id: int, cantidad: float):
    """Adds stock (called after purchase registration)"""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO inventario (mp_id, cantidad_actual, updated_at)
        VALUES (%s, %s, NOW())
        ON CONFLICT (mp_id) DO UPDATE SET cantidad_actual = inventario.cantidad_actual + %s, updated_at = NOW()
    """, (mp_id, cantidad, cantidad))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

@router.put("/inventario/{mp_id}/stock-minimo")
def update_stock_minimo(mp_id: int, data: InventarioAjusteIn):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO inventario (mp_id, cantidad_actual, stock_minimo, updated_at)
        VALUES (%s, 0, %s, NOW())
        ON CONFLICT (mp_id) DO UPDATE SET stock_minimo=%s, updated_at=NOW()
    """, (mp_id, data.stock_minimo, data.stock_minimo))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

# ═══════════════════════════════════════════════════════════════
# RECETA — Escalar (calculadora de ingredientes)
# ═══════════════════════════════════════════════════════════════

@router.get("/recetas/{receta_id}/escalar")
def escalar_receta(receta_id: int, porciones: float = 1):
    """Calcula ingredientes necesarios para N porciones."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT porciones FROM recetas WHERE id=%s", (receta_id,))
    r = cur.fetchone()
    if not r:
        raise HTTPException(404, "Receta no encontrada")
    porciones_base = float(r[0]) if r[0] else 1
    factor = porciones / porciones_base if porciones_base > 0 else 1

    cur.execute("""
        SELECT mp.nombre, mp.unidad, ri.cantidad,
               COALESCE(ult.precio_unit, 0) AS precio_unit,
               COALESCE(inv.cantidad_actual, 0) AS stock
        FROM receta_ingredientes ri
        JOIN materias_primas mp ON mp.id = ri.mp_id
        LEFT JOIN LATERAL (
            SELECT precio_unit FROM compras_mp WHERE mp_id = ri.mp_id ORDER BY fecha DESC, id DESC LIMIT 1
        ) ult ON TRUE
        LEFT JOIN inventario inv ON inv.mp_id = ri.mp_id
        WHERE ri.receta_id = %s
    """, (receta_id,))
    ingredientes = []
    for i in cur.fetchall():
        cant_necesaria = float(i[2]) * factor
        precio = float(i[3])
        stock = float(i[4])
        ingredientes.append({
            "nombre": i[0], "unidad": i[1],
            "cantidad_base": float(i[2]),
            "cantidad_necesaria": round(cant_necesaria, 4),
            "costo": round(cant_necesaria * precio, 2),
            "stock_disponible": stock,
            "alcanza": stock >= cant_necesaria
        })
    cur.close(); conn.close()
    costo_total = sum(i["costo"] for i in ingredientes)
    return {
        "porciones_solicitadas": porciones,
        "factor": round(factor, 4),
        "ingredientes": ingredientes,
        "costo_total": costo_total,
        "costo_porcion": round(costo_total / porciones, 2) if porciones > 0 else 0
    }

# ═══════════════════════════════════════════════════════════════
# TRAZABILIDAD — Rastreo de lotes
# ═══════════════════════════════════════════════════════════════

@router.get("/trazabilidad/produccion/{lote}")
def trazar_produccion(lote: str):
    """Traza un lote de producción hasta sus materias primas y proveedores."""
    conn = get_conn()
    cur = conn.cursor()
    # Buscar producción por lote
    cur.execute("""
        SELECT p.id, p.fecha, p.porciones, p.operario, p.notas,
               r.nombre AS receta, r.id AS receta_id
        FROM produccion p
        LEFT JOIN recetas r ON r.id = p.receta_id
        WHERE p.lote = %s
    """, (lote,))
    prod = cur.fetchone()
    if not prod:
        raise HTTPException(404, "Lote de producción no encontrado")

    # Ingredientes de la receta con sus últimas compras
    cur.execute("""
        SELECT mp.id, mp.nombre, mp.unidad, ri.cantidad,
               c.fecha AS compra_fecha, c.proveedor, c.precio_unit
        FROM receta_ingredientes ri
        JOIN materias_primas mp ON mp.id = ri.mp_id
        LEFT JOIN LATERAL (
            SELECT fecha, proveedor, precio_unit FROM compras_mp
            WHERE mp_id = ri.mp_id AND fecha <= %s
            ORDER BY fecha DESC, id DESC LIMIT 1
        ) c ON TRUE
        WHERE ri.receta_id = %s
    """, (prod[1], prod[6]))
    ingredientes_base = cur.fetchall()

    # Try to get lote info from remision_items (column may not exist in older DBs)
    lote_info = {}
    try:
        mp_ids = [i[0] for i in ingredientes_base]
        if mp_ids:
            cur.execute("""
                SELECT ri2.mp_id, ri2.lote, rem.numero
                FROM remision_items ri2
                LEFT JOIN remisiones rem ON rem.id = ri2.remision_id
                WHERE ri2.mp_id = ANY(%s)
                ORDER BY ri2.id DESC
            """, (mp_ids,))
            for row in cur.fetchall():
                if row[0] not in lote_info:
                    lote_info[row[0]] = {"lote_mp": row[1], "remision": row[2]}
    except Exception:
        conn.rollback()

    ingredientes = []
    for i in ingredientes_base:
        li = lote_info.get(i[0], {})
        ingredientes.append({
            "mp_nombre": i[1], "unidad": i[2], "cantidad": float(i[3]),
            "compra_fecha": str(i[4]) if i[4] else None,
            "proveedor": i[5], "precio_unit": float(i[6]) if i[6] else 0,
            "lote_mp": li.get("lote_mp"), "remision": li.get("remision"),
            "rem_proveedor": None
        })
    cur.close(); conn.close()
    return {
        "lote": lote,
        "fecha": str(prod[1]),
        "receta": prod[5],
        "porciones": float(prod[2]),
        "operario": prod[3],
        "notas": prod[4],
        "ingredientes": ingredientes
    }

@router.get("/trazabilidad/buscar")
def buscar_lotes(q: str = ""):
    """Busca lotes de producción por texto."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT p.lote, p.fecha, r.nombre, p.porciones, p.operario
        FROM produccion p LEFT JOIN recetas r ON r.id = p.receta_id
        WHERE p.lote ILIKE %s OR r.nombre ILIKE %s
        ORDER BY p.fecha DESC LIMIT 20
    """, (f"%{q}%", f"%{q}%"))
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"lote": r[0], "fecha": str(r[1]), "receta": r[2],
             "porciones": float(r[3]), "operario": r[4]} for r in rows]

# ═══════════════════════════════════════════════════════════════
# ÓRDENES DE COMPRA
# ═══════════════════════════════════════════════════════════════

class OrdenItemIn(BaseModel):
    mp_id: int
    mp_nombre: str
    cantidad: float
    precio_est: float = 0
    notas: Optional[str] = None

class OrdenCompraIn(BaseModel):
    fecha: date
    proveedor: str
    notas: Optional[str] = None
    creado_por: Optional[str] = None
    items: List[OrdenItemIn]

def _gen_numero_orden(new_id: int) -> str:
    return f"OC-{datetime.now().strftime('%Y%m%d')}-{new_id:06d}"

@router.get("/ordenes-compra")
def listar_ordenes(estado: Optional[str] = None, limit: int = 100):
    conn = get_conn()
    cur = conn.cursor()
    if estado:
        cur.execute("""
            SELECT id, numero, fecha, proveedor, estado, notas, creado_por, creado_en
            FROM ordenes_compra WHERE estado=%s ORDER BY creado_en DESC LIMIT %s
        """, (estado, limit))
    else:
        cur.execute("""
            SELECT id, numero, fecha, proveedor, estado, notas, creado_por, creado_en
            FROM ordenes_compra ORDER BY creado_en DESC LIMIT %s
        """, (limit,))
    rows = cur.fetchall()
    if not rows:
        cur.close(); conn.close()
        return []
    orden_ids = [r[0] for r in rows]
    cur.execute("""
        SELECT orden_id, mp_nombre, cantidad, precio_est
        FROM orden_items WHERE orden_id = ANY(%s) ORDER BY orden_id, id
    """, (orden_ids,))
    items_by_orden = {}
    for i in cur.fetchall():
        items_by_orden.setdefault(i[0], []).append({
            "mp_nombre": i[1], "cantidad": float(i[2]), "precio_est": float(i[3]) if i[3] else 0
        })
    result = []
    for r in rows:
        items = items_by_orden.get(r[0], [])
        total = sum(i["cantidad"] * i["precio_est"] for i in items)
        result.append({"id": r[0], "numero": r[1], "fecha": str(r[2]), "proveedor": r[3],
                       "estado": r[4], "notas": r[5], "creado_por": r[6],
                       "creado_en": str(r[7]), "items": items, "total": total})
    cur.close(); conn.close()
    return result

@router.post("/ordenes-compra")
def crear_orden(data: OrdenCompraIn):
    if not data.items:
        raise HTTPException(400, "Debe agregar al menos un ítem")
    conn = get_conn()
    cur = conn.cursor()
    temp = f"TEMP-{uuid.uuid4()}"
    cur.execute("""
        INSERT INTO ordenes_compra (numero, fecha, proveedor, notas, creado_por)
        VALUES (%s,%s,%s,%s,%s) RETURNING id
    """, (temp, data.fecha, data.proveedor, data.notas, data.creado_por))
    orden_id = cur.fetchone()[0]
    numero = _gen_numero_orden(orden_id)
    cur.execute("UPDATE ordenes_compra SET numero=%s WHERE id=%s", (numero, orden_id))
    for item in data.items:
        cur.execute("""
            INSERT INTO orden_items (orden_id, mp_id, mp_nombre, cantidad, precio_est, notas)
            VALUES (%s,%s,%s,%s,%s,%s)
        """, (orden_id, item.mp_id, item.mp_nombre, item.cantidad, item.precio_est, item.notas))
    conn.commit(); cur.close(); conn.close()
    return {"id": orden_id, "numero": numero, "ok": True}

@router.put("/ordenes-compra/{orden_id}/estado")
def cambiar_estado_orden(orden_id: int, estado: str):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("UPDATE ordenes_compra SET estado=%s, actualizado_en=NOW() WHERE id=%s", (estado, orden_id))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

@router.delete("/ordenes-compra/{orden_id}")
def eliminar_orden(orden_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM ordenes_compra WHERE id=%s", (orden_id,))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

@router.get("/ordenes-compra/generar-desde-inventario")
def generar_orden_desde_inventario():
    """Sugiere items para orden de compra basado en stock bajo."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT mp.id, mp.nombre, mp.unidad, inv.cantidad_actual, inv.stock_minimo,
               COALESCE(lc.precio_unit, 0) AS precio_est,
               COALESCE(lc.proveedor, '') AS proveedor
        FROM inventario inv
        JOIN materias_primas mp ON mp.id = inv.mp_id AND mp.activo = TRUE
        LEFT JOIN LATERAL (
            SELECT precio_unit, proveedor FROM compras_mp WHERE mp_id = mp.id ORDER BY fecha DESC, id DESC LIMIT 1
        ) lc ON TRUE
        WHERE inv.stock_minimo > 0 AND inv.cantidad_actual <= inv.stock_minimo
        ORDER BY mp.nombre
    """)
    rows = cur.fetchall()
    cur.close(); conn.close()
    # Agrupar por proveedor
    por_proveedor = {}
    for r in rows:
        prov = r[6] or 'Sin proveedor'
        faltante = float(r[4]) * 2 - float(r[3])  # Pedir hasta 2x el mínimo
        if faltante < 0: faltante = float(r[4])
        item = {"mp_id": r[0], "mp_nombre": r[1], "unidad": r[2],
                "stock_actual": float(r[3]), "stock_minimo": float(r[4]),
                "cantidad_sugerida": round(faltante, 2),
                "precio_est": float(r[5])}
        if prov not in por_proveedor:
            por_proveedor[prov] = []
        por_proveedor[prov].append(item)
    return por_proveedor

# ═══════════════════════════════════════════════════════════════
# INVIMA — Programas Sanitarios
# ═══════════════════════════════════════════════════════════════

class InvimaProgramaIn(BaseModel):
    nombre: str
    codigo: Optional[str] = None
    descripcion: Optional[str] = None
    responsable: Optional[str] = None
    frecuencia: str = "Mensual"

class InvimaRegistroIn(BaseModel):
    programa_id: int
    fecha: date
    descripcion: str
    resultado: str = "Conforme"
    responsable: Optional[str] = None
    observaciones: Optional[str] = None
    proxima_revision: Optional[date] = None

@router.get("/invima/programas")
def listar_invima_programas():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT p.id, p.codigo, p.nombre, p.descripcion, p.responsable, p.frecuencia, p.activo,
               (SELECT COUNT(*) FROM invima_registros WHERE programa_id=p.id) AS total_registros,
               (SELECT MAX(fecha) FROM invima_registros WHERE programa_id=p.id) AS ultimo_registro,
               (SELECT MIN(proxima_revision) FROM invima_registros
                WHERE programa_id=p.id AND proxima_revision >= CURRENT_DATE) AS prox_revision
        FROM invima_programas p WHERE p.activo=TRUE
        ORDER BY p.codigo, p.nombre
    """)
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id": r[0], "codigo": r[1], "nombre": r[2], "descripcion": r[3],
             "responsable": r[4], "frecuencia": r[5], "activo": r[6],
             "total_registros": r[7],
             "ultimo_registro": str(r[8]) if r[8] else None,
             "prox_revision": str(r[9]) if r[9] else None} for r in rows]

@router.post("/invima/programas")
def crear_invima_programa(data: InvimaProgramaIn):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO invima_programas (codigo, nombre, descripcion, responsable, frecuencia)
        VALUES (%s,%s,%s,%s,%s) RETURNING id
    """, (data.codigo, data.nombre, data.descripcion, data.responsable, data.frecuencia))
    new_id = cur.fetchone()[0]
    conn.commit(); cur.close(); conn.close()
    return {"id": new_id, "ok": True}

@router.put("/invima/programas/{prog_id}")
def editar_invima_programa(prog_id: int, data: InvimaProgramaIn):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        UPDATE invima_programas SET codigo=%s, nombre=%s, descripcion=%s,
               responsable=%s, frecuencia=%s WHERE id=%s
    """, (data.codigo, data.nombre, data.descripcion, data.responsable, data.frecuencia, prog_id))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

@router.delete("/invima/programas/{prog_id}")
def eliminar_invima_programa(prog_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("UPDATE invima_programas SET activo=FALSE WHERE id=%s", (prog_id,))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

@router.get("/invima/registros")
def listar_invima_registros(programa_id: Optional[int] = None, limit: int = 100):
    conn = get_conn()
    cur = conn.cursor()
    if programa_id:
        cur.execute("""
            SELECT r.id, r.fecha, p.nombre AS programa, p.codigo, r.descripcion,
                   r.resultado, r.responsable, r.observaciones, r.proxima_revision
            FROM invima_registros r JOIN invima_programas p ON p.id=r.programa_id
            WHERE r.programa_id=%s ORDER BY r.fecha DESC LIMIT %s
        """, (programa_id, limit))
    else:
        cur.execute("""
            SELECT r.id, r.fecha, p.nombre AS programa, p.codigo, r.descripcion,
                   r.resultado, r.responsable, r.observaciones, r.proxima_revision
            FROM invima_registros r JOIN invima_programas p ON p.id=r.programa_id
            ORDER BY r.fecha DESC LIMIT %s
        """, (limit,))
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id": r[0], "fecha": str(r[1]), "programa": r[2], "codigo": r[3],
             "descripcion": r[4], "resultado": r[5], "responsable": r[6],
             "observaciones": r[7],
             "proxima_revision": str(r[8]) if r[8] else None} for r in rows]

@router.post("/invima/registros")
def crear_invima_registro(data: InvimaRegistroIn):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO invima_registros (programa_id, fecha, descripcion, resultado, responsable, observaciones, proxima_revision)
        VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id
    """, (data.programa_id, data.fecha, data.descripcion, data.resultado,
          data.responsable, data.observaciones, data.proxima_revision))
    new_id = cur.fetchone()[0]
    conn.commit(); cur.close(); conn.close()
    return {"id": new_id, "ok": True}

@router.delete("/invima/registros/{reg_id}")
def eliminar_invima_registro(reg_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM invima_registros WHERE id=%s", (reg_id,))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

# ═══════════════════════════════════════════════════════════════
# PROYECCIONES DE DEMANDA
# ═══════════════════════════════════════════════════════════════

@router.get("/proyecciones")
def proyecciones(dias: int = 30):
    """Forecast simple basado en producción histórica por receta."""
    conn = get_conn()
    cur = conn.cursor()
    # Producción por receta últimos 90 días (para calcular promedio)
    cur.execute("""
        SELECT r.id, r.nombre, r.categoria,
               COUNT(p.id) AS veces_producido,
               COALESCE(SUM(p.porciones), 0) AS total_porciones,
               COALESCE(AVG(p.porciones), 0) AS promedio_porciones,
               MAX(p.fecha) AS ultima_produccion,
               COALESCE(SUM(ri.cantidad * COALESCE(lc.precio_unit, 0)), 0) / NULLIF(r.porciones, 0) AS costo_porcion
        FROM recetas r
        LEFT JOIN produccion p ON p.receta_id = r.id AND p.fecha >= CURRENT_DATE - INTERVAL '90 days'
        LEFT JOIN receta_ingredientes ri ON ri.receta_id = r.id
        LEFT JOIN LATERAL (
            SELECT precio_unit FROM compras_mp WHERE mp_id = ri.mp_id ORDER BY fecha DESC, id DESC LIMIT 1
        ) lc ON TRUE
        WHERE r.activo = TRUE
        GROUP BY r.id, r.nombre, r.categoria, r.porciones
        HAVING COUNT(p.id) > 0
        ORDER BY total_porciones DESC
    """)
    rows = cur.fetchall()

    # Producción por semana últimas 12 semanas para tendencia
    cur.execute("""
        SELECT DATE_TRUNC('week', fecha) AS semana, COALESCE(SUM(porciones), 0)
        FROM produccion
        WHERE fecha >= CURRENT_DATE - INTERVAL '12 weeks'
        GROUP BY semana ORDER BY semana
    """)
    tendencia_semanal = [{"semana": r[0].strftime('%Y-%m-%d'), "porciones": float(r[1])} for r in cur.fetchall()]

    cur.close(); conn.close()

    recetas_forecast = []
    for r in rows:
        veces = r[3]
        total = float(r[4])
        promedio = float(r[5])
        # Proyección simple: (promedio diario) * días
        promedio_diario = total / 90.0
        proyeccion = round(promedio_diario * dias, 1)
        recetas_forecast.append({
            "id": r[0], "nombre": r[1], "categoria": r[2],
            "veces_90d": veces, "total_90d": total,
            "promedio_produccion": round(promedio, 1),
            "ultima_produccion": str(r[6]) if r[6] else None,
            "costo_porcion": round(float(r[7]), 2) if r[7] else 0,
            "proyeccion_porciones": proyeccion,
            "proyeccion_costo": round(proyeccion * (float(r[7]) if r[7] else 0), 2)
        })

    return {
        "dias_proyeccion": dias,
        "recetas": recetas_forecast,
        "tendencia_semanal": tendencia_semanal
    }

# ═══════════════════════════════════════════════════════════════
# SIIGO — Productos y Ventas
# ═══════════════════════════════════════════════════════════════

@router.get("/siigo/productos")
def siigo_productos(tipo: str = "terminado"):
    """tipo: terminado | mp | todos"""
    try:
        from siigo import fetch_products
        return fetch_products(tipo=tipo)
    except Exception as e:
        raise HTTPException(500, f"Error Siigo: {str(e)}")

@router.get("/siigo/ventas-semana")
def siigo_ventas_semana(semanas: int = 8):
    try:
        from siigo import sales_by_product_weekly
        return sales_by_product_weekly(semanas)
    except Exception as e:
        raise HTTPException(500, f"Error Siigo: {str(e)}")

# ═══════════════════════════════════════════════════════════════
# FINANZAS — Contabilidad desde Siigo
# ═══════════════════════════════════════════════════════════════

@router.post("/finanzas/sync")
def sync_contabilidad():
    try:
        from siigo_contabilidad import sync_journals
        result = sync_journals()
        return {"ok": True, **result}
    except Exception as e:
        raise HTTPException(500, f"Error sync: {str(e)}")

@router.get("/finanzas/balance-prueba")
def balance_prueba(anio: int = 2026, mes: int = 3):
    try:
        from siigo_contabilidad import get_balance_prueba
        return get_balance_prueba(anio, mes)
    except Exception as e:
        raise HTTPException(500, f"Error: {str(e)}")

@router.get("/finanzas/estado-resultados")
def estado_resultados(anio: int = 2026, mes_inicio: int = 1, mes_fin: int = 3):
    try:
        from siigo_contabilidad import get_estado_resultados
        return get_estado_resultados(anio, mes_inicio, mes_fin)
    except Exception as e:
        raise HTTPException(500, f"Error: {str(e)}")

@router.get("/finanzas/indicadores")
def indicadores_financieros(anio: int = 2026, mes: int = 3):
    try:
        from siigo_contabilidad import get_indicadores
        return get_indicadores(anio, mes)
    except Exception as e:
        raise HTTPException(500, f"Error: {str(e)}")

@router.get("/finanzas/tendencia")
def tendencia_financiera(anio: int = 2026):
    try:
        from siigo_contabilidad import get_tendencia_mensual_from_invoices
        return get_tendencia_mensual_from_invoices(anio)
    except Exception as e:
        raise HTTPException(500, f"Error: {str(e)}")

@router.get("/finanzas/diagnostico")
def diagnostico_financiero():
    """Debug: check what data sources are in DB and their 5xxx totals."""
    import json
    from collections import defaultdict
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM siigo_journals WHERE id LIKE 'VCH_%'")
    vch = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM siigo_journals WHERE id LIKE 'JRN_%'")
    jrn = cur.fetchone()[0]

    # Gastos 5xxx from vouchers only
    gastos_vch = defaultdict(float)
    gastos_jrn = defaultdict(float)
    cur.execute("SELECT id, items FROM siigo_journals")
    for row_id, items_json in cur.fetchall():
        items = json.loads(items_json) if isinstance(items_json, str) else items_json
        for item in items:
            if not isinstance(item, dict) or 'account' not in item:
                continue
            code = item['account']['code']
            if code.startswith('5') and item['account']['movement'] == 'Debit':
                if row_id.startswith('VCH_'):
                    gastos_vch[code[:4]] += item['value']
                elif row_id.startswith('JRN_'):
                    gastos_jrn[code[:4]] += item['value']

    cur.close(); conn.close()
    return {
        "vouchers_in_db": vch, "journals_in_db": jrn,
        "gastos_5xxx_vouchers": round(sum(gastos_vch.values()), 2),
        "gastos_5xxx_journals": round(sum(gastos_jrn.values()), 2),
        "top_vch_5xxx": {k: round(v,2) for k,v in sorted(gastos_vch.items(), key=lambda x:-x[1])[:10]},
        "top_jrn_5xxx": {k: round(v,2) for k,v in sorted(gastos_jrn.items(), key=lambda x:-x[1])[:10]},
    }

@router.get("/finanzas/balance-general")
def balance_general(anio: int = 2026, mes: int = 3):
    try:
        from siigo_contabilidad import get_balance_general
        return get_balance_general(anio, mes)
    except Exception as e:
        raise HTTPException(500, f"Error: {str(e)}")

@router.get("/finanzas/comparativo")
def comparativo_pl(anio1: int = 2026, m1i: int = 1, m1f: int = 3,
                   anio2: int = 2025, m2i: int = 1, m2f: int = 3):
    try:
        from siigo_contabilidad import get_comparativo
        return get_comparativo(anio1, m1i, m1f, anio2, m2i, m2f)
    except Exception as e:
        raise HTTPException(500, f"Error: {str(e)}")

@router.get("/finanzas/export-excel")
def export_excel_eeff(anio: int = 2026, mes_inicio: int = 1, mes_fin: int = 12):
    from fastapi.responses import Response
    try:
        from siigo_contabilidad import export_eeff_excel
        content = export_eeff_excel(anio, mes_inicio, mes_fin)
        meses = ['','Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
        filename = f"EEFF_Daily_{meses[mes_inicio]}-{meses[mes_fin]}_{anio}.xlsx"
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except Exception as e:
        raise HTTPException(500, f"Error: {str(e)}")

@router.get("/finanzas/presupuesto-vs-real")
def presupuesto_vs_real(anio: int = 2026, mes: int = 3):
    try:
        from siigo_contabilidad import get_presupuesto_vs_real
        return get_presupuesto_vs_real(anio, mes)
    except Exception as e:
        raise HTTPException(500, f"Error: {str(e)}")

@router.get("/finanzas/sync-log")
def sync_log():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id, tipo, fecha, registros, detalle FROM sync_log ORDER BY fecha DESC LIMIT 20")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id": r[0], "tipo": r[1], "fecha": str(r[2]), "registros": r[3], "detalle": r[4]} for r in rows]

class PresupuestoIn(BaseModel):
    cuenta: str
    cuenta_nombre: Optional[str] = None
    anio: int
    mes: int
    monto: float
    notas: Optional[str] = None

@router.post("/finanzas/presupuestos")
def guardar_presupuesto(data: PresupuestoIn):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO presupuestos (cuenta, cuenta_nombre, anio, mes, monto, notas)
        VALUES (%s, %s, %s, %s, %s, %s)
        ON CONFLICT (cuenta, anio, mes) DO UPDATE SET monto=%s, notas=%s, cuenta_nombre=%s
    """, (data.cuenta, data.cuenta_nombre, data.anio, data.mes, data.monto, data.notas,
          data.monto, data.notas, data.cuenta_nombre))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

@router.get("/finanzas/presupuestos")
def listar_presupuestos(anio: int = 2026):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT cuenta, cuenta_nombre, anio, mes, monto, notas
        FROM presupuestos WHERE anio=%s ORDER BY cuenta, mes
    """, (anio,))
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"cuenta": r[0], "nombre": r[1], "anio": r[2], "mes": r[3],
             "monto": float(r[4]), "notas": r[5]} for r in rows]

# ═══════════════════════════════════════════════════════════════
# PRODUCTO ↔ RECETA (ligado Siigo → Daily)
# ═══════════════════════════════════════════════════════════════

class ProductoRecetaIn(BaseModel):
    siigo_code: str
    siigo_name: str
    siigo_group: Optional[str] = None
    receta_id: Optional[int] = None

@router.get("/producto-receta")
def listar_producto_receta():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT pr.id, pr.siigo_code, pr.siigo_name, pr.siigo_group,
               CASE WHEN r.activo = TRUE THEN pr.receta_id ELSE NULL END AS receta_id,
               CASE WHEN r.activo = TRUE THEN r.nombre ELSE NULL END AS receta_nombre,
               pr.activo, COALESCE(pr.precio_venta, 0)
        FROM producto_receta pr
        LEFT JOIN recetas r ON r.id = pr.receta_id
        WHERE pr.activo = TRUE
        ORDER BY pr.siigo_group, pr.siigo_name
    """)
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id": r[0], "siigo_code": r[1], "siigo_name": r[2], "siigo_group": r[3],
             "receta_id": r[4], "receta_nombre": r[5], "activo": r[6],
             "precio_venta": float(r[7])} for r in rows]

@router.post("/producto-receta")
def crear_producto_receta(data: ProductoRecetaIn):
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO producto_receta (siigo_code, siigo_name, siigo_group, receta_id)
            VALUES (%s,%s,%s,%s)
            ON CONFLICT (siigo_code) WHERE siigo_code IS NOT NULL
            DO UPDATE SET siigo_name=%s, siigo_group=%s, receta_id=%s
            RETURNING id
        """, (data.siigo_code, data.siigo_name, data.siigo_group, data.receta_id,
              data.siigo_name, data.siigo_group, data.receta_id))
        new_id = cur.fetchone()[0]
        conn.commit()
    except Exception as e:
        conn.rollback()
        raise HTTPException(500, str(e))
    finally:
        cur.close(); conn.close()
    return {"id": new_id, "ok": True}

@router.put("/producto-receta/{pr_id}/receta")
def ligar_receta(pr_id: int, receta_id: Optional[int] = None):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("UPDATE producto_receta SET receta_id=%s WHERE id=%s", (receta_id, pr_id))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

@router.post("/producto-receta/sync")
def sync_productos_siigo():
    """Sincroniza productos terminados de Siigo con la tabla producto_receta + actualiza precios en recetas."""
    try:
        from siigo import fetch_products
        productos = fetch_products(tipo="terminado")
    except Exception as e:
        raise HTTPException(500, f"Error Siigo: {str(e)}")
    conn = get_conn()
    cur = conn.cursor()
    synced = 0
    precios_actualizados = 0
    for p in productos:
        precio = p.get("precio_venta", 0)
        cur.execute("""
            INSERT INTO producto_receta (siigo_code, siigo_name, siigo_group, precio_venta)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (siigo_code) WHERE siigo_code IS NOT NULL
            DO UPDATE SET siigo_name=%s, siigo_group=%s, precio_venta=%s
            RETURNING id, receta_id
        """, (p["code"], p["name"], p["group"], precio,
              p["name"], p["group"], precio))
        row = cur.fetchone()
        synced += 1
        # Si tiene receta ligada, actualizar precio_venta de la receta
        if row and row[1] and precio > 0:
            cur.execute("UPDATE recetas SET precio_venta=%s WHERE id=%s", (precio, row[1]))
            precios_actualizados += 1
    conn.commit(); cur.close(); conn.close()
    return {"synced": synced, "precios_actualizados": precios_actualizados, "ok": True}

# ═══════════════════════════════════════════════════════════════
# REGLAS DE PRODUCCIÓN (condicionales)
# ═══════════════════════════════════════════════════════════════

class ReglaIn(BaseModel):
    tipo: str          # dias_recepcion, vida_util, no_fin_semana, lead_time, stock_seguridad, produccion_dia, capacidad_max
    entidad: Optional[str] = None   # nombre de MP, proveedor, receta, o 'planta'
    entidad_id: Optional[int] = None
    parametro: str
    valor: str
    descripcion: Optional[str] = None

@router.get("/reglas")
def listar_reglas():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, tipo, entidad, entidad_id, parametro, valor, descripcion, activo
        FROM reglas_produccion WHERE activo=TRUE ORDER BY tipo, entidad
    """)
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id": r[0], "tipo": r[1], "entidad": r[2], "entidad_id": r[3],
             "parametro": r[4], "valor": r[5], "descripcion": r[6], "activo": r[7]} for r in rows]

@router.post("/reglas")
def crear_regla(data: ReglaIn):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO reglas_produccion (tipo, entidad, entidad_id, parametro, valor, descripcion)
        VALUES (%s,%s,%s,%s,%s,%s) RETURNING id
    """, (data.tipo, data.entidad, data.entidad_id, data.parametro, data.valor, data.descripcion))
    new_id = cur.fetchone()[0]
    conn.commit(); cur.close(); conn.close()
    return {"id": new_id, "ok": True}

@router.put("/reglas/{regla_id}")
def editar_regla(regla_id: int, data: ReglaIn):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        UPDATE reglas_produccion SET tipo=%s, entidad=%s, entidad_id=%s,
               parametro=%s, valor=%s, descripcion=%s WHERE id=%s
    """, (data.tipo, data.entidad, data.entidad_id, data.parametro, data.valor, data.descripcion, regla_id))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

@router.delete("/reglas/{regla_id}")
def eliminar_regla(regla_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("UPDATE reglas_produccion SET activo=FALSE WHERE id=%s", (regla_id,))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

# ═══════════════════════════════════════════════════════════════
# CONFIGURACIÓN DE PLANTA
# ═══════════════════════════════════════════════════════════════

@router.get("/config-planta")
def get_config_planta():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT parametro, valor, descripcion FROM config_planta ORDER BY parametro")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return {r[0]: {"valor": r[1], "descripcion": r[2]} for r in rows}

class ConfigPlantaIn(BaseModel):
    parametro: str
    valor: str

@router.put("/config-planta")
def update_config_planta(data: ConfigPlantaIn):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO config_planta (parametro, valor) VALUES (%s, %s)
        ON CONFLICT (parametro) DO UPDATE SET valor=%s
    """, (data.parametro, data.valor, data.valor))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

# ═══════════════════════════════════════════════════════════════
# PLAN DE PRODUCCIÓN SEMANAL (motor de planificación v2)
# ═══════════════════════════════════════════════════════════════

import math

@router.get("/plan-produccion")
def plan_produccion(semanas_historico: int = 4):
    """
    Motor MRP:
    1. Ventas Siigo → demanda semanal promedio
    2. Stock mínimo = demanda × factor (default 2)
    3. Necesidad = stock_mínimo - stock_actual
    4. Decisión batch: completo vs parcial basado en vida útil y velocidad de venta
    5. Costo por unidad: MP + (mano_obra + servicios) / batch_size
    6. Cascada a MP → plan de compras con reglas
    """
    # 1. Ventas Siigo
    try:
        from siigo import sales_by_product_weekly
        ventas = sales_by_product_weekly(semanas_historico)
    except Exception as e:
        raise HTTPException(500, f"Error Siigo: {str(e)}")

    conn = get_conn()
    cur = conn.cursor()

    # Config planta
    cur.execute("SELECT parametro, valor FROM config_planta")
    config = {r[0]: r[1] for r in cur.fetchall()}
    horas_dia = float(config.get("horas_productivas_dia", "8"))
    minutos_dia = horas_dia * 60
    factor_stock = float(config.get("factor_stock_minimo", "2"))

    # 2. Ligados producto ↔ receta con config de producción
    cur.execute("""
        SELECT pr.siigo_code, pr.siigo_name, pr.receta_id,
               r.nombre, r.porciones,
               COALESCE(r.batch_maximo, 0), COALESCE(r.tiempo_batch_min, 0),
               COALESCE(r.vida_util_dias, 30), COALESCE(r.costo_mano_obra, 0),
               COALESCE(r.costo_servicios, 0)
        FROM producto_receta pr
        JOIN recetas r ON r.id = pr.receta_id
        WHERE pr.activo = TRUE AND pr.receta_id IS NOT NULL
    """)
    ligados = {}
    for r in cur.fetchall():
        ligados[r[0]] = {
            "siigo_name": r[1], "receta_id": r[2], "receta_nombre": r[3],
            "porciones_base": float(r[4]) if r[4] else 1,
            "batch_maximo": float(r[5]), "tiempo_batch_min": int(r[6]),
            "vida_util_dias": int(r[7]), "costo_mano_obra": float(r[8]),
            "costo_servicios": float(r[9])
        }

    # 3. Reglas
    cur.execute("SELECT tipo, entidad, parametro, valor FROM reglas_produccion WHERE activo=TRUE")
    reglas = {}
    for tipo, entidad, param, valor in cur.fetchall():
        key = f"{tipo}:{entidad or 'global'}"
        reglas[key] = valor

    # 4. Calcular plan por producto
    num_weeks = max(len(ventas.get("weeks", [])), 1)
    plan_productos = []
    total_mp_necesaria = {}
    tiempo_total_min = 0

    for prod_venta in ventas.get("products", []):
        code = prod_venta["code"]
        if code not in ligados:
            continue
        lig = ligados[code]
        avg_qty_semana = prod_venta["total_qty"] / num_weeks
        avg_revenue = prod_venta["total_revenue"] / num_weeks

        # Stock mínimo = venta semanal × factor
        stock_minimo = avg_qty_semana * factor_stock
        # TODO: stock_actual de producto terminado (por ahora = 0)
        stock_actual_pt = 0
        necesidad = max(0, stock_minimo - stock_actual_pt)

        if necesidad <= 0:
            continue

        # Decisión batch completo vs parcial
        batch_max = lig["batch_maximo"]
        vida_util = lig["vida_util_dias"]
        venta_diaria = avg_qty_semana / 7 if avg_qty_semana > 0 else 0.1
        costo_mo = lig["costo_mano_obra"]
        costo_sv = lig["costo_servicios"]
        costo_fijo = costo_mo + costo_sv
        tiempo_batch = lig["tiempo_batch_min"]

        if batch_max > 0:
            batches_necesarios = math.ceil(necesidad / batch_max)
            produccion_batch_completo = batches_necesarios * batch_max
            sobrante = produccion_batch_completo - necesidad
            dias_para_vender_sobrante = sobrante / venta_diaria if venta_diaria > 0 else 999

            if dias_para_vender_sobrante <= vida_util:
                # Batch completo conviene — sobrante se vende a tiempo
                producir = produccion_batch_completo
                batches = batches_necesarios
                decision = "batch_completo"
                razon = f"Sobrante ({round(sobrante)}) se vende en {round(dias_para_vender_sobrante,1)} dias (vida util: {vida_util}d)"
            else:
                # Sobrante se daña — producir solo lo necesario
                producir = round(necesidad)
                batches = necesidad / batch_max
                decision = "parcial"
                razon = f"Batch completo dejaria {round(sobrante)} sobrantes que tardan {round(dias_para_vender_sobrante,1)} dias en venderse (vida util: {vida_util}d)"

            # Costo por unidad
            costo_fijo_und = costo_fijo / batch_max  # Diluido en batch completo
            if decision == "parcial" and necesidad > 0:
                costo_fijo_und = costo_fijo / necesidad  # Menos diluido
        else:
            # Sin batch configurado — producir exacto
            producir = round(necesidad)
            batches = 1
            decision = "sin_config"
            razon = "Sin batch maximo configurado"
            costo_fijo_und = 0

        tiempo_produccion = round(batches * tiempo_batch) if tiempo_batch > 0 else 0
        tiempo_total_min += tiempo_produccion

        # Ingredientes necesarios (basado en producción real)
        porciones_base = lig["porciones_base"]
        factor_mp = producir / porciones_base if porciones_base > 0 else producir

        cur.execute("""
            SELECT ri.mp_id, mp.nombre, mp.unidad, ri.cantidad,
                   COALESCE(lc.precio_unit, 0), COALESCE(lc.proveedor, '')
            FROM receta_ingredientes ri
            JOIN materias_primas mp ON mp.id = ri.mp_id
            LEFT JOIN LATERAL (
                SELECT precio_unit, proveedor FROM compras_mp WHERE mp_id=ri.mp_id ORDER BY fecha DESC, id DESC LIMIT 1
            ) lc ON TRUE
            WHERE ri.receta_id = %s
        """, (lig["receta_id"],))
        ingredientes = []
        for ing in cur.fetchall():
            mp_id, mp_nombre, mp_unidad, cant_base, precio, proveedor = ing
            cant_necesaria = float(cant_base) * factor_mp

            mp_reglas = []
            dr = reglas.get(f"dias_recepcion:{mp_nombre}", "")
            vu = reglas.get(f"vida_util:{mp_nombre}", "")
            nf = reglas.get(f"no_fin_semana:{mp_nombre}", "")
            lt = reglas.get(f"dias_entrega:{proveedor}", "") if proveedor else ""
            if dr: mp_reglas.append(f"Solo llega: {dr}")
            if vu: mp_reglas.append(f"Vida util: {vu} dias")
            if nf: mp_reglas.append("No pedir para finde")
            if lt: mp_reglas.append(f"Lead time: {lt}d ({proveedor})")

            ingredientes.append({
                "mp_id": mp_id, "mp_nombre": mp_nombre, "unidad": mp_unidad,
                "cantidad_semanal": round(cant_necesaria, 2),
                "costo_semanal": round(cant_necesaria * float(precio), 2),
                "proveedor": proveedor, "reglas": mp_reglas
            })
            if mp_id not in total_mp_necesaria:
                total_mp_necesaria[mp_id] = {"nombre": mp_nombre, "unidad": mp_unidad,
                                              "cantidad": 0, "costo": 0, "proveedor": proveedor,
                                              "reglas": mp_reglas}
            total_mp_necesaria[mp_id]["cantidad"] += cant_necesaria
            total_mp_necesaria[mp_id]["costo"] += cant_necesaria * float(precio)

        plan_productos.append({
            "siigo_code": code, "producto": prod_venta["name"],
            "receta": lig["receta_nombre"],
            "venta_semanal": round(avg_qty_semana, 1),
            "stock_minimo": round(stock_minimo, 0),
            "stock_actual": stock_actual_pt,
            "necesidad": round(necesidad, 0),
            "producir": producir,
            "batches": round(batches, 1) if isinstance(batches, float) else batches,
            "batch_maximo": batch_max,
            "decision": decision, "razon": razon,
            "tiempo_min": tiempo_produccion,
            "costo_mp_und": round(sum(i["costo_semanal"] for i in ingredientes) / producir, 2) if producir > 0 else 0,
            "costo_mo_und": round(costo_mo / batch_max, 2) if batch_max > 0 else 0,
            "costo_sv_und": round(costo_sv / batch_max, 2) if batch_max > 0 else 0,
            "costo_total_und": round((sum(i["costo_semanal"] for i in ingredientes) / producir if producir > 0 else 0) + costo_fijo_und, 2),
            "revenue_semanal": round(avg_revenue, 2),
            "ingredientes": ingredientes
        })

    # 5. Plan de compras
    cur.execute("""
        SELECT mp.id, COALESCE(inv.cantidad_actual, 0)
        FROM materias_primas mp LEFT JOIN inventario inv ON inv.mp_id = mp.id
        WHERE mp.activo = TRUE
    """)
    stock_mp = {r[0]: float(r[1]) for r in cur.fetchall()}

    plan_compras = []
    for mp_id, d in sorted(total_mp_necesaria.items(), key=lambda x: -x[1]["costo"]):
        stock = stock_mp.get(mp_id, 0)
        necesita = round(d["cantidad"], 2)
        a_pedir = max(0, round(necesita - stock, 2))
        plan_compras.append({
            "mp_id": mp_id, "nombre": d["nombre"], "unidad": d["unidad"],
            "necesario_semanal": necesita, "stock_actual": stock,
            "a_pedir": a_pedir,
            "costo_estimado": round(d["costo"], 2),
            "proveedor": d["proveedor"], "reglas": d["reglas"]
        })

    cur.close(); conn.close()

    dias_produccion = len(config.get("dias_produccion", "lunes,martes,miercoles,jueves,viernes").split(","))
    return {
        "semanas_analizadas": num_weeks,
        "factor_stock_minimo": factor_stock,
        "plan_produccion": plan_productos,
        "plan_compras": plan_compras,
        "tiempo_total_min": tiempo_total_min,
        "minutos_disponibles_semana": minutos_dia * dias_produccion,
        "capacidad_usada_pct": round(tiempo_total_min / (minutos_dia * dias_produccion) * 100, 1) if minutos_dia > 0 else 0,
        "total_costo_mp": round(sum(m["costo_estimado"] for m in plan_compras), 2),
        "total_costo_fijos": round(sum(p.get("costo_mo_und", 0) * p["producir"] + p.get("costo_sv_und", 0) * p["producir"] for p in plan_productos), 2),
    }


# ═══════════════════════════════════════════════════════════════
# CLIENTES (CRM)
# ═══════════════════════════════════════════════════════════════

class ClienteIn(BaseModel):
    nombre: str
    direccion: Optional[str] = None
    apto: Optional[str] = None
    info_adicional: Optional[str] = None
    zona: Optional[str] = None
    telefono: Optional[str] = None
    cedula: Optional[str] = None
    email: Optional[str] = None

@router.get("/clientes")
def listar_clientes(q: Optional[str] = None, segmento: Optional[str] = None):
    conn = get_conn(); cur = conn.cursor()
    cur.execute("""
        SELECT c.id, c.nombre, c.direccion, c.apto, c.zona, c.telefono, c.cedula, c.email,
               c.shopify_customer_id, c.origen, c.fecha_registro,
               COUNT(v.id) AS num_pedidos,
               COALESCE(SUM(v.valor),0) AS total_gastado,
               MAX(v.fecha) AS ultimo_pedido
        FROM clientes c
        LEFT JOIN ventas v ON v.cliente_id = c.id
        WHERE c.activo = TRUE
        GROUP BY c.id
        ORDER BY ultimo_pedido DESC NULLS LAST, c.nombre
    """)
    rows = cur.fetchall()
    cur.close(); conn.close()
    result = []
    for r in rows:
        num = int(r[11]); total = float(r[12]); ultimo = str(r[13]) if r[13] else None
        promedio = total / num if num > 0 else 0
        # Calcular días sin comprar
        if r[13]:
            dias_sin = (date.today() - r[13]).days
        else:
            dias_sin = None
        # Segmento
        if num == 0:
            seg = 'prospect'
        elif num == 1:
            seg = 'unica'
        elif dias_sin and dias_sin > 60 and num >= 2:
            seg = 'inactivo'
        elif num >= 5:
            seg = 'vip'
        elif num >= 2:
            seg = 'frecuente'
        else:
            seg = 'unica'
        item = {
            "id": r[0], "nombre": r[1], "direccion": r[2], "apto": r[3],
            "zona": r[4], "telefono": r[5], "cedula": r[6], "email": r[7],
            "shopify_customer_id": r[8], "origen": r[9], "fecha_registro": str(r[10]) if r[10] else None,
            "num_pedidos": num, "total_gastado": total, "pedido_promedio": round(promedio, 2),
            "ultimo_pedido": ultimo, "dias_sin_compra": dias_sin, "segmento": seg
        }
        # Filtro por texto
        if q:
            ql = q.lower()
            if not (ql in (r[1] or '').lower() or ql in (r[5] or '').lower()
                    or ql in (r[7] or '').lower() or ql in (r[6] or '').lower()):
                continue
        # Filtro por segmento
        if segmento and segmento != 'todos' and seg != segmento:
            continue
        result.append(item)
    return result

@router.get("/clientes/{cliente_id}/perfil")
def perfil_cliente(cliente_id: int):
    conn = get_conn(); cur = conn.cursor()
    cur.execute("""
        SELECT id, nombre, direccion, apto, info_adicional, zona, telefono,
               cedula, email, shopify_customer_id, origen, fecha_registro
        FROM clientes WHERE id=%s
    """, (cliente_id,))
    c = cur.fetchone()
    if not c:
        raise HTTPException(404, "Cliente no encontrado")
    cur.execute("""
        SELECT id, fecha, factura, valor, estado, medio_pago, canal, notas,
               shopify_order_id, shopify_order_name
        FROM ventas WHERE cliente_id=%s ORDER BY fecha DESC
    """, (cliente_id,))
    ventas_list = [{"id": v[0], "fecha": str(v[1]), "factura": v[2], "valor": float(v[3]),
                    "estado": v[4], "medio_pago": v[5], "canal": v[6], "notas": v[7],
                    "shopify_order_id": v[8], "shopify_order_name": v[9]}
                   for v in cur.fetchall()]
    cur.close(); conn.close()

    num = len(ventas_list)
    total = sum(v["valor"] for v in ventas_list)
    promedio = total / num if num > 0 else 0
    # Frecuencia promedio (días entre pedidos)
    frecuencia_dias = None
    if num >= 2:
        fechas = sorted([datetime.strptime(v["fecha"], "%Y-%m-%d").date() for v in ventas_list])
        diffs = [(fechas[i+1] - fechas[i]).days for i in range(len(fechas)-1)]
        frecuencia_dias = round(sum(diffs) / len(diffs), 1) if diffs else None
    # Días sin comprar
    dias_sin = None
    if ventas_list:
        ultima = datetime.strptime(ventas_list[0]["fecha"], "%Y-%m-%d").date()
        dias_sin = (date.today() - ultima).days
    # Medio de pago favorito
    medios = {}
    canales = {}
    for v in ventas_list:
        if v["medio_pago"]:
            medios[v["medio_pago"]] = medios.get(v["medio_pago"], 0) + 1
        if v["canal"]:
            canales[v["canal"]] = canales.get(v["canal"], 0) + 1
    medio_fav = max(medios, key=medios.get) if medios else None
    canal_fav = max(canales, key=canales.get) if canales else None

    return {
        "id": c[0], "nombre": c[1], "direccion": c[2], "apto": c[3],
        "info_adicional": c[4], "zona": c[5], "telefono": c[6],
        "cedula": c[7], "email": c[8], "shopify_customer_id": c[9],
        "origen": c[10], "fecha_registro": str(c[11]) if c[11] else None,
        "metricas": {
            "total_gastado": round(total, 2),
            "num_pedidos": num,
            "pedido_promedio": round(promedio, 2),
            "frecuencia_dias": frecuencia_dias,
            "dias_sin_compra": dias_sin,
            "medio_pago_favorito": medio_fav,
            "canal_favorito": canal_fav
        },
        "ventas": ventas_list
    }

@router.post("/clientes")
def crear_cliente(data: ClienteIn):
    conn = get_conn(); cur = conn.cursor()
    cur.execute("""
        INSERT INTO clientes (nombre, direccion, apto, info_adicional, zona, telefono, cedula, email)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id
    """, (data.nombre, data.direccion, data.apto, data.info_adicional,
          data.zona, data.telefono, data.cedula, data.email))
    new_id = cur.fetchone()[0]
    conn.commit(); cur.close(); conn.close()
    return {"id": new_id, "ok": True}

@router.put("/clientes/{cliente_id}")
def editar_cliente(cliente_id: int, data: ClienteIn):
    conn = get_conn(); cur = conn.cursor()
    cur.execute("""
        UPDATE clientes SET nombre=%s, direccion=%s, apto=%s, info_adicional=%s,
               zona=%s, telefono=%s, cedula=%s, email=%s
        WHERE id=%s
    """, (data.nombre, data.direccion, data.apto, data.info_adicional,
          data.zona, data.telefono, data.cedula, data.email, cliente_id))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}


# ═══════════════════════════════════════════════════════════════
# VENTAS
# ═══════════════════════════════════════════════════════════════

class VentaIn(BaseModel):
    fecha: Optional[str] = None
    cliente_nombre: str
    cliente_id: Optional[int] = None
    factura: Optional[str] = None
    valor: float = 0
    estado: Optional[str] = "pendiente"
    medio_pago: Optional[str] = None
    canal: Optional[str] = None
    notas: Optional[str] = None

@router.get("/ventas")
def listar_ventas(fecha_desde: Optional[str] = None, fecha_hasta: Optional[str] = None,
                  estado: Optional[str] = None, medio_pago: Optional[str] = None,
                  canal: Optional[str] = None, q: Optional[str] = None, limit: int = 200):
    conn = get_conn(); cur = conn.cursor()
    where = ["1=1"]
    params = []
    if fecha_desde:
        where.append("v.fecha >= %s"); params.append(fecha_desde)
    if fecha_hasta:
        where.append("v.fecha <= %s"); params.append(fecha_hasta)
    if estado:
        where.append("v.estado = %s"); params.append(estado)
    if medio_pago:
        where.append("v.medio_pago = %s"); params.append(medio_pago)
    if canal:
        where.append("v.canal = %s"); params.append(canal)
    if q:
        where.append("(v.cliente_nombre ILIKE %s OR v.factura ILIKE %s)")
        params.extend([f"%{q}%", f"%{q}%"])
    params.append(limit)
    cur.execute(f"""
        SELECT v.id, v.fecha, v.cliente_nombre, v.cliente_id, v.factura, v.valor,
               v.estado, v.medio_pago, v.canal, v.notas,
               v.shopify_order_id, v.shopify_order_name
        FROM ventas v
        WHERE {' AND '.join(where)}
        ORDER BY v.fecha DESC, v.id DESC
        LIMIT %s
    """, params)
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id": r[0], "fecha": str(r[1]), "cliente_nombre": r[2], "cliente_id": r[3],
             "factura": r[4], "valor": float(r[5]), "estado": r[6], "medio_pago": r[7],
             "canal": r[8], "notas": r[9], "shopify_order_id": r[10],
             "shopify_order_name": r[11]} for r in rows]

@router.post("/ventas")
def crear_venta(data: VentaIn):
    conn = get_conn(); cur = conn.cursor()
    fecha = data.fecha or str(date.today())
    # Auto-vincular cliente si hay nombre exacto
    cliente_id = data.cliente_id
    if not cliente_id and data.cliente_nombre:
        cur.execute("SELECT id FROM clientes WHERE LOWER(nombre) = LOWER(%s) LIMIT 1",
                    (data.cliente_nombre,))
        match = cur.fetchone()
        if match:
            cliente_id = match[0]
    cur.execute("""
        INSERT INTO ventas (fecha, cliente_id, cliente_nombre, factura, valor, estado, medio_pago, canal, notas)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id
    """, (fecha, cliente_id, data.cliente_nombre, data.factura, data.valor,
          data.estado or 'pendiente', data.medio_pago, data.canal, data.notas))
    new_id = cur.fetchone()[0]
    conn.commit(); cur.close(); conn.close()
    return {"id": new_id, "ok": True}

@router.put("/ventas/{venta_id}")
def editar_venta(venta_id: int, data: VentaIn):
    conn = get_conn(); cur = conn.cursor()
    cur.execute("""
        UPDATE ventas SET fecha=%s, cliente_nombre=%s, cliente_id=%s, factura=%s,
               valor=%s, estado=%s, medio_pago=%s, canal=%s, notas=%s
        WHERE id=%s
    """, (data.fecha, data.cliente_nombre, data.cliente_id, data.factura,
          data.valor, data.estado, data.medio_pago, data.canal, data.notas, venta_id))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

@router.put("/ventas/{venta_id}/estado")
def cambiar_estado_venta(venta_id: int, estado: str):
    conn = get_conn(); cur = conn.cursor()
    cur.execute("UPDATE ventas SET estado=%s WHERE id=%s", (estado, venta_id))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

@router.delete("/ventas/{venta_id}")
def eliminar_venta(venta_id: int):
    conn = get_conn(); cur = conn.cursor()
    cur.execute("DELETE FROM ventas WHERE id=%s", (venta_id,))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}


# ═══════════════════════════════════════════════════════════════
# DASHBOARD VENTAS
# ═══════════════════════════════════════════════════════════════

@router.get("/ventas/dashboard")
def dashboard_ventas():
    conn = get_conn(); cur = conn.cursor()
    # KPIs del mes actual
    cur.execute("""
        SELECT COALESCE(SUM(valor),0), COUNT(*)
        FROM ventas WHERE fecha >= date_trunc('month', CURRENT_DATE)
    """)
    ventas_mes, ventas_mes_count = cur.fetchone()

    cur.execute("SELECT COALESCE(SUM(valor),0), COUNT(*) FROM ventas WHERE estado='pendiente'")
    cartera_pendiente, cartera_count = cur.fetchone()

    cur.execute("SELECT COUNT(DISTINCT cliente_nombre) FROM ventas")
    clientes_unicos = cur.fetchone()[0]

    # Recompra: clientes con 2+ compras / total clientes con compras
    cur.execute("""
        SELECT COUNT(*) FROM (
            SELECT cliente_nombre FROM ventas GROUP BY cliente_nombre HAVING COUNT(*) >= 2
        ) sub
    """)
    clientes_recompra = cur.fetchone()[0]
    pct_recompra = round(clientes_recompra / clientes_unicos * 100, 1) if clientes_unicos > 0 else 0

    cur.execute("SELECT COALESCE(AVG(valor),0) FROM ventas")
    pedido_promedio = float(cur.fetchone()[0])

    cur.execute("""
        SELECT COUNT(DISTINCT cliente_nombre) FROM ventas
        WHERE fecha >= date_trunc('month', CURRENT_DATE)
          AND cliente_nombre NOT IN (
              SELECT DISTINCT cliente_nombre FROM ventas WHERE fecha < date_trunc('month', CURRENT_DATE)
          )
    """)
    clientes_nuevos_mes = cur.fetchone()[0]

    # Ventas últimos 30 días
    cur.execute("""
        SELECT fecha, COALESCE(SUM(valor),0), COUNT(*)
        FROM ventas WHERE fecha >= CURRENT_DATE - INTERVAL '30 days'
        GROUP BY fecha ORDER BY fecha
    """)
    ventas_30d = [{"fecha": str(r[0]), "valor": float(r[1]), "count": int(r[2])} for r in cur.fetchall()]

    # Ventas por medio de pago
    cur.execute("""
        SELECT COALESCE(medio_pago,'Sin definir'), COUNT(*), COALESCE(SUM(valor),0)
        FROM ventas GROUP BY medio_pago ORDER BY COUNT(*) DESC
    """)
    por_medio = [{"medio": r[0], "count": int(r[1]), "valor": float(r[2])} for r in cur.fetchall()]

    # Ventas por canal
    cur.execute("""
        SELECT COALESCE(canal,'Sin definir'), COUNT(*), COALESCE(SUM(valor),0)
        FROM ventas GROUP BY canal ORDER BY COUNT(*) DESC
    """)
    por_canal = [{"canal": r[0], "count": int(r[1]), "valor": float(r[2])} for r in cur.fetchall()]

    # Segmentación clientes
    cur.execute("""
        SELECT segmento, COUNT(*) FROM (
            SELECT
                CASE
                    WHEN cnt = 0 THEN 'prospect'
                    WHEN cnt = 1 THEN 'unica'
                    WHEN cnt >= 5 AND dias_sin <= 60 THEN 'vip'
                    WHEN cnt >= 2 AND dias_sin <= 60 THEN 'frecuente'
                    WHEN cnt >= 2 AND dias_sin > 60 THEN 'inactivo'
                    ELSE 'unica'
                END AS segmento
            FROM (
                SELECT c.id,
                       COUNT(v.id) AS cnt,
                       COALESCE(CURRENT_DATE - MAX(v.fecha), 999) AS dias_sin
                FROM clientes c LEFT JOIN ventas v ON v.cliente_id = c.id
                WHERE c.activo = TRUE
                GROUP BY c.id
            ) sub
        ) seg
        GROUP BY segmento
    """)
    segmentacion = {r[0]: int(r[1]) for r in cur.fetchall()}

    # Top 10 clientes por valor
    cur.execute("""
        SELECT cliente_nombre, COUNT(*) AS cnt, SUM(valor) AS total, MAX(fecha) AS ultimo
        FROM ventas GROUP BY cliente_nombre
        ORDER BY total DESC LIMIT 10
    """)
    top_clientes = [{"nombre": r[0], "num_pedidos": int(r[1]),
                     "total": float(r[2]), "ultimo_pedido": str(r[3])} for r in cur.fetchall()]

    # Alertas: clientes que antes compraban frecuente pero llevan 30+ días sin comprar
    cur.execute("""
        SELECT c.id, c.nombre, c.telefono, c.email,
               COUNT(v.id) AS num_pedidos,
               SUM(v.valor) AS total_gastado,
               MAX(v.fecha) AS ultimo,
               CURRENT_DATE - MAX(v.fecha) AS dias_sin
        FROM clientes c
        JOIN ventas v ON v.cliente_id = c.id
        WHERE c.activo = TRUE
        GROUP BY c.id
        HAVING COUNT(v.id) >= 2 AND CURRENT_DATE - MAX(v.fecha) > 30
        ORDER BY SUM(v.valor) DESC
        LIMIT 10
    """)
    alertas_inactivos = [{"id": r[0], "nombre": r[1], "telefono": r[2], "email": r[3],
                          "num_pedidos": int(r[4]), "total_gastado": float(r[5]),
                          "ultimo_pedido": str(r[6]), "dias_sin": int(r[7])}
                         for r in cur.fetchall()]

    # Última sync Shopify
    cur.execute("SELECT fecha, tipo, registros_nuevos FROM shopify_sync_log ORDER BY fecha DESC LIMIT 1")
    sync_row = cur.fetchone()
    ultima_sync = {"fecha": str(sync_row[0]), "tipo": sync_row[1],
                   "registros": int(sync_row[2])} if sync_row else None

    cur.close(); conn.close()
    return {
        "ventas_mes": float(ventas_mes),
        "ventas_mes_count": int(ventas_mes_count),
        "cartera_pendiente": float(cartera_pendiente),
        "cartera_count": int(cartera_count),
        "clientes_unicos": int(clientes_unicos),
        "pct_recompra": pct_recompra,
        "clientes_recompra": int(clientes_recompra),
        "pedido_promedio": round(pedido_promedio, 2),
        "clientes_nuevos_mes": int(clientes_nuevos_mes),
        "ventas_30d": ventas_30d,
        "por_medio_pago": por_medio,
        "por_canal": por_canal,
        "segmentacion": segmentacion,
        "top_clientes": top_clientes,
        "alertas_inactivos": alertas_inactivos,
        "ultima_sync": ultima_sync
    }


# ═══════════════════════════════════════════════════════════════
# IMPORTACIÓN CSV (Ventas + Clientes desde Google Sheets)
# ═══════════════════════════════════════════════════════════════

class ImportClienteRow(BaseModel):
    nombre: str
    direccion: Optional[str] = None
    apto: Optional[str] = None
    info_adicional: Optional[str] = None
    zona: Optional[str] = None
    telefono: Optional[str] = None
    cedula: Optional[str] = None
    email: Optional[str] = None

class ImportVentaRow(BaseModel):
    fecha: Optional[str] = None
    cliente_nombre: str
    factura: Optional[str] = None
    valor: float = 0
    estado: Optional[str] = "pendiente"
    medio_pago: Optional[str] = None
    canal: Optional[str] = None

@router.post("/importar/clientes")
def importar_clientes(rows: List[ImportClienteRow]):
    conn = get_conn(); cur = conn.cursor()
    nuevos = 0; existentes = 0
    for r in rows:
        # Check si ya existe por nombre (case insensitive)
        cur.execute("SELECT id FROM clientes WHERE LOWER(nombre) = LOWER(%s) LIMIT 1", (r.nombre,))
        match = cur.fetchone()
        if match:
            existentes += 1
            continue
        cur.execute("""
            INSERT INTO clientes (nombre, direccion, apto, info_adicional, zona, telefono, cedula, email, origen)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'sheet_import')
        """, (r.nombre, r.direccion, r.apto, r.info_adicional, r.zona, r.telefono, r.cedula, r.email))
        nuevos += 1
    conn.commit(); cur.close(); conn.close()
    return {"ok": True, "nuevos": nuevos, "existentes": existentes}

@router.post("/importar/ventas")
def importar_ventas(rows: List[ImportVentaRow]):
    conn = get_conn(); cur = conn.cursor()
    nuevos = 0; dup = 0
    for r in rows:
        # Check duplicado por factura
        if r.factura:
            cur.execute("SELECT id FROM ventas WHERE factura = %s LIMIT 1", (r.factura,))
            if cur.fetchone():
                dup += 1; continue
        # Auto-vincular cliente
        cliente_id = None
        cur.execute("SELECT id FROM clientes WHERE LOWER(nombre) = LOWER(%s) LIMIT 1", (r.cliente_nombre,))
        cm = cur.fetchone()
        if cm:
            cliente_id = cm[0]
        else:
            # Crear cliente auto
            cur.execute("INSERT INTO clientes (nombre, origen) VALUES (%s, 'sheet_import') RETURNING id",
                        (r.cliente_nombre,))
            cliente_id = cur.fetchone()[0]
        cur.execute("""
            INSERT INTO ventas (fecha, cliente_id, cliente_nombre, factura, valor, estado, medio_pago, canal)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
        """, (r.fecha or str(date.today()), cliente_id, r.cliente_nombre, r.factura,
              r.valor, r.estado or 'pendiente', r.medio_pago, r.canal))
        nuevos += 1
    conn.commit(); cur.close(); conn.close()
    return {"ok": True, "nuevos": nuevos, "duplicados": dup}


# ═══════════════════════════════════════════════════════════════
# SHOPIFY SYNC (preparado para credenciales)
# ═══════════════════════════════════════════════════════════════

@router.get("/shopify/status")
def shopify_status():
    """Check if Shopify credentials are configured."""
    import os
    store = os.environ.get("SHOPIFY_STORE_URL", "")
    token = os.environ.get("SHOPIFY_ACCESS_TOKEN", "")
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT fecha, tipo, registros_nuevos, registros_actualizados FROM shopify_sync_log ORDER BY fecha DESC LIMIT 1")
    last = cur.fetchone()
    cur.close(); conn.close()
    return {
        "configured": bool(store and token),
        "store_url": store[:30] + "..." if len(store) > 30 else store,
        "last_sync": {
            "fecha": str(last[0]), "tipo": last[1],
            "nuevos": int(last[2]), "actualizados": int(last[3])
        } if last else None
    }

@router.post("/shopify/sync")
def shopify_sync():
    """Sync orders and customers from Shopify. Requires SHOPIFY_STORE_URL and SHOPIFY_ACCESS_TOKEN env vars."""
    import os, httpx
    store = os.environ.get("SHOPIFY_STORE_URL", "")
    token = os.environ.get("SHOPIFY_ACCESS_TOKEN", "")
    if not store or not token:
        raise HTTPException(400, "Shopify no configurado. Agregar SHOPIFY_STORE_URL y SHOPIFY_ACCESS_TOKEN en Railway.")

    headers = {"X-Shopify-Access-Token": token, "Content-Type": "application/json"}
    base = f"https://{store}/admin/api/2024-01"
    conn = get_conn(); cur = conn.cursor()

    # Get last sync date
    cur.execute("SELECT MAX(fecha) FROM shopify_sync_log WHERE tipo IN ('orders','full')")
    last_sync_row = cur.fetchone()
    last_sync = last_sync_row[0] if last_sync_row and last_sync_row[0] else None

    nuevos = 0; actualizados = 0; errores = 0

    try:
        # ── Sync Orders ──
        url = f"{base}/orders.json?status=any&limit=250"
        if last_sync:
            url += f"&updated_at_min={last_sync.isoformat()}"
        resp = httpx.get(url, headers=headers, timeout=30)
        resp.raise_for_status()
        orders = resp.json().get("orders", [])

        for order in orders:
            shopify_id = str(order["id"])
            cur.execute("SELECT id FROM ventas WHERE shopify_order_id = %s", (shopify_id,))
            existing = cur.fetchone()
            nombre = ""
            if order.get("customer"):
                nombre = f"{order['customer'].get('first_name','')} {order['customer'].get('last_name','')}".strip()
            if not nombre:
                nombre = order.get("email", "Sin nombre")
            fecha = order["created_at"][:10]
            valor = float(order.get("total_price", 0))
            financial = order.get("financial_status", "pending")
            estado = "pagado" if financial == "paid" else "pendiente"
            gateway = order.get("gateway", order.get("payment_gateway_names", [""])[0] if order.get("payment_gateway_names") else "")

            if existing:
                cur.execute("UPDATE ventas SET estado=%s WHERE id=%s", (estado, existing[0]))
                actualizados += 1
            else:
                # Match cliente: shopify_id → cédula → email → teléfono → nombre
                cliente_id = None
                if order.get("customer"):
                    cust = order["customer"]
                    cust_id = str(cust["id"])
                    email = (cust.get("email") or "").strip()
                    phone = re.sub(r"[^0-9+]", "", cust.get("phone") or "")
                    # Shopify puede guardar cédula en note_attributes
                    cedula = None
                    for attr in (cust.get("note_attributes") or []):
                        n = (attr.get("name") or "").lower()
                        if "cedula" in n or "nit" in n or "documento" in n or "identificacion" in n:
                            cedula = re.sub(r"[^0-9a-zA-Z\-]", "", str(attr.get("value") or ""))
                            break

                    # 1) Ya vinculado
                    cur.execute("SELECT id FROM clientes WHERE shopify_customer_id = %s", (cust_id,))
                    cm = cur.fetchone()
                    if cm:
                        cliente_id = cm[0]
                    else:
                        # 2) Por cédula (clave principal de cruce con historial)
                        if cedula:
                            cur.execute("SELECT id FROM clientes WHERE cedula = %s", (cedula,))
                            cm = cur.fetchone()
                        # 3) Por email
                        if not cm and email:
                            cur.execute("SELECT id FROM clientes WHERE email ILIKE %s", (email,))
                            cm = cur.fetchone()
                        # 4) Por teléfono
                        if not cm and phone:
                            cur.execute("SELECT id FROM clientes WHERE telefono = %s", (phone,))
                            cm = cur.fetchone()
                        # 5) Por nombre exacto
                        if not cm:
                            cur.execute("SELECT id FROM clientes WHERE nombre ILIKE %s LIMIT 1", (nombre,))
                            cm = cur.fetchone()

                        if cm:
                            cur.execute("UPDATE clientes SET shopify_customer_id=%s WHERE id=%s", (cust_id, cm[0]))
                            cliente_id = cm[0]
                        else:
                            cur.execute("""
                                INSERT INTO clientes (nombre, email, telefono, cedula, shopify_customer_id, origen)
                                VALUES (%s,%s,%s,%s,%s,'shopify') RETURNING id
                            """, (nombre, email, phone or None, cedula, cust_id))
                            cliente_id = cur.fetchone()[0]
                cur.execute("""
                    INSERT INTO ventas (fecha, cliente_id, cliente_nombre, factura, valor, estado,
                                        medio_pago, shopify_order_id, shopify_order_name)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (fecha, cliente_id, nombre, None, valor, estado,
                      gateway, shopify_id, order.get("name", "")))
                nuevos += 1

        conn.commit()
    except httpx.HTTPError as e:
        errores += 1
        conn.rollback()
        raise HTTPException(502, f"Error conectando a Shopify: {str(e)}")
    finally:
        cur.execute("""
            INSERT INTO shopify_sync_log (tipo, registros_nuevos, registros_actualizados, errores)
            VALUES ('orders', %s, %s, %s)
        """, (nuevos, actualizados, errores))
        conn.commit(); cur.close(); conn.close()

    return {"ok": True, "nuevos": nuevos, "actualizados": actualizados, "errores": errores}


# ═══════════════════════════════════════════════════════════════
# IMPORTACIÓN HISTÓRICA — LIBRO DE CARTERA DAILY.xlsx
# ═══════════════════════════════════════════════════════════════
import re, io
from datetime import date as date_type

def _clean_str(v):
    if v is None: return None
    s = str(v).strip().strip('\n').strip()
    return s if s else None

def _clean_cedula(v):
    if v is None: return None
    s = re.sub(r'\.0$', '', str(v).strip())
    s = re.sub(r'[^0-9a-zA-Z\-]', '', s)
    return s if s else None

def _clean_phone(v):
    if v is None: return None
    s = re.sub(r'\.0$', '', str(v).strip())
    s = re.sub(r'[^0-9+]', '', s)
    return s if s else None

def _to_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date_type): return v
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(str(v).strip(), fmt).date()
        except ValueError: pass
    return None

def _clean_valor(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    s = re.sub(r'[$,. ]', '', str(v).strip())
    try: return float(s)
    except ValueError: return None

def _normalize_medio(medio):
    if not medio: return medio
    m = medio.upper()
    if 'BOGOT' in m: return 'BANCO DE BOGOTA'
    if 'BANCOLOMBIA' in m: return 'BANCOLOMBIA'
    if 'LINK' in m or 'PAGO' in m: return 'LINK'
    if 'EFECTIVO' in m: return 'EFECTIVO'
    return medio.strip()


@router.post("/importar/excel-historico")
async def importar_excel_historico(file: UploadFile = File(...)):
    """Importa el LIBRO DE CARTERA DAILY.xlsx completo al sistema."""
    import openpyxl
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    conn = get_conn()
    cur = conn.cursor()
    result = {}

    # ── 1. BASE DATOS → clientes ──────────────────────────────────
    if ' BASE DATOS' in wb.sheetnames:
        ws = wb[' BASE DATOS']
        ins = upd = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            nombre = _clean_str(row[0])
            if not nombre: continue
            direccion = _clean_str(row[1])
            apto      = _clean_str(row[2])
            info_adic = _clean_str(row[3])
            zona      = _clean_str(row[4])
            telefono  = _clean_phone(row[5])
            cedula    = _clean_cedula(row[6])
            email     = _clean_str(row[7])

            if cedula:
                cur.execute("SELECT id FROM clientes WHERE cedula = %s", (cedula,))
                ex = cur.fetchone()
                if ex:
                    cur.execute("UPDATE clientes SET nombre=%s,direccion=%s,apto=%s,info_adicional=%s,zona=%s,telefono=%s,email=%s WHERE id=%s",
                                (nombre,direccion,apto,info_adic,zona,telefono,email,ex[0]))
                    upd += 1; continue

            cur.execute("SELECT id FROM clientes WHERE nombre ILIKE %s LIMIT 1", (nombre,))
            ex = cur.fetchone()
            if ex:
                cur.execute("""UPDATE clientes SET
                    direccion=COALESCE(clientes.direccion,%s), apto=COALESCE(clientes.apto,%s),
                    zona=COALESCE(clientes.zona,%s), telefono=COALESCE(clientes.telefono,%s),
                    cedula=COALESCE(clientes.cedula,%s), email=COALESCE(clientes.email,%s) WHERE id=%s""",
                    (direccion,apto,zona,telefono,cedula,email,ex[0]))
                upd += 1; continue

            cur.execute("INSERT INTO clientes (nombre,direccion,apto,info_adicional,zona,telefono,cedula,email,origen) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'sheet_import')",
                        (nombre,direccion,apto,info_adic,zona,telefono,cedula,email))
            ins += 1
        conn.commit()
        result['clientes_base_datos'] = {'insertados': ins, 'actualizados': upd}

    # ── 2. CLIENTES SIN COMPRA → prospects ───────────────────────
    if 'CLIENTES SIN COMPRA ' in wb.sheetnames:
        ws = wb['CLIENTES SIN COMPRA ']
        ins = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            nombre = _clean_str(row[0])
            if not nombre: continue
            cur.execute("SELECT id FROM clientes WHERE nombre ILIKE %s LIMIT 1", (nombre,))
            if cur.fetchone(): continue
            direccion = _clean_str(row[1])
            apto      = _clean_str(row[2])
            zona      = _clean_str(row[4]) if len(row) > 4 else None
            telefono  = _clean_phone(row[5]) if len(row) > 5 else None
            cur.execute("INSERT INTO clientes (nombre,direccion,apto,zona,telefono,origen) VALUES (%s,%s,%s,%s,%s,'prospect')",
                        (nombre,direccion,apto,zona,telefono))
            ins += 1
        conn.commit()
        result['prospects'] = {'insertados': ins}

    # ── Helper importar ventas ────────────────────────────────────
    def _import_ventas_sheet(ws, canal_default=None, fallback_year=2025):
        ins = skip = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            cliente_nombre = _clean_str(row[1])
            if not cliente_nombre: skip += 1; continue
            fecha   = _to_date(row[0]) or date_type(fallback_year, 12, 1)
            factura = _clean_str(row[2])
            valor   = _clean_valor(row[3])
            medio   = _normalize_medio(_clean_str(row[5]))
            canal   = _clean_str(row[6]) if len(row) > 6 else None
            if not canal: canal = canal_default
            if not valor or valor <= 0: skip += 1; continue
            if factura:
                cur.execute("SELECT id FROM ventas WHERE factura=%s", (factura,))
                if cur.fetchone(): skip += 1; continue
            cur.execute("SELECT id FROM clientes WHERE nombre ILIKE %s LIMIT 1", (cliente_nombre.strip(),))
            c = cur.fetchone()
            cur.execute("INSERT INTO ventas (fecha,cliente_id,cliente_nombre,factura,valor,estado,medio_pago,canal,notas) VALUES (%s,%s,%s,%s,%s,'pagado',%s,%s,'importado del historial')",
                        (fecha, c[0] if c else None, cliente_nombre, factura, valor, medio, canal))
            ins += 1
        conn.commit()
        return ins, skip

    # ── 3. SEPTIEMBRE 2025 ────────────────────────────────────────
    if 'SEPTIEMBRE 2025' in wb.sheetnames:
        ins, sk = _import_ventas_sheet(wb['SEPTIEMBRE 2025'], fallback_year=2025)
        result['septiembre_2025'] = {'insertadas': ins, 'ignoradas': sk}

    # ── 4. VENTAS NAVIDAD ─────────────────────────────────────────
    if 'VENTAS NAVIDAD' in wb.sheetnames:
        ins, sk = _import_ventas_sheet(wb['VENTAS NAVIDAD'], canal_default='NAVIDAD', fallback_year=2025)
        result['ventas_navidad'] = {'insertadas': ins, 'ignoradas': sk}

    # ── 5. VENTAS DAILY 2025 ──────────────────────────────────────
    if 'VENTAS DAILY 2025' in wb.sheetnames:
        ins, sk = _import_ventas_sheet(wb['VENTAS DAILY 2025'], fallback_year=2025)
        result['ventas_2025'] = {'insertadas': ins, 'ignoradas': sk}

    # ── 6. VENTAS DAILY 2026 ──────────────────────────────────────
    if 'VENTAS DAILY 2026' in wb.sheetnames:
        ins, sk = _import_ventas_sheet(wb['VENTAS DAILY 2026'], fallback_year=2026)
        result['ventas_2026'] = {'insertadas': ins, 'ignoradas': sk}

    cur.close(); conn.close()
    return {"ok": True, "resultado": result}


@router.post("/admin/migrate-clientes")
def admin_migrate_clientes():
    """Endpoint de emergencia: agrega columnas faltantes a la tabla clientes."""
    conn = get_conn()
    cur = conn.cursor()
    results = []
    cols = [
        ("direccion", "TEXT"),
        ("apto", "TEXT"),
        ("info_adicional", "TEXT"),
        ("zona", "TEXT"),
        ("telefono", "TEXT"),
        ("email", "TEXT"),
        ("cedula", "TEXT"),
        ("shopify_customer_id", "TEXT"),
        ("origen", "TEXT DEFAULT 'manual'"),
        ("fecha_registro", "DATE DEFAULT CURRENT_DATE"),
        ("activo", "BOOLEAN DEFAULT TRUE"),
    ]
    for col, defn in cols:
        try:
            cur.execute("SAVEPOINT sp")
            cur.execute(f"ALTER TABLE clientes ADD COLUMN IF NOT EXISTS {col} {defn}")
            cur.execute("RELEASE SAVEPOINT sp")
            results.append(f"{col}: ok")
        except Exception as e:
            cur.execute("ROLLBACK TO SAVEPOINT sp")
            results.append(f"{col}: {e}")
    conn.commit()
    cur.close()
    conn.close()


# ═══════════════════════════════════════════════════════════════
# VENTAS DAILY
# ═══════════════════════════════════════════════════════════════

class VentaIn(BaseModel):
    fecha_despacho: Optional[str] = None
    fecha_pago: Optional[str] = None
    cliente: str
    numero_factura: Optional[str] = None
    valor: Optional[float] = None
    medio_pago: Optional[str] = None
    canal: Optional[str] = None
    conciliacion: Optional[str] = None
    notas: Optional[str] = None

@router.post("/ventas-daily")
def crear_venta(data: VentaIn):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO ventas_daily
            (fecha_despacho, fecha_pago, cliente, numero_factura, valor, medio_pago, canal, conciliacion, notas)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id
    """, (data.fecha_despacho or None, data.fecha_pago or None, data.cliente,
          data.numero_factura, data.valor, data.medio_pago, data.canal, data.conciliacion, data.notas))
    new_id = cur.fetchone()[0]
    conn.commit(); cur.close(); conn.close()
    return {"id": new_id, "ok": True}

@router.put("/ventas-daily/{venta_id}")
def actualizar_venta(venta_id: int, data: VentaIn):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        UPDATE ventas_daily
        SET fecha_despacho=%s, fecha_pago=%s, cliente=%s, numero_factura=%s,
            valor=%s, medio_pago=%s, canal=%s, conciliacion=%s, notas=%s
        WHERE id=%s
    """, (data.fecha_despacho or None, data.fecha_pago or None, data.cliente,
          data.numero_factura, data.valor, data.medio_pago, data.canal, data.conciliacion, data.notas, venta_id))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

@router.delete("/ventas-daily/{venta_id}")
def eliminar_venta(venta_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM ventas_daily WHERE id=%s", (venta_id,))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

@router.post("/ventas-daily/import-sheet")
def importar_ventas_sheet():
    import re as _re, json, os
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        raise HTTPException(status_code=500, detail="gspread no instalado")

    creds_json = os.environ.get("GOOGLE_CREDENTIALS_JSON")
    if not creds_json:
        raise HTTPException(status_code=500, detail="GOOGLE_CREDENTIALS_JSON no configurado")
    import tempfile
    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as f:
        f.write(creds_json); tmp_path = f.name
    try:
        creds = Credentials.from_service_account_file(tmp_path, scopes=['https://www.googleapis.com/auth/spreadsheets.readonly'])
        gc = gspread.authorize(creds)
        sh = gc.open_by_key('1fMC3syH7ObP6-kf96zXruMYIYnBkxwyi4u1AxenFm7Q')
        ws = sh.worksheet('VENTAS DAILY 2026')
        rows = ws.get_all_values()
    finally:
        os.unlink(tmp_path)

    skip_clients = {'FEBRERO', 'MARZO', 'ABRIL', ' ', ''}

    def parse_date(s):
        s = s.strip()
        if not s: return None
        m = _re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', s)
        if m: return f'{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}'
        return None

    def parse_valor(s):
        s = s.strip().replace('$','').replace(',','').replace(' ','')
        try: return float(s)
        except: return None

    def map_medio(s):
        s = s.strip().upper()
        if not s: return None
        if s.startswith('LINK'): return 'LINK'
        if 'BANCOLOMBIA' in s: return 'BANCOLOMBIA'
        if 'EFECTIVO' in s and 'BANCOLOMBIA' not in s: return 'EFECTIVO'
        if 'EFECTIVO' in s: return 'BANCOLOMBIA'
        if 'BANCO DE BOGOTA' in s: return 'BANCO DE BOGOTA'
        if 'CRUCE' in s: return 'CRUCE'
        if 'AVALPAY' in s: return 'AVALPAY'
        if 'MARIA INES' in s or s == 'S.MARIA INES': return 'MARIA INES'
        return None

    def map_canal(s):
        s = s.strip().upper()
        if s == 'CLIENTE DAILY': return 'CLIENTE DAILY'
        if s in ('PAGINA','PÁGINA'): return 'PÁGINA'
        if s == 'INSTAGRAM': return 'INSTAGRAM'
        if s == 'TIKTOK': return 'TIK TOK'
        if s in ('REFERIDO','REFERID@'): return 'REFERIDO'
        if s == 'SIN INFORMACION': return 'SIN INFORMACIÓN'
        return None

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM ventas_daily")
    existing = cur.fetchone()[0]
    if existing > 0:
        cur.close(); conn.close()
        return {"ok": False, "detail": f"Ya hay {existing} registros. Usa force=true para reimportar.", "existing": existing}

    inserted = 0; skipped = 0
    for r in rows[1:]:
        cliente = r[2].strip() if len(r)>2 else ''
        if not cliente or cliente in skip_clients: skipped += 1; continue
        fd = parse_date(r[0]) if len(r)>0 else None
        fp = parse_date(r[1]) if len(r)>1 else None
        factura = r[3].strip() if len(r)>3 else None
        valor   = parse_valor(r[4]) if len(r)>4 else None
        medio   = map_medio(r[6]) if len(r)>6 else None
        canal   = map_canal(r[7]) if len(r)>7 else None
        cur.execute(
            "INSERT INTO ventas_daily (fecha_despacho, fecha_pago, cliente, numero_factura, valor, medio_pago, canal) VALUES (%s,%s,%s,%s,%s,%s,%s)",
            (fd, fp, cliente, factura or None, valor, medio, canal)
        )
        inserted += 1

    conn.commit(); cur.close(); conn.close()
    return {"ok": True, "inserted": inserted, "skipped": skipped}


@router.post("/ventas-daily/sync-desde-siigo")
def sync_ventas_desde_siigo():
    """Sincroniza ventas_daily desde crm_facturas (Siigo es fuente de verdad).
    - Facturas con match en ventas_daily → vincula factura_id y actualiza cliente/valor/fecha
    - Facturas sin match → crea fila nueva en ventas_daily
    - No sobreescribe: canal, medio_pago, notas del comercial
    """
    import re as _re
    conn = get_conn()
    cur = conn.cursor()

    # Traer todas las facturas de Siigo
    cur.execute("""
        SELECT f.id, f.numero, f.prefix, f.fecha, f.total, f.balance,
               COALESCE(NULLIF(TRIM(c.nombre),''), f.cliente_nombre) as nombre,
               f.estado_pago
        FROM crm_facturas f
        LEFT JOIN crm_clientes c ON c.id = f.cliente_id
        ORDER BY f.fecha DESC
    """)
    facturas = cur.fetchall()

    # Traer ventas_daily existentes con su número de factura y factura_id
    cur.execute("SELECT id, numero_factura, factura_id FROM ventas_daily")
    ventas = cur.fetchall()

    # Índice por número extraído (ej: "No. FE 3539" → 3539)
    def extraer_numero(s):
        if not s: return None
        m = _re.search(r'(\d{3,6})', str(s))
        return int(m.group(1)) if m else None

    ventas_por_num = {}
    for v in ventas:
        n = extraer_numero(v[1])
        if n: ventas_por_num[n] = v[0]  # num → venta_id

    vinculadas = 0; creadas = 0

    for f in facturas:
        fid, numero, prefix, fecha, total, balance, nombre, estado = f
        num = numero  # ya es int desde Siigo

        if num in ventas_por_num:
            # Actualizar campos de Siigo, respetar campos del comercial
            cur.execute("""
                UPDATE ventas_daily
                SET factura_id = %s,
                    cliente = COALESCE(NULLIF(cliente,''), %s),
                    valor = %s,
                    fecha_despacho = COALESCE(fecha_despacho, %s)
                WHERE id = %s AND (factura_id IS NULL OR factura_id = %s)
            """, (fid, nombre, float(total), fecha, ventas_por_num[num], fid))
            vinculadas += 1
        else:
            # Crear nueva fila desde Siigo
            factura_str = f"No. {prefix}-{numero}" if prefix else f"No. {numero}"
            cur.execute("""
                INSERT INTO ventas_daily (factura_id, cliente, numero_factura, valor, fecha_despacho)
                VALUES (%s, %s, %s, %s, %s)
                ON CONFLICT DO NOTHING
            """, (fid, nombre or '(Sin nombre)', factura_str, float(total), fecha))
            creadas += 1

    conn.commit(); cur.close(); conn.close()
    return {"ok": True, "vinculadas": vinculadas, "creadas": creadas, "total_facturas": len(facturas)}


@router.get("/ventas-daily")
def listar_ventas_daily_v2(
    desde: Optional[str] = None,
    hasta: Optional[str] = None,
    canal: Optional[str] = None,
    pagado: Optional[bool] = None,
    limit: int = 2000
):
    """Versión enriquecida: incluye estado de conciliación desde crm_facturas."""
    conn = get_conn()
    cur = conn.cursor()
    where = ["1=1"]
    params = []
    if desde:
        where.append("v.fecha_despacho >= %s"); params.append(desde)
    if hasta:
        where.append("v.fecha_despacho <= %s"); params.append(hasta)
    if canal:
        where.append("v.canal = %s"); params.append(canal)
    if pagado is True:
        where.append("v.fecha_pago IS NOT NULL")
    elif pagado is False:
        where.append("v.fecha_pago IS NULL")
    params.append(limit)
    cur.execute(f"""
        SELECT v.id, v.fecha_despacho, v.fecha_pago, v.cliente, v.numero_factura,
               v.valor, v.medio_pago, v.canal, v.conciliacion, v.notas, v.factura_id,
               f.estado_pago, f.rc_numero, f.balance,
               m.fecha AS fecha_pago_banco, m.banco AS banco_pago
        FROM ventas_daily v
        LEFT JOIN crm_facturas f ON v.factura_id = f.id
        LEFT JOIN movimientos_bancarios m ON f.movimiento_id = m.id
        WHERE {' AND '.join(where)}
        ORDER BY v.fecha_despacho DESC NULLS LAST, v.id DESC
        LIMIT %s
    """, params)
    rows = cur.fetchall()
    cur.close(); conn.close()

    result = []
    for r in rows:
        estado_pago = r[11]
        rc_numero = r[12]
        fecha_pago_banco = r[14]
        if rc_numero:
            estado_concil = "con_rc"
        elif estado_pago == 'pagado':
            estado_concil = "conciliado"
        elif r[10]:  # tiene factura_id
            estado_concil = "pendiente"
        else:
            estado_concil = "sin_factura"

        result.append({
            "id": r[0],
            "fecha_despacho": str(r[1]) if r[1] else None,
            "fecha_pago": str(r[2]) if r[2] else None,
            "cliente": r[3], "numero_factura": r[4],
            "valor": float(r[5]) if r[5] else None,
            "medio_pago": r[6], "canal": r[7],
            "conciliacion": r[8], "notas": r[9],
            "factura_id": r[10],
            "estado_conciliacion": estado_concil,
            "rc_numero": rc_numero,
            "balance": float(r[13]) if r[13] else None,
            "fecha_pago_banco": str(fecha_pago_banco) if fecha_pago_banco else None,
            "banco_pago": r[15]
        })
    return result
