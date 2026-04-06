"""
Script de importación histórica: LIBRO DE CARTERA DAILY.xlsx → PostgreSQL
Correr con: railway run python import_excel.py

Hojas importadas:
  - BASE DATOS         → clientes (origen='sheet_import')
  - CLIENTES SIN COMPRA → clientes (origen='prospect')
  - VENTAS DAILY 2025  → ventas
  - VENTAS NAVIDAD     → ventas (canal='NAVIDAD')
  - VENTAS DAILY 2026  → ventas
  - SEPTIEMBRE 2025    → ventas
  (AGOSTO 2025 excluida por instrucción del usuario)

Match clave para Shopify: cédula (columna 'Cedula' en BASE DATOS)
"""

import os, sys, re
import openpyxl
import psycopg2
from datetime import date, datetime

EXCEL = os.path.join(os.path.dirname(__file__), "LIBRO DE CARTERA DAILY.xlsx")

def get_conn():
    url = os.environ.get("DATABASE_URL")
    if not url:
        raise RuntimeError("DATABASE_URL no configurada")
    return psycopg2.connect(url)

def clean_str(v):
    if v is None:
        return None
    s = str(v).strip().strip('\n').strip()
    return s if s else None

def clean_cedula(v):
    if v is None:
        return None
    s = str(v).strip()
    # quitar decimales si viene como float
    s = re.sub(r'\.0$', '', s)
    s = re.sub(r'[^0-9a-zA-Z-]', '', s)
    return s if s else None

def clean_phone(v):
    if v is None:
        return None
    s = str(v).strip()
    s = re.sub(r'\.0$', '', s)   # 3100000000.0 → 3100000000
    s = re.sub(r'[^0-9+]', '', s)
    return s if s else None

def to_date(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.date() if isinstance(v, datetime) else v
    s = str(v).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None

def clean_valor(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace('$', '').replace(',', '').replace('.', '').strip()
    try:
        return float(s)
    except ValueError:
        return None

# ─────────────────────────────────────────────
# 1. IMPORTAR CLIENTES — BASE DATOS
# ─────────────────────────────────────────────
def import_clientes(conn, ws, origen='sheet_import'):
    cur = conn.cursor()
    inserted = 0
    updated = 0
    skip = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        nombre = clean_str(row[0])
        if not nombre:
            skip += 1
            continue

        direccion  = clean_str(row[1])
        apto       = clean_str(row[2])
        info_adic  = clean_str(row[3])
        zona       = clean_str(row[4])
        telefono   = clean_phone(row[5])
        cedula     = clean_cedula(row[6])
        email      = clean_str(row[7])

        # ¿Ya existe por cédula?
        if cedula:
            cur.execute("SELECT id FROM clientes WHERE cedula = %s", (cedula,))
            existing = cur.fetchone()
            if existing:
                cur.execute("""
                    UPDATE clientes SET nombre=%s, direccion=%s, apto=%s,
                    info_adicional=%s, zona=%s, telefono=%s, email=%s
                    WHERE id=%s
                """, (nombre, direccion, apto, info_adic, zona, telefono, email, existing[0]))
                updated += 1
                continue

        # ¿Ya existe por nombre exacto?
        cur.execute("SELECT id FROM clientes WHERE nombre ILIKE %s", (nombre,))
        existing = cur.fetchone()
        if existing:
            # Actualizar datos si faltaban
            cur.execute("""
                UPDATE clientes SET
                  direccion   = COALESCE(clientes.direccion, %s),
                  apto        = COALESCE(clientes.apto, %s),
                  zona        = COALESCE(clientes.zona, %s),
                  telefono    = COALESCE(clientes.telefono, %s),
                  cedula      = COALESCE(clientes.cedula, %s),
                  email       = COALESCE(clientes.email, %s)
                WHERE id=%s
            """, (direccion, apto, zona, telefono, cedula, email, existing[0]))
            updated += 1
            continue

        cur.execute("""
            INSERT INTO clientes
              (nombre, direccion, apto, info_adicional, zona, telefono, cedula, email, origen)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (nombre, direccion, apto, info_adic, zona, telefono, cedula, email, origen))
        inserted += 1

    conn.commit()
    return inserted, updated, skip

# ─────────────────────────────────────────────
# 2. IMPORTAR VENTAS — varias hojas
# ─────────────────────────────────────────────
def import_ventas(conn, ws, canal_default=None, fecha_header_col=0):
    """
    Columnas esperadas (0-indexed):
      0: Fecha
      1: Cliente
      2: # Factura
      3: Valor
      4: Valor Nota Crédito (ignorar)
      5: Medio de Pago
      6: Canal (PROVIENE CLIENTE) — opcional
    """
    cur = conn.cursor()
    inserted = 0
    skip = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        cliente_nombre = clean_str(row[1])
        if not cliente_nombre:
            skip += 1
            continue

        fecha   = to_date(row[fecha_header_col])
        factura = clean_str(row[2])
        valor   = clean_valor(row[3])
        medio   = clean_str(row[5])
        canal   = clean_str(row[6]) if len(row) > 6 else None

        if not canal:
            canal = canal_default

        if valor is None or valor <= 0:
            skip += 1
            continue

        if fecha is None:
            fecha = date(2025, 12, 1)   # fallback para VENTAS NAVIDAD sin fecha

        # Normalizar medio de pago
        if medio:
            medio = medio.upper()
            if 'BANCO' in medio and 'BOGOT' in medio:
                medio = 'BANCO DE BOGOTA'
            elif 'BANCOLOMBIA' in medio:
                medio = 'BANCOLOMBIA'
            elif 'LINK' in medio or 'PAGO' in medio:
                medio = 'LINK'
            elif 'EFECTIVO' in medio:
                medio = 'EFECTIVO'

        # Evitar duplicados por factura
        if factura:
            cur.execute("SELECT id FROM ventas WHERE factura = %s", (factura,))
            if cur.fetchone():
                skip += 1
                continue

        # Buscar cliente_id
        cur.execute("SELECT id FROM clientes WHERE nombre ILIKE %s LIMIT 1",
                    (cliente_nombre.strip(),))
        row_c = cur.fetchone()
        cliente_id = row_c[0] if row_c else None

        cur.execute("""
            INSERT INTO ventas
              (fecha, cliente_id, cliente_nombre, factura, valor, estado, medio_pago, canal, notas)
            VALUES (%s,%s,%s,%s,%s,'pagado',%s,%s,'importado del historial')
        """, (fecha, cliente_id, cliente_nombre, factura, valor, medio, canal))
        inserted += 1

    conn.commit()
    return inserted, skip

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    print(f"Leyendo {EXCEL}...")
    wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)
    conn = get_conn()

    # ── Clientes: BASE DATOS ──
    print("\n[1/5] Importando BASE DATOS → clientes...")
    ws = wb[' BASE DATOS']
    ins, upd, sk = import_clientes(conn, ws, origen='sheet_import')
    print(f"  ✓ {ins} nuevos, {upd} actualizados, {sk} vacíos ignorados")

    # ── Clientes: CLIENTES SIN COMPRA ──
    print("\n[2/5] Importando CLIENTES SIN COMPRA → prospects...")
    ws = wb['CLIENTES SIN COMPRA ']
    # Cols: 0=nombre, 1=dirección, 2=apto, 3=?, 4=zona, 5=teléfono
    # No tienen cédula ni email en esta hoja — adaptar manualmente
    cur = conn.cursor()
    ins2 = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        nombre = clean_str(row[0])
        if not nombre:
            continue
        direccion = clean_str(row[1])
        apto      = clean_str(row[2])
        zona      = clean_str(row[4]) if len(row) > 4 else None
        telefono  = clean_phone(row[5]) if len(row) > 5 else None

        cur.execute("SELECT id FROM clientes WHERE nombre ILIKE %s", (nombre,))
        if cur.fetchone():
            continue
        cur.execute("""
            INSERT INTO clientes (nombre, direccion, apto, zona, telefono, origen)
            VALUES (%s,%s,%s,%s,%s,'prospect')
        """, (nombre, direccion, apto, zona, telefono))
        ins2 += 1
    conn.commit()
    print(f"  ✓ {ins2} prospects nuevos")

    # ── Ventas: SEPTIEMBRE 2025 ──
    print("\n[3/5] Importando SEPTIEMBRE 2025...")
    ws = wb['SEPTIEMBRE 2025']
    ins, sk = import_ventas(conn, ws, canal_default=None)
    print(f"  ✓ {ins} ventas, {sk} ignoradas")

    # ── Ventas: VENTAS NAVIDAD ──
    print("\n[4/5] Importando VENTAS NAVIDAD...")
    ws = wb['VENTAS NAVIDAD']
    ins, sk = import_ventas(conn, ws, canal_default='NAVIDAD')
    print(f"  ✓ {ins} ventas, {sk} ignoradas")

    # ── Ventas: VENTAS DAILY 2025 ──
    print("\n[5/6] Importando VENTAS DAILY 2025...")
    ws = wb['VENTAS DAILY 2025']
    ins, sk = import_ventas(conn, ws, canal_default=None)
    print(f"  ✓ {ins} ventas, {sk} ignoradas")

    # ── Ventas: VENTAS DAILY 2026 ──
    print("\n[6/6] Importando VENTAS DAILY 2026...")
    ws = wb['VENTAS DAILY 2026']
    ins, sk = import_ventas(conn, ws, canal_default=None)
    print(f"  ✓ {ins} ventas, {sk} ignoradas")

    conn.close()
    print("\n✅ Importación completa.")

if __name__ == '__main__':
    main()
