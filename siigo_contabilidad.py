"""Siigo accounting sync — journals → trial balance → financial statements."""

import json
from collections import defaultdict
from datetime import datetime
from database import get_conn
from siigo import get_token, _headers, _paginate, SIIGO_BASE
import requests

# PUC classification for Colombian chart of accounts
PUC_CLASES = {
    '1': ('Activo', 'debit'),
    '2': ('Pasivo', 'credit'),
    '3': ('Patrimonio', 'credit'),
    '4': ('Ingresos', 'credit'),
    '5': ('Gastos', 'debit'),
    '6': ('Costo de Ventas', 'debit'),
    '7': ('Costos de Producción', 'debit'),
    '8': ('Cuentas de Orden Deudoras', 'debit'),
    '9': ('Cuentas de Orden Acreedoras', 'credit'),
}

PUC_GRUPOS = {
    '11': 'Efectivo y equivalentes', '12': 'Inversiones', '13': 'Deudores',
    '14': 'Inventarios', '15': 'Propiedad planta y equipo', '16': 'Intangibles',
    '17': 'Diferidos', '18': 'Otros activos',
    '21': 'Obligaciones financieras', '22': 'Proveedores', '23': 'Cuentas por pagar',
    '24': 'Impuestos por pagar', '25': 'Obligaciones laborales', '26': 'Pasivos estimados',
    '27': 'Diferidos pasivo', '28': 'Otros pasivos', '29': 'Bonos y papeles comerciales',
    '31': 'Capital social', '32': 'Superávit de capital', '33': 'Reservas',
    '34': 'Revalorización del patrimonio', '36': 'Resultados del ejercicio',
    '37': 'Resultados de ejercicios anteriores', '38': 'Superávit por valorizaciones',
    '41': 'Ingresos operacionales', '42': 'Ingresos no operacionales',
    '51': 'Gastos operacionales de administración', '52': 'Gastos operacionales de ventas',
    '53': 'Gastos no operacionales', '54': 'Impuesto de renta',
    '61': 'Costo de ventas', '62': 'Compras', '71': 'Costos de producción',
    '81': 'Derechos contingentes', '82': 'Deudoras fiscales', '83': 'Deudoras de control',
    '91': 'Responsabilidades contingentes', '92': 'Acreedoras fiscales',
    '93': 'Acreedoras de control',
}


def classify_account(code: str):
    """Classify a PUC account code."""
    clase_code = code[0] if code else '0'
    grupo_code = code[:2] if len(code) >= 2 else code
    clase_info = PUC_CLASES.get(clase_code, ('Otra', 'debit'))
    grupo_name = PUC_GRUPOS.get(grupo_code, '')
    return {
        'clase': clase_info[0],
        'naturaleza': clase_info[1],
        'grupo_puc': grupo_name
    }


def sync_journals():
    """
    Download accounting data from Siigo.
    Strategy:
    - Journals + Vouchers = real accounting entries with PUC codes (the source of truth)
    - Invoices = revenue detail (used for sales analytics, NOT for accounting entries —
      the journals/vouchers already contain the accounting impact of invoices)
    - Purchases = cost detail (same — already reflected in journals)
    """
    conn = get_conn()
    cur = conn.cursor()

    # Clear old synthetic entries (INV_, PUR_, CN_ prefixed) from previous bad sync
    cur.execute("DELETE FROM siigo_journals WHERE id LIKE 'INV_%' OR id LIKE 'PUR_%' OR id LIKE 'CN_%'")

    cuentas_seen = {}
    total = 0

    def process_doc(doc, source):
        nonlocal total
        jid = f"{source}_{doc['id']}"
        fecha = doc.get('date', '')[:10]
        items = doc.get('items', [])
        if not fecha or not items:
            return
        # Only items with account codes (real accounting entries)
        normalized = [item for item in items if 'account' in item and 'code' in item.get('account', {})]
        if not normalized:
            return
        cur.execute("""
            INSERT INTO siigo_journals (id, name, fecha, items, synced_at)
            VALUES (%s, %s, %s, %s, NOW())
            ON CONFLICT (id) DO UPDATE SET name=%s, fecha=%s, items=%s, synced_at=NOW()
        """, (jid, doc.get('name', ''), fecha, json.dumps(normalized),
              doc.get('name', ''), fecha, json.dumps(normalized)))
        for item in normalized:
            code = item['account']['code']
            if code not in cuentas_seen:
                cuentas_seen[code] = classify_account(code)
        total += 1

    # 1. Journals (manual entries, adjustments, closing entries — 552 docs)
    for doc in _paginate("/journals"):
        process_doc(doc, "JRN")

    # 2. Vouchers (in small batches with delays to avoid rate limits)
    from siigo import fetch_vouchers_paginated
    import time
    vch_page = 1
    vch_total = 0
    while True:
        time.sleep(1.5)  # Respect rate limits
        try:
            data = fetch_vouchers_paginated(page=vch_page, page_size=50)
        except Exception:
            break
        results = data.get('results', [])
        if not results:
            break
        for doc in results:
            process_doc(doc, "VCH")
            vch_total += 1
        total_expected = data.get('pagination', {}).get('total_results', 0)
        if vch_total >= total_expected:
            break
        vch_page += 1

    # Upsert accounts
    for code, info in cuentas_seen.items():
        cur.execute("""
            INSERT INTO siigo_cuentas (codigo, clase, grupo_puc, naturaleza)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (codigo) DO UPDATE SET clase=%s, grupo_puc=%s, naturaleza=%s
        """, (code, info['clase'], info['grupo_puc'], info['naturaleza'],
              info['clase'], info['grupo_puc'], info['naturaleza']))

    # Rebuild monthly balances
    _rebuild_saldos_mensuales(cur)

    # Log
    cur.execute("INSERT INTO sync_log (tipo, registros, detalle) VALUES ('full_sync', %s, %s)",
                (total, f"{len(cuentas_seen)} cuentas"))
    conn.commit()
    cur.close()
    conn.close()
    return {"registros": total, "cuentas": len(cuentas_seen)}


def _rebuild_saldos_mensuales(cur):
    """Rebuild monthly balances from all stored journals."""
    cur.execute("DELETE FROM saldos_mensuales")

    cur.execute("SELECT fecha, items FROM siigo_journals ORDER BY fecha")
    saldos = defaultdict(lambda: defaultdict(lambda: {'debito': 0, 'credito': 0}))

    for fecha, items_json in cur.fetchall():
        if not fecha:
            continue
        anio = fecha.year
        mes = fecha.month
        items = json.loads(items_json) if isinstance(items_json, str) else items_json
        for item in items:
            code = item['account']['code']
            val = float(item['value'])
            mov = item['account']['movement']
            key = (code, anio, mes)
            if mov == 'Debit':
                saldos[code][(anio, mes)]['debito'] += val
            else:
                saldos[code][(anio, mes)]['credito'] += val

    for code, periodos in saldos.items():
        for (anio, mes), vals in periodos.items():
            debito = round(vals['debito'], 2)
            credito = round(vals['credito'], 2)
            # Get naturaleza
            info = classify_account(code)
            if info['naturaleza'] == 'debit':
                saldo = debito - credito
            else:
                saldo = credito - debito
            cur.execute("""
                INSERT INTO saldos_mensuales (cuenta, anio, mes, debito, credito, saldo)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (code, anio, mes, debito, credito, saldo))


def get_balance_prueba(anio: int, mes: int):
    """Get trial balance for a specific month (movements of that month only)."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT s.cuenta, COALESCE(c.nombre, s.cuenta), c.clase, c.grupo_puc, c.naturaleza,
               s.debito, s.credito, s.saldo
        FROM saldos_mensuales s
        LEFT JOIN siigo_cuentas c ON c.codigo = s.cuenta
        WHERE s.anio = %s AND s.mes = %s
        ORDER BY s.cuenta
    """, (anio, mes))
    rows = cur.fetchall()
    cur.close(); conn.close()

    result = []
    total_debito = 0
    total_credito = 0
    for r in rows:
        total_debito += float(r[5])
        total_credito += float(r[6])
        result.append({
            "cuenta": r[0], "nombre": r[1] or r[0], "clase": r[2],
            "grupo_puc": r[3], "naturaleza": r[4],
            "debito": float(r[5]), "credito": float(r[6]), "saldo": float(r[7])
        })
    return {
        "anio": anio, "mes": mes,
        "cuentas": result,
        "total_debito": round(total_debito, 2),
        "total_credito": round(total_credito, 2)
    }


def get_estado_resultados(anio: int, mes_inicio: int = 1, mes_fin: int = 12):
    """
    P&L combining:
    - Revenue from invoices - credit notes - taxes (impoconsumo)
    - Costs from purchases (6xxx + 7xxx)
    - Expenses from purchases (5xxx) + vouchers/journals (nómina, financieros)
    """
    from siigo import fetch_invoices, fetch_purchases, fetch_credit_notes

    start = f"{anio}-{mes_inicio:02d}-01"
    last_day = 28 if mes_fin == 2 else 30 if mes_fin in (4,6,9,11) else 31
    end = f"{anio}-{mes_fin:02d}-{last_day}"

    # 1. Revenue from invoices (net of taxes)
    invoices = fetch_invoices(start, end)
    ingresos_brutos = 0
    impoconsumo = 0
    for inv in invoices:
        if inv.get('annulled'):
            continue
        ingresos_brutos += float(inv.get('total', 0))
        for item in inv.get('items', []):
            for tax in item.get('taxes', []):
                impoconsumo += float(tax.get('value', 0))

    # 2. Credit notes (devoluciones)
    credit_notes = fetch_credit_notes(start, end)
    devoluciones = sum(float(cn.get('total', 0)) for cn in credit_notes if not cn.get('annulled'))

    ingresos_netos = ingresos_brutos - impoconsumo - devoluciones

    # 3. Costs and expenses from purchases
    purchases = fetch_purchases(start, end)
    costos_map = defaultdict(float)
    gastos_admin_map = defaultdict(float)
    gastos_ventas_map = defaultdict(float)
    gastos_no_op_map = defaultdict(float)

    for pur in purchases:
        if pur.get('annulled'):
            continue
        for item in pur.get('items', []):
            code = item.get('code', '')
            total_val = float(item.get('total', 0))
            desc = item.get('description', code)
            if code.startswith('6') or code.startswith('7'):
                costos_map[f"{code} {desc}"] += total_val
            elif code.startswith('51'):
                gastos_admin_map[f"{code} {desc}"] += total_val
            elif code.startswith('52'):
                gastos_ventas_map[f"{code} {desc}"] += total_val
            elif code.startswith('53') or code.startswith('54'):
                gastos_no_op_map[f"{code} {desc}"] += total_val

    # Note: P&L uses ONLY purchases as source for costs/expenses.
    # Journals contain annual closing entries that would duplicate amounts.
    # Nómina and other expenses registered via vouchers are reflected in purchases.

    def to_items(m):
        return sorted([{"cuenta": k.split(' ')[0], "nombre": ' '.join(k.split(' ')[1:]) or k.split(' ')[0],
                        "clase": "", "grupo_puc": "", "saldo": round(v, 2)}
                       for k, v in m.items() if v > 0], key=lambda x: -x['saldo'])

    costos = to_items(costos_map)
    gastos_admin = to_items(gastos_admin_map)
    gastos_ventas = to_items(gastos_ventas_map)
    gastos_no_op = to_items(gastos_no_op_map)

    total_costos = sum(i['saldo'] for i in costos)
    total_gastos_admin = sum(i['saldo'] for i in gastos_admin)
    total_gastos_ventas = sum(i['saldo'] for i in gastos_ventas)
    total_gastos_no_op = sum(i['saldo'] for i in gastos_no_op)
    utilidad_bruta = ingresos_netos - total_costos
    utilidad_operacional = utilidad_bruta - total_gastos_admin - total_gastos_ventas
    utilidad_antes_imp = utilidad_operacional - total_gastos_no_op

    ingresos_items = [
        {"cuenta": "Facturas", "nombre": "Ingresos brutos (facturas)", "clase": "Ingresos",
         "grupo_puc": "", "saldo": round(ingresos_brutos, 2)},
        {"cuenta": "Impoconsumo", "nombre": "(-) Impoconsumo", "clase": "Ingresos",
         "grupo_puc": "", "saldo": round(-impoconsumo, 2)},
        {"cuenta": "NC", "nombre": "(-) Devoluciones (notas crédito)", "clase": "Ingresos",
         "grupo_puc": "", "saldo": round(-devoluciones, 2)},
    ]

    return {
        "anio": anio, "mes_inicio": mes_inicio, "mes_fin": mes_fin,
        "ingresos": {"items": ingresos_items, "total": round(ingresos_netos, 2)},
        "ingresos_brutos": round(ingresos_brutos, 2),
        "impoconsumo": round(impoconsumo, 2),
        "devoluciones": round(devoluciones, 2),
        "costos_ventas": {"items": costos, "total": round(total_costos, 2)},
        "utilidad_bruta": round(utilidad_bruta, 2),
        "margen_bruto_pct": round(utilidad_bruta / ingresos_netos * 100, 1) if ingresos_netos > 0 else 0,
        "gastos_admin": {"items": gastos_admin, "total": round(total_gastos_admin, 2)},
        "gastos_ventas": {"items": gastos_ventas, "total": round(total_gastos_ventas, 2)},
        "utilidad_operacional": round(utilidad_operacional, 2),
        "margen_operacional_pct": round(utilidad_operacional / ingresos_netos * 100, 1) if ingresos_netos > 0 else 0,
        "gastos_no_operacionales": {"items": gastos_no_op, "total": round(total_gastos_no_op, 2)},
        "utilidad_neta": round(utilidad_antes_imp, 2),
        "margen_neto_pct": round(utilidad_antes_imp / ingresos_netos * 100, 1) if ingresos_netos > 0 else 0,
    }


def get_indicadores(anio: int, mes: int):
    """Key financial indicators from invoices, credit notes, purchases, journals."""
    from siigo import fetch_invoices, fetch_purchases, fetch_credit_notes

    start = f"{anio}-01-01"
    last_day = 28 if mes == 2 else 30 if mes in (4,6,9,11) else 31
    end = f"{anio}-{mes:02d}-{last_day}"

    # Revenue (net of taxes and credit notes)
    invoices = fetch_invoices(start, end)
    ingresos_brutos = 0
    impuestos = 0
    for inv in invoices:
        if inv.get('annulled'): continue
        ingresos_brutos += float(inv.get('total', 0))
        for item in inv.get('items', []):
            for tax in item.get('taxes', []):
                impuestos += float(tax.get('value', 0))
    credit_notes = fetch_credit_notes(start, end)
    devoluciones = sum(float(cn.get('total', 0)) for cn in credit_notes if not cn.get('annulled'))
    ingresos = ingresos_brutos - impuestos - devoluciones

    # Costs/expenses from purchases
    purchases = fetch_purchases(start, end)
    costos = 0
    gastos_compras = 0
    for pur in purchases:
        if pur.get('annulled'): continue
        for item in pur.get('items', []):
            code = item.get('code', '')
            val = float(item.get('total', 0))
            if code.startswith('6') or code.startswith('7'):
                costos += val
            elif code.startswith('5'):
                gastos_compras += val

    # Only use purchases for expenses (journals have closing entries that duplicate)
    gastos = gastos_compras

    # Balance from journal saldos
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT s.cuenta, SUM(s.debito) as total_d, SUM(s.credito) as total_c
        FROM saldos_mensuales s
        WHERE (s.anio < %s) OR (s.anio = %s AND s.mes <= %s)
        GROUP BY s.cuenta
    """, (anio, anio, mes))
    raw = {r[0]: {"debito": float(r[1]), "credito": float(r[2])} for r in cur.fetchall()}
    cur.close(); conn.close()

    # Calculate balance from raw debits/credits
    def balance_grupo(prefixes):
        total = 0
        for code, vals in raw.items():
            if any(code.startswith(p) for p in prefixes):
                info = classify_account(code)
                if info['naturaleza'] == 'debit':
                    total += vals['debito'] - vals['credito']
                else:
                    total += vals['credito'] - vals['debito']
        return total

    activo_corriente = balance_grupo(['11', '12', '13', '14'])
    activo_total = balance_grupo(['1'])
    pasivo_corriente = balance_grupo(['21', '22', '23', '24', '25'])
    pasivo_total = balance_grupo(['2'])
    patrimonio = balance_grupo(['3'])

    utilidad_bruta = ingresos - costos
    utilidad_operacional = utilidad_bruta - gastos

    return {
        "anio": anio, "mes": mes,
        "balance": {
            "activo_corriente": round(activo_corriente, 2),
            "pasivo_corriente": round(pasivo_corriente, 2),
            "activo_total": round(activo_total, 2),
            "pasivo_total": round(pasivo_total, 2),
            "patrimonio": round(patrimonio, 2),
        },
        "resultados": {
            "ingresos": round(ingresos, 2),
            "costos": round(costos, 2),
            "gastos": round(gastos, 2),
            "utilidad_bruta": round(utilidad_bruta, 2),
            "utilidad_operacional": round(utilidad_operacional, 2),
            "ebitda": round(utilidad_operacional, 2),
        },
        "ratios": {
            "liquidez": round(activo_corriente / pasivo_corriente, 2) if pasivo_corriente > 0 else 0,
            "endeudamiento_pct": round(pasivo_total / activo_total * 100, 1) if activo_total > 0 else 0,
            "margen_bruto_pct": round(utilidad_bruta / ingresos * 100, 1) if ingresos > 0 else 0,
            "margen_operacional_pct": round(utilidad_operacional / ingresos * 100, 1) if ingresos > 0 else 0,
            "margen_neto_pct": round(utilidad_operacional / ingresos * 100, 1) if ingresos > 0 else 0,
            "roe_pct": round(utilidad_operacional / patrimonio * 100, 1) if patrimonio > 0 else 0,
            "roa_pct": round(utilidad_operacional / activo_total * 100, 1) if activo_total > 0 else 0,
        }
    }


def get_tendencia_mensual_from_invoices(anio: int):
    """Monthly P&L trend from invoices, credit notes, purchases, and journal expenses."""
    from siigo import fetch_invoices, fetch_purchases, fetch_credit_notes

    meses = {}
    for mes in range(1, 13):
        meses[mes] = {"ingresos_brutos": 0, "impuestos": 0, "devoluciones": 0,
                      "costos": 0, "gastos": 0}

    start = f"{anio}-01-01"
    end = f"{anio}-12-31"

    # Invoices = revenue (net of taxes)
    invoices = fetch_invoices(start, end)
    for inv in invoices:
        if inv.get('annulled'):
            continue
        fecha = inv.get('date', '')
        if not fecha: continue
        mes = int(fecha[5:7])
        meses[mes]["ingresos_brutos"] += float(inv.get('total', 0))
        for item in inv.get('items', []):
            for tax in item.get('taxes', []):
                meses[mes]["impuestos"] += float(tax.get('value', 0))

    # Credit notes = devoluciones
    credit_notes = fetch_credit_notes(start, end)
    for cn in credit_notes:
        if cn.get('annulled'): continue
        fecha = cn.get('date', '')
        if not fecha: continue
        mes = int(fecha[5:7])
        meses[mes]["devoluciones"] += float(cn.get('total', 0))

    # Purchases = costs (6+7xxx) and expenses (5xxx)
    purchases = fetch_purchases(start, end)
    for pur in purchases:
        if pur.get('annulled'): continue
        fecha = pur.get('date', '')
        if not fecha: continue
        mes = int(fecha[5:7])
        for item in pur.get('items', []):
            code = item.get('code', '')
            val = float(item.get('total', 0))
            if code.startswith('6') or code.startswith('7'):
                meses[mes]["costos"] += val
            elif code.startswith('5'):
                meses[mes]["gastos"] += val

    # Expenses come only from purchases (journals have annual closing entries that duplicate)
    cur.close(); conn.close()

    result = []
    for mes in range(1, 13):
        m = meses[mes]
        ingresos = m["ingresos_brutos"] - m["impuestos"] - m["devoluciones"]
        if ingresos == 0 and m["costos"] == 0 and m["gastos"] == 0:
            continue
        utilidad_bruta = ingresos - m["costos"]
        utilidad_neta = utilidad_bruta - m["gastos"]
        result.append({
            "mes": mes,
            "ingresos": round(ingresos, 2),
            "costos": round(m["costos"], 2),
            "gastos": round(m["gastos"], 2),
            "utilidad_bruta": round(utilidad_bruta, 2),
            "utilidad_neta": round(utilidad_neta, 2),
            "margen_bruto_pct": round(utilidad_bruta / ingresos * 100, 1) if ingresos > 0 else 0,
            "margen_neto_pct": round(utilidad_neta / ingresos * 100, 1) if ingresos > 0 else 0,
        })
    return {"anio": anio, "meses": result}


def get_tendencia_mensual(anio: int):
    """Monthly P&L trend for a full year."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT s.mes,
               SUM(CASE WHEN s.cuenta LIKE '4%%' THEN s.saldo ELSE 0 END) as ingresos,
               SUM(CASE WHEN s.cuenta LIKE '6%%' THEN s.saldo ELSE 0 END) as costos,
               SUM(CASE WHEN s.cuenta LIKE '5%%' THEN s.saldo ELSE 0 END) as gastos
        FROM saldos_mensuales s
        WHERE s.anio = %s
        GROUP BY s.mes ORDER BY s.mes
    """, (anio,))
    rows = cur.fetchall()
    cur.close(); conn.close()

    meses = []
    for r in rows:
        ingresos = abs(float(r[1]))
        costos = abs(float(r[2]))
        gastos = abs(float(r[3]))
        utilidad_bruta = ingresos - costos
        utilidad_neta = utilidad_bruta - gastos
        meses.append({
            "mes": r[0],
            "ingresos": round(ingresos, 2),
            "costos": round(costos, 2),
            "gastos": round(gastos, 2),
            "utilidad_bruta": round(utilidad_bruta, 2),
            "utilidad_neta": round(utilidad_neta, 2),
            "margen_bruto_pct": round(utilidad_bruta / ingresos * 100, 1) if ingresos > 0 else 0,
            "margen_neto_pct": round(utilidad_neta / ingresos * 100, 1) if ingresos > 0 else 0,
        })
    return {"anio": anio, "meses": meses}


def get_presupuesto_vs_real(anio: int, mes: int):
    """Compare budget vs actual for a given month."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT p.cuenta, p.cuenta_nombre, p.monto AS presupuesto,
               COALESCE(s.saldo, 0) AS real,
               c.clase, c.grupo_puc
        FROM presupuestos p
        LEFT JOIN saldos_mensuales s ON s.cuenta = p.cuenta AND s.anio = p.anio AND s.mes = p.mes
        LEFT JOIN siigo_cuentas c ON c.codigo = p.cuenta
        WHERE p.anio = %s AND p.mes = %s
        ORDER BY p.cuenta
    """, (anio, mes))
    rows = cur.fetchall()
    cur.close(); conn.close()

    result = []
    for r in rows:
        presupuesto = float(r[2])
        real = abs(float(r[3]))
        desviacion = real - presupuesto
        desviacion_pct = (desviacion / presupuesto * 100) if presupuesto != 0 else 0
        # For expenses: over budget is bad. For revenue: under budget is bad.
        es_gasto = (r[4] or '').startswith('Gasto') or (r[4] or '').startswith('Costo')
        alerta = ''
        if es_gasto and desviacion_pct > 10:
            alerta = 'sobre_presupuesto'
        elif not es_gasto and desviacion_pct < -10:
            alerta = 'bajo_presupuesto'
        result.append({
            "cuenta": r[0], "nombre": r[1] or r[0],
            "presupuesto": round(presupuesto, 2), "real": round(real, 2),
            "desviacion": round(desviacion, 2),
            "desviacion_pct": round(desviacion_pct, 1),
            "clase": r[4], "grupo_puc": r[5], "alerta": alerta
        })
    return {"anio": anio, "mes": mes, "items": result}


def get_balance_general(anio: int, mes: int):
    """Balance General formal con estructura NIIF."""
    conn = get_conn()
    cur = conn.cursor()
    # Saldos acumulados hasta el mes (solo cuentas 1-3xxx)
    cur.execute("""
        SELECT s.cuenta, COALESCE(c.nombre, s.cuenta), c.clase, c.grupo_puc, c.naturaleza,
               SUM(s.debito), SUM(s.credito)
        FROM saldos_mensuales s
        LEFT JOIN siigo_cuentas c ON c.codigo = s.cuenta
        WHERE (s.anio < %s OR (s.anio = %s AND s.mes <= %s))
          AND s.cuenta LIKE ANY(ARRAY['1%%','2%%','3%%'])
        GROUP BY s.cuenta, c.nombre, c.clase, c.grupo_puc, c.naturaleza
        ORDER BY s.cuenta
    """, (anio, anio, mes))
    rows = cur.fetchall()
    cur.close(); conn.close()

    activo_corriente = []
    activo_no_corriente = []
    pasivo_corriente = []
    pasivo_no_corriente = []
    patrimonio_items = []

    for r in rows:
        code = r[0]
        nat = r[4] or ('debit' if code[0] == '1' else 'credit')
        debito = float(r[5])
        credito = float(r[6])
        saldo = (debito - credito) if nat == 'debit' else (credito - debito)
        if abs(saldo) < 0.01:
            continue
        item = {"cuenta": code, "nombre": r[1] or code, "grupo_puc": r[3] or '', "saldo": round(saldo, 2)}
        prefix = code[:2]
        if code[0] == '1':
            if prefix in ('11', '12', '13', '14'):
                activo_corriente.append(item)
            else:
                activo_no_corriente.append(item)
        elif code[0] == '2':
            if prefix in ('21', '22', '23', '24', '25'):
                pasivo_corriente.append(item)
            else:
                pasivo_no_corriente.append(item)
        elif code[0] == '3':
            patrimonio_items.append(item)

    t_ac = sum(i['saldo'] for i in activo_corriente)
    t_anc = sum(i['saldo'] for i in activo_no_corriente)
    t_activo = t_ac + t_anc
    t_pc = sum(i['saldo'] for i in pasivo_corriente)
    t_pnc = sum(i['saldo'] for i in pasivo_no_corriente)
    t_pasivo = t_pc + t_pnc
    t_patrimonio = sum(i['saldo'] for i in patrimonio_items)

    return {
        "anio": anio, "mes": mes,
        "activo_corriente": {"items": activo_corriente, "total": round(t_ac, 2)},
        "activo_no_corriente": {"items": activo_no_corriente, "total": round(t_anc, 2)},
        "total_activo": round(t_activo, 2),
        "pasivo_corriente": {"items": pasivo_corriente, "total": round(t_pc, 2)},
        "pasivo_no_corriente": {"items": pasivo_no_corriente, "total": round(t_pnc, 2)},
        "total_pasivo": round(t_pasivo, 2),
        "patrimonio": {"items": patrimonio_items, "total": round(t_patrimonio, 2)},
        "total_pasivo_patrimonio": round(t_pasivo + t_patrimonio, 2),
        "cuadre": round(abs(t_activo - (t_pasivo + t_patrimonio)), 2)
    }


def get_comparativo(anio1: int, mes_inicio1: int, mes_fin1: int,
                    anio2: int, mes_inicio2: int, mes_fin2: int):
    """Comparativo de P&L entre dos períodos."""
    pl1 = get_estado_resultados(anio1, mes_inicio1, mes_fin1)
    pl2 = get_estado_resultados(anio2, mes_inicio2, mes_fin2)

    def var(v1, v2):
        diff = v1 - v2
        pct = (diff / abs(v2) * 100) if v2 != 0 else 0
        return {"valor": round(diff, 2), "pct": round(pct, 1)}

    return {
        "periodo1": {"anio": anio1, "mes_inicio": mes_inicio1, "mes_fin": mes_fin1, "data": pl1},
        "periodo2": {"anio": anio2, "mes_inicio": mes_inicio2, "mes_fin": mes_fin2, "data": pl2},
        "variaciones": {
            "ingresos": var(pl1["ingresos"]["total"], pl2["ingresos"]["total"]),
            "costos": var(pl1["costos_ventas"]["total"], pl2["costos_ventas"]["total"]),
            "utilidad_bruta": var(pl1["utilidad_bruta"], pl2["utilidad_bruta"]),
            "gastos_admin": var(pl1["gastos_admin"]["total"], pl2["gastos_admin"]["total"]),
            "gastos_ventas": var(pl1["gastos_ventas"]["total"], pl2["gastos_ventas"]["total"]),
            "utilidad_operacional": var(pl1["utilidad_operacional"], pl2["utilidad_operacional"]),
            "utilidad_neta": var(pl1["utilidad_neta"], pl2["utilidad_neta"]),
        }
    }


def export_eeff_excel(anio: int, mes_inicio: int, mes_fin: int):
    """Generate Excel file with formal financial statements."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers

    wb = Workbook()
    bold = Font(bold=True)
    bold_big = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    subtotal_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    total_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    total_font = Font(bold=True, size=12, color="FFFFFF")
    num_fmt = '#,##0'
    thin_border = Border(bottom=Side(style='thin'))
    meses_nombres = ['','Enero','Febrero','Marzo','Abril','Mayo','Junio',
                     'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']

    periodo_str = f"{meses_nombres[mes_inicio]} a {meses_nombres[mes_fin]} {anio}"

    def write_header(ws, titulo):
        ws.append(["DAILY FOOD SOLUTIONS SAS"])
        ws['A1'].font = bold_big
        ws.append(["NIT 901.871.505-2"])
        ws.append([titulo])
        ws['A3'].font = Font(bold=True, size=12)
        ws.append([f"Del 1 de {meses_nombres[mes_inicio]} al {mes_fin if mes_fin < 28 else ''} de {meses_nombres[mes_fin]} de {anio}"])
        ws.append(["(Expresados en pesos colombianos)"])
        ws.append([])

    def fmt_row(ws, row, val, is_subtotal=False, is_total=False, indent=0):
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value=("  " * indent) + row)
        cell = ws.cell(row=r, column=2, value=val)
        cell.number_format = num_fmt
        if is_total:
            ws.cell(row=r, column=1).font = total_font
            cell.font = total_font
            ws.cell(row=r, column=1).fill = total_fill
            cell.fill = total_fill
        elif is_subtotal:
            ws.cell(row=r, column=1).font = bold
            cell.font = bold
            ws.cell(row=r, column=1).fill = subtotal_fill
            cell.fill = subtotal_fill

    # === Sheet 1: Estado de Resultados ===
    ws_pl = wb.active
    ws_pl.title = "Estado de Resultados"
    ws_pl.column_dimensions['A'].width = 45
    ws_pl.column_dimensions['B'].width = 20

    write_header(ws_pl, "ESTADO DE RESULTADOS")
    pl = get_estado_resultados(anio, mes_inicio, mes_fin)

    # Header row
    r = ws_pl.max_row + 1
    ws_pl.cell(row=r, column=1, value="Concepto").font = header_font
    ws_pl.cell(row=r, column=1).fill = header_fill
    ws_pl.cell(row=r, column=2, value=periodo_str).font = header_font
    ws_pl.cell(row=r, column=2).fill = header_fill

    fmt_row(ws_pl, "Ingresos brutos", pl.get("ingresos_brutos", pl["ingresos"]["total"]))
    if pl.get("impoconsumo", 0) > 0:
        fmt_row(ws_pl, "(-) Impoconsumo", -pl["impoconsumo"], indent=1)
    if pl.get("devoluciones", 0) > 0:
        fmt_row(ws_pl, "(-) Devoluciones", -pl["devoluciones"], indent=1)
    fmt_row(ws_pl, "= INGRESOS NETOS", pl["ingresos"]["total"], is_subtotal=True)
    ws_pl.append([])
    for item in pl["costos_ventas"]["items"]:
        fmt_row(ws_pl, f"  {item['nombre']}", -item["saldo"], indent=1)
    fmt_row(ws_pl, "(-) TOTAL COSTO DE VENTAS", -pl["costos_ventas"]["total"], is_subtotal=True)
    ws_pl.append([])
    fmt_row(ws_pl, f"= UTILIDAD BRUTA ({pl['margen_bruto_pct']}%)", pl["utilidad_bruta"], is_subtotal=True)
    ws_pl.append([])
    for item in pl["gastos_admin"]["items"][:10]:
        fmt_row(ws_pl, f"  {item['nombre']}", -item["saldo"], indent=1)
    fmt_row(ws_pl, "(-) TOTAL GASTOS ADMINISTRACIÓN", -pl["gastos_admin"]["total"], is_subtotal=True)
    for item in pl["gastos_ventas"]["items"][:10]:
        fmt_row(ws_pl, f"  {item['nombre']}", -item["saldo"], indent=1)
    fmt_row(ws_pl, "(-) TOTAL GASTOS DE VENTAS", -pl["gastos_ventas"]["total"], is_subtotal=True)
    ws_pl.append([])
    fmt_row(ws_pl, f"= UTILIDAD OPERACIONAL ({pl['margen_operacional_pct']}%)", pl["utilidad_operacional"], is_subtotal=True)
    if pl["gastos_no_operacionales"]["total"] > 0:
        fmt_row(ws_pl, "(-) GASTOS NO OPERACIONALES", -pl["gastos_no_operacionales"]["total"])
    ws_pl.append([])
    fmt_row(ws_pl, f"= UTILIDAD NETA ({pl['margen_neto_pct']}%)", pl["utilidad_neta"], is_total=True)

    # === Sheet 2: Balance General ===
    ws_bg = wb.create_sheet("Balance General")
    ws_bg.column_dimensions['A'].width = 45
    ws_bg.column_dimensions['B'].width = 20

    write_header(ws_bg, "ESTADO DE SITUACIÓN FINANCIERA")
    bg = get_balance_general(anio, mes_fin)

    r = ws_bg.max_row + 1
    ws_bg.cell(row=r, column=1, value="Concepto").font = header_font
    ws_bg.cell(row=r, column=1).fill = header_fill
    ws_bg.cell(row=r, column=2, value=f"A {meses_nombres[mes_fin]} {anio}").font = header_font
    ws_bg.cell(row=r, column=2).fill = header_fill

    fmt_row(ws_bg, "ACTIVOS", None, is_subtotal=True)
    fmt_row(ws_bg, "Activo Corriente", None)
    for item in bg["activo_corriente"]["items"]:
        fmt_row(ws_bg, f"  {item['nombre']}", item["saldo"], indent=1)
    fmt_row(ws_bg, "Total Activo Corriente", bg["activo_corriente"]["total"], is_subtotal=True)
    ws_bg.append([])
    fmt_row(ws_bg, "Activo No Corriente", None)
    for item in bg["activo_no_corriente"]["items"]:
        fmt_row(ws_bg, f"  {item['nombre']}", item["saldo"], indent=1)
    fmt_row(ws_bg, "Total Activo No Corriente", bg["activo_no_corriente"]["total"], is_subtotal=True)
    fmt_row(ws_bg, "TOTAL ACTIVO", bg["total_activo"], is_total=True)
    ws_bg.append([])

    fmt_row(ws_bg, "PASIVOS", None, is_subtotal=True)
    fmt_row(ws_bg, "Pasivo Corriente", None)
    for item in bg["pasivo_corriente"]["items"]:
        fmt_row(ws_bg, f"  {item['nombre']}", item["saldo"], indent=1)
    fmt_row(ws_bg, "Total Pasivo Corriente", bg["pasivo_corriente"]["total"], is_subtotal=True)
    ws_bg.append([])
    fmt_row(ws_bg, "Pasivo No Corriente", None)
    for item in bg["pasivo_no_corriente"]["items"]:
        fmt_row(ws_bg, f"  {item['nombre']}", item["saldo"], indent=1)
    fmt_row(ws_bg, "Total Pasivo No Corriente", bg["pasivo_no_corriente"]["total"], is_subtotal=True)
    fmt_row(ws_bg, "TOTAL PASIVO", bg["total_pasivo"], is_subtotal=True)
    ws_bg.append([])

    fmt_row(ws_bg, "PATRIMONIO", None, is_subtotal=True)
    for item in bg["patrimonio"]["items"]:
        fmt_row(ws_bg, f"  {item['nombre']}", item["saldo"], indent=1)
    fmt_row(ws_bg, "TOTAL PATRIMONIO", bg["patrimonio"]["total"], is_subtotal=True)
    ws_bg.append([])
    fmt_row(ws_bg, "TOTAL PASIVO + PATRIMONIO", bg["total_pasivo_patrimonio"], is_total=True)

    # === Sheet 3: Indicadores ===
    ws_ind = wb.create_sheet("Indicadores")
    ws_ind.column_dimensions['A'].width = 35
    ws_ind.column_dimensions['B'].width = 20

    write_header(ws_ind, "INDICADORES FINANCIEROS")
    ind = get_indicadores(anio, mes_fin)

    r = ws_ind.max_row + 1
    ws_ind.cell(row=r, column=1, value="Indicador").font = header_font
    ws_ind.cell(row=r, column=1).fill = header_fill
    ws_ind.cell(row=r, column=2, value="Valor").font = header_font
    ws_ind.cell(row=r, column=2).fill = header_fill

    fmt_row(ws_ind, "LIQUIDEZ", None, is_subtotal=True)
    fmt_row(ws_ind, "Razón corriente", ind["ratios"]["liquidez"])
    fmt_row(ws_ind, "Capital de trabajo", ind["balance"]["activo_corriente"] - ind["balance"]["pasivo_corriente"])
    ws_ind.append([])
    fmt_row(ws_ind, "ENDEUDAMIENTO", None, is_subtotal=True)
    fmt_row(ws_ind, "Endeudamiento total (%)", ind["ratios"]["endeudamiento_pct"])
    ws_ind.append([])
    fmt_row(ws_ind, "RENTABILIDAD", None, is_subtotal=True)
    fmt_row(ws_ind, "Margen bruto (%)", ind["ratios"]["margen_bruto_pct"])
    fmt_row(ws_ind, "Margen operacional (%)", ind["ratios"]["margen_operacional_pct"])
    fmt_row(ws_ind, "Margen neto (%)", ind["ratios"]["margen_neto_pct"])
    fmt_row(ws_ind, "ROE (%)", ind["ratios"]["roe_pct"])
    fmt_row(ws_ind, "ROA (%)", ind["ratios"]["roa_pct"])

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
