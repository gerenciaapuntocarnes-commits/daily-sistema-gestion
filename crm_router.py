"""
CRM / Cartera module — Daily Sistema de Gestión
Fuentes: Siigo (clientes + facturas), Excel (datos contacto), Google Sheets (movimientos bancarios)
"""

import os
import json
import re
import io
import requests
from datetime import date, datetime, timedelta
from typing import Optional

from fastapi import APIRouter, HTTPException, UploadFile, File, BackgroundTasks
from pydantic import BaseModel

from database import get_conn
from siigo import get_token, _headers as siigo_headers, SIIGO_BASE, _paginate

router = APIRouter(prefix="/crm")

# In-memory status for long-running sync jobs (single Railway instance)
_sync_jobs: dict = {
    "siigo": {"running": False, "ok": None, "msg": "", "step": "", "result": None},
    "rc":    {"running": False, "ok": None, "msg": "", "step": "", "result": None},
}

# ═══════════════════════════════════════════════════════════════
# CONFIG helpers
# ═══════════════════════════════════════════════════════════════

def _get_config(key: str, default: str = "") -> str:
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT valor FROM config_planta WHERE parametro = %s", (key,))
        row = cur.fetchone()
        cur.close(); conn.close()
        return row[0] if row else default
    except Exception:
        return default

def _set_config(key: str, value: str, desc: str = ""):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO config_planta (parametro, valor, descripcion)
        VALUES (%s, %s, %s)
        ON CONFLICT (parametro) DO UPDATE SET valor = EXCLUDED.valor
    """, (key, value, desc))
    conn.commit()
    cur.close(); conn.close()

def _modo_prueba() -> bool:
    return _get_config("crm_modo_prueba", "true").lower() in ("true", "1", "yes")

# ═══════════════════════════════════════════════════════════════
# GOOGLE SHEETS helper
# ═══════════════════════════════════════════════════════════════

SHEET_ID = "1TPNSQWHHZbGJNrDwVbiSXmWRnQNQNRX97ciSWjfEXwU"

# Tabs: banco -> (tab_name, banco_code)
SHEET_TABS = [
    ("BDB2024",           "BDB"),
    ("BDB2025",           "BDB"),
    ("BANCOLOMBIA 2025",  "BANCOLOMBIA"),
    ("BANCOLOMBIA 2024",  "BANCOLOMBIA"),
    ("AVALPAY",           "AVALPAY"),
]

# Mapeo medio_pago → bancos a buscar en sugerencias
_MEDIO_A_BANCOS = {
    "BANCOLOMBIA":    ["BANCOLOMBIA"],
    "BANCO DE BOGOTA":["BDB"],
    "LINK":           ["AVALPAY", "BDB"],   # AVALPAY primero (identificación), BDB para conciliar
    "AVALPAY":        ["AVALPAY", "BDB"],
    "EFECTIVO":       [],
    "CRUCE":          [],
    "MARIA INES":     [],
}

def _get_sheets_service():
    """Build Google Sheets service from env var or local credentials file."""
    from google.oauth2 import service_account
    from googleapiclient.discovery import build

    creds_json = os.environ.get("GOOGLE_CREDENTIALS_JSON")
    if creds_json:
        info = json.loads(creds_json)
    else:
        # Fall back to local file (dev)
        paths = [
            os.path.join(os.path.dirname(__file__), "credentials.json"),
            "/Users/cav/Proyectos/trackmeat/credentials.json",
        ]
        for p in paths:
            if os.path.exists(p):
                with open(p) as f:
                    info = json.load(f)
                break
        else:
            raise RuntimeError("Google credentials not found. Set GOOGLE_CREDENTIALS_JSON env var.")

    scopes = ["https://www.googleapis.com/auth/spreadsheets"]  # lectura + escritura
    creds = service_account.Credentials.from_service_account_info(info, scopes=scopes)
    return build("sheets", "v4", credentials=creds, cache_discovery=False)


def _clean_valor_sheet(v, us_format: bool = False) -> Optional[float]:
    """
    Parsea valores del Sheet.
    us_format=False (BDB): punto=miles, coma=decimal  → "406.080,00" → 406080.0
    us_format=True (Bancolombia): coma=miles, punto=decimal → "179,000.00" → 179000.0
    """
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace("\xa0", "").strip()
    if not s:
        return None
    if us_format:
        # Formato americano: "179,000.00" → quitar comas, punto=decimal
        s = s.replace(",", "")
    else:
        if "," in s and "." in s:
            # Formato colombiano: "1.234.567,50" → quitar puntos, coma→punto
            s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            parts = s.split(",")
            if len(parts) == 2 and len(parts[1]) == 3 and parts[1].isdigit():
                s = s.replace(",", "")  # coma como miles
            else:
                s = s.replace(",", ".")  # coma como decimal
        elif "." in s:
            parts = s.split(".")
            if len(parts) > 2:
                s = s.replace(".", "")  # múltiples puntos = miles
            elif len(parts) == 2 and len(parts[1]) == 3 and parts[1].isdigit():
                s = s.replace(".", "")  # un punto como miles: "406.080"
    try:
        return float(s)
    except ValueError:
        return None


def _parse_date_sheet(v) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y%m%d", "%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y",
                "%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


_sheet_name_cache = {}  # tab_name_normalizado -> nombre exacto del Sheet
_sheet_name_cache_ts: float = 0.0  # timestamp de la última carga

def _get_exact_tab_name(service, tab_name: str) -> str:
    """Resuelve el nombre exacto de una pestaña normalizando espacios y mayúsculas. TTL 1h."""
    global _sheet_name_cache, _sheet_name_cache_ts
    import time
    if not _sheet_name_cache or (time.time() - _sheet_name_cache_ts) > 3600:
        try:
            meta = service.spreadsheets().get(spreadsheetId=SHEET_ID).execute()
            new_cache = {}
            for s in meta.get("sheets", []):
                title = s["properties"]["title"]
                key = title.strip().upper().replace("\xa0", " ")
                new_cache[key] = title
            _sheet_name_cache = new_cache
            _sheet_name_cache_ts = time.time()
        except Exception as e:
            print(f"Error obteniendo metadatos del Sheet: {e}")
    key = tab_name.strip().upper().replace("\xa0", " ")
    return _sheet_name_cache.get(key, tab_name)


def _fetch_sheet_tab(service, tab_name: str) -> list:
    """Fetch all rows from a sheet tab usando el nombre exacto del Sheet."""
    try:
        exact = _get_exact_tab_name(service, tab_name)
        safe = exact.replace("'", "\\'")
        result = service.spreadsheets().values().get(
            spreadsheetId=SHEET_ID,
            range=f"'{safe}'!A:J"
        ).execute()
        return result.get("values", [])
    except Exception as e:
        print(f"Error fetching tab {tab_name}: {e}")
        return []

# ═══════════════════════════════════════════════════════════════
# SIIGO CUSTOMERS — fetch all
# ═══════════════════════════════════════════════════════════════

def _fetch_siigo_customers() -> list:
    """Fetch all customers from Siigo, paginated."""
    all_results = []
    page = 1
    while True:
        resp = requests.get(
            f"{SIIGO_BASE}/customers",
            headers=siigo_headers(),
            params={"page": page, "page_size": 100},
            timeout=30
        )
        if resp.status_code == 429:
            import time; time.sleep(3)
            continue
        resp.raise_for_status()
        data = resp.json()
        results = data.get("results", [])
        all_results.extend(results)
        total = data.get("pagination", {}).get("total_results", 0)
        if len(all_results) >= total or not results:
            break
        page += 1
    return all_results


def _fetch_siigo_invoices_all() -> list:
    """Fetch all invoices from Siigo (last 3 years). Uses date chunking to avoid limits."""
    import time
    today = date.today()
    start = date(today.year - 3, 1, 1)
    all_invoices = []
    # Chunk by month to avoid hitting API limits
    cursor = start
    while cursor <= today:
        month_end = (cursor.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
        month_end = min(month_end, today)
        try:
            chunk = _paginate("/invoices", {
                "date_start": cursor.isoformat(),
                "date_end": month_end.isoformat(),
            }, max_pages=20)
            all_invoices.extend(chunk)
            time.sleep(0.2)
        except Exception as e:
            print(f"Error fetching invoices {cursor} - {month_end}: {e}")
        cursor = month_end + timedelta(days=1)
    return all_invoices

# ═══════════════════════════════════════════════════════════════
# ENDPOINTS — CONFIG
# ═══════════════════════════════════════════════════════════════

@router.get("/config")
def get_crm_config():
    modo_prueba = _modo_prueba()
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT tipo, fecha, registros, nuevos, actualizados FROM crm_sync_log ORDER BY fecha DESC LIMIT 1")
    last = cur.fetchone()
    cur.execute("SELECT COUNT(*) FROM crm_clientes")
    n_clientes = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM crm_facturas")
    n_facturas = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM movimientos_bancarios")
    n_movimientos = cur.fetchone()[0]
    cur.close(); conn.close()
    return {
        "modo_prueba": modo_prueba,
        "last_sync": {"tipo": last[0], "fecha": last[1].isoformat() if last else None, "registros": last[2]} if last else None,
        "stats": {"clientes": n_clientes, "facturas": n_facturas, "movimientos": n_movimientos}
    }


class ConfigIn(BaseModel):
    modo_prueba: bool

@router.put("/config")
def set_crm_config(data: ConfigIn):
    _set_config("crm_modo_prueba", str(data.modo_prueba).lower(), "Modo prueba CRM — no enviar a Siigo")
    return {"ok": True, "modo_prueba": data.modo_prueba}


# ═══════════════════════════════════════════════════════════════
# ENDPOINTS — RESET
# ═══════════════════════════════════════════════════════════════

@router.post("/reset")
def reset_crm():
    """Elimina todos los datos del CRM para empezar de cero."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM movimientos_bancarios")
    # Limpiar FK en ventas_daily antes de borrar crm_facturas
    cur.execute("UPDATE ventas_daily SET factura_id = NULL WHERE factura_id IS NOT NULL")
    cur.execute("DELETE FROM crm_facturas")
    cur.execute("DELETE FROM crm_clientes")
    cur.execute("DELETE FROM crm_sync_log")
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True, "mensaje": "Datos CRM eliminados. Listo para sincronizar desde Siigo."}


# Columnas por banco: {banco: (col_estado_idx, col_cliente_idx)} — 0-based
_SHEET_COLS = {
    "BDB":         {"estado": 4, "rc": 5, "cliente": 6},   # E, F, G
    "BANCOLOMBIA": {"estado": 3, "rc": 4, "cliente": 5},   # D, E, F
}

def _col_letter(idx: int) -> str:
    """Convierte índice 0-based a letra de columna: 0→A, 4→E, etc."""
    return chr(65 + idx)


@router.post("/movimientos/{mov_id}/actualizar-sheet")
def actualizar_movimiento_sheet(mov_id: int, rc_numero: Optional[str] = None):
    """Escribe CONCILIADO + nombre del cliente + RC en el Sheet para el movimiento dado."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT m.banco, m.sheet_tab, m.sheet_row, m.factura_id,
               c.nombre AS cliente_nombre, f.rc_numero
        FROM movimientos_bancarios m
        LEFT JOIN crm_facturas f ON f.id = m.factura_id
        LEFT JOIN crm_clientes c ON c.id = f.cliente_id
        WHERE m.id = %s
    """, (mov_id,))
    row = cur.fetchone()
    cur.close(); conn.close()

    if not row:
        raise HTTPException(404, "Movimiento no encontrado")

    banco, sheet_tab, sheet_row, factura_id, cliente_nombre, db_rc_numero = row
    rc_final = rc_numero or db_rc_numero

    if not sheet_tab or not sheet_row:
        raise HTTPException(400, "Movimiento sin referencia al Sheet")

    cols = _SHEET_COLS.get(banco, {"estado": 3, "rc": 4, "cliente": 5})
    col_estado  = _col_letter(cols["estado"])
    col_rc      = _col_letter(cols["rc"])
    col_cliente = _col_letter(cols["cliente"])

    try:
        service = _get_sheets_service()
        exact_tab = _get_exact_tab_name(service, sheet_tab)
        safe_tab  = exact_tab.replace("'", "\\'")

        # Una sola llamada batchUpdate para las 3 celdas
        batch_data = [{"range": f"'{safe_tab}'!{col_estado}{sheet_row}", "values": [["CONCILIADO"]]}]
        if rc_final:
            batch_data.append({"range": f"'{safe_tab}'!{col_rc}{sheet_row}", "values": [[rc_final]]})
        if cliente_nombre:
            batch_data.append({"range": f"'{safe_tab}'!{col_cliente}{sheet_row}", "values": [[cliente_nombre]]})
        service.spreadsheets().values().batchUpdate(
            spreadsheetId=SHEET_ID,
            body={"valueInputOption": "RAW", "data": batch_data}
        ).execute()

        return {"ok": True, "tab": exact_tab, "fila": sheet_row,
                "estado": "CONCILIADO", "rc": rc_final, "cliente": cliente_nombre}
    except Exception as e:
        raise HTTPException(500, f"Error actualizando Sheet: {str(e)}")


@router.post("/fix/conciliados-estado")
def fix_conciliados_estado():
    """Marca como conciliados los movimientos cuyo estado indica que ya están conciliados en el Sheet."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        UPDATE movimientos_bancarios
        SET conciliado = TRUE
        WHERE conciliado = FALSE
          AND (
            estado ILIKE '%CONCIL%' OR
            estado ILIKE '%MEDIO DE PAGO%' OR
            estado ILIKE '%QUEDO CON MEDIO%'
          )
    """)
    marcados = cur.rowcount
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True, "marcados": marcados}


class LimpiarSheetIn(BaseModel):
    tab: str
    clientes: list  # list of client name substrings to match

@router.post("/fix/limpiar-sheet")
def limpiar_sheet(data: LimpiarSheetIn):
    """Busca filas en un tab del Sheet por nombre de cliente y limpia ESTADO, RC, CLIENTE."""
    service = _get_sheets_service()
    exact_tab = _get_exact_tab_name(service, data.tab)
    safe_tab = exact_tab.replace("'", "\\'")

    banco = "BDB" if "BDB" in exact_tab.upper() else "BANCOLOMBIA"
    cols = _SHEET_COLS.get(banco, {"estado": 3, "rc": 4, "cliente": 5})
    col_cli = _col_letter(cols["cliente"])

    result = service.spreadsheets().values().get(
        spreadsheetId=SHEET_ID,
        range=f"'{safe_tab}'!{col_cli}1:{col_cli}10000"
    ).execute()
    rows = result.get("values", [])

    # Encontrar filas que matchean
    filas_limpiar = []
    for row_idx, row in enumerate(rows, start=1):
        cell = (row[0] if row else "").strip().lower()
        if not cell:
            continue
        for cli_name in data.clientes:
            if cli_name.lower() in cell:
                filas_limpiar.append({"fila": row_idx, "cliente": row[0].strip()})
                break

    # Batch update: una sola llamada para todas las celdas
    if filas_limpiar:
        batch_data = []
        for item in filas_limpiar:
            r = item["fila"]
            for col_key in ("estado", "rc", "cliente"):
                col_letter = _col_letter(cols[col_key])
                batch_data.append({
                    "range": f"'{safe_tab}'!{col_letter}{r}",
                    "values": [[""]]
                })
        service.spreadsheets().values().batchUpdate(
            spreadsheetId=SHEET_ID,
            body={"valueInputOption": "RAW", "data": batch_data}
        ).execute()

    return {"ok": True, "tab": exact_tab, "limpiados": filas_limpiar}


@router.post("/reset/bancos")
def reset_bancos():
    """Elimina solo los movimientos bancarios, conservando clientes y facturas."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM movimientos_bancarios")
    # También limpiar movimiento_id en facturas para no dejar referencias huérfanas
    cur.execute("UPDATE crm_facturas SET movimiento_id = NULL WHERE movimiento_id IS NOT NULL")
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}


# ═══════════════════════════════════════════════════════════════
# ENDPOINTS — SYNC
# ═══════════════════════════════════════════════════════════════

def _parse_customers(customers: list) -> list:
    rows = []
    for c in customers:
        sid = str(c.get("id", ""))
        if not sid:
            continue
        ident = str(c.get("identification", "")).strip() or None
        names = c.get("name", [])
        nombre = (" ".join(n for n in names if n) if isinstance(names, list) else str(names or "")).strip() or "(Sin nombre)"
        phones = c.get("phones", [])
        telefono = phones[0].get("number", "") if phones else None
        contacts = c.get("contacts", [])
        email = (contacts[0].get("email", "") if contacts else "") or None
        address_obj = c.get("address", {}) or {}
        direccion = address_obj.get("address", "") or None
        ciudad_obj = address_obj.get("city", {})
        ciudad = ciudad_obj.get("city_name", "") if ciudad_obj else None
        # Tipo de documento y persona para RC
        id_type_obj = c.get("id_type", {}) or {}
        id_type_code = str(id_type_obj.get("code", "13")) or "13"
        person_type = c.get("person_type", "Person") or "Person"
        rows.append((sid, ident, nombre, telefono, email, direccion, ciudad, id_type_code, person_type))
    return rows


def _parse_invoices(invoices: list) -> list:
    rows = []
    for inv in invoices:
        if inv.get("annulled"):
            continue
        siigo_id = str(inv.get("id", ""))
        if not siigo_id:
            continue
        prefix = inv.get("prefix", "") or ""
        number = inv.get("number", 0) or 0
        # Siigo name contiene el prefijo contable: "FV-2-4892"
        # Si prefix es el electrónico (FE) o está vacío, lo reemplazamos con el contable
        inv_name = inv.get("name", "") or ""
        if inv_name and "-" in inv_name:
            parts = inv_name.rsplit("-", 1)
            if len(parts) == 2 and parts[1].isdigit() and len(parts[1]) >= 3:
                prefix = parts[0]          # "FV-2"
                number = int(parts[1])     # 4892
        fecha_str = (inv.get("date") or "")[:10]
        try:
            fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date() if fecha_str else date.today()
        except ValueError:
            fecha = date.today()
        total = float(inv.get("total", 0) or 0)
        balance = float(inv.get("balance", 0) or 0)
        estado = "pendiente" if balance > 0 else "pagado"
        cust = inv.get("customer", {}) or {}
        cedula = str(cust.get("identification", "") or "").strip() or None
        cust_names = cust.get("name", [])
        cli_nombre = (" ".join(n for n in cust_names if n) if isinstance(cust_names, list) else str(cust_names or "")).strip() or "(Sin nombre)"
        siigo_cust_id = str(cust.get("id", "") or "") or None
        # Extraer medio de pago desde Siigo (payments[0].name si existe)
        payments = inv.get("payments") or []
        medio_pago_siigo = payments[0].get("name", "").strip() if payments else None
        rows.append((siigo_id, number, prefix, cli_nombre, cedula, fecha, total, balance, estado, siigo_cust_id, medio_pago_siigo))
    return rows


def _run_sync_siigo():
    """Tarea en background: sincroniza clientes y facturas de Siigo usando upsert batch."""
    import time
    job = _sync_jobs["siigo"]
    job["running"] = True
    job["ok"] = None
    job["msg"] = ""
    job["result"] = None
    job["started_at"] = time.time()
    try:
        _run_sync_siigo_inner(job)
    except Exception as e:
        job["running"] = False
        job["ok"] = False
        job["msg"] = f"Error inesperado: {str(e)}"
        job["step"] = "Error"


def _run_sync_siigo_inner(job):

    try:
        job["step"] = "Obteniendo clientes de Siigo..."
        customers = _fetch_siigo_customers()
    except Exception as e:
        job["running"] = False; job["ok"] = False
        job["msg"] = f"Error obteniendo clientes: {str(e)}"
        return

    job["step"] = f"Guardando {len(customers)} clientes..."
    cli_rows = _parse_customers(customers)

    conn = get_conn()
    cur = conn.cursor()

    # Batch upsert clientes — un solo query
    if cli_rows:
        from psycopg2.extras import execute_values
        execute_values(cur, """
            INSERT INTO crm_clientes (siigo_id, cedula, nombre, telefono, email, direccion, ciudad, id_type_code, person_type)
            VALUES %s
            ON CONFLICT (siigo_id) DO UPDATE SET
              cedula = EXCLUDED.cedula,
              nombre = EXCLUDED.nombre,
              telefono = EXCLUDED.telefono,
              email = EXCLUDED.email,
              direccion = EXCLUDED.direccion,
              ciudad = EXCLUDED.ciudad,
              id_type_code = EXCLUDED.id_type_code,
              person_type = EXCLUDED.person_type,
              actualizado_en = NOW()
        """, cli_rows)
        conn.commit()

    # Count nuevos vs actualizados
    cur.execute("SELECT COUNT(*) FROM crm_clientes")
    cli_total = cur.fetchone()[0]
    cli_new = max(0, cli_total - max(0, cli_total - len(cli_rows)))

    try:
        job["step"] = "Obteniendo facturas de Siigo (últimos 3 años)..."
        invoices = _fetch_siigo_invoices_all()
    except Exception as e:
        conn.close()
        job["running"] = False; job["ok"] = False
        job["msg"] = f"Error obteniendo facturas: {str(e)}"
        return

    job["step"] = f"Guardando {len(invoices)} facturas..."
    inv_rows = _parse_invoices(invoices)

    # Build siigo_id -> internal id map for cliente lookup
    cur.execute("SELECT siigo_id, id FROM crm_clientes")
    cust_map = {row[0]: row[1] for row in cur.fetchall()}

    # Resolve cliente_id for each invoice
    final_rows = []
    for r in inv_rows:
        siigo_id, number, prefix, cli_nombre, cedula, fecha, total, balance, estado, siigo_cust_id, medio_pago_siigo = r
        cliente_id = cust_map.get(siigo_cust_id) if siigo_cust_id else None
        final_rows.append((siigo_id, number, prefix, cliente_id, cli_nombre, cedula, fecha, total, balance, estado, medio_pago_siigo))

    if final_rows:
        execute_values(cur, """
            INSERT INTO crm_facturas
              (siigo_invoice_id, numero, prefix, cliente_id, cliente_nombre, cliente_cedula,
               fecha, total, balance, estado_pago, medio_pago)
            VALUES %s
            ON CONFLICT (siigo_invoice_id) DO UPDATE SET
              total = EXCLUDED.total,
              balance = EXCLUDED.balance,
              prefix = EXCLUDED.prefix,
              numero = EXCLUDED.numero,
              cliente_id = EXCLUDED.cliente_id,
              cliente_nombre = EXCLUDED.cliente_nombre,
              cliente_cedula = EXCLUDED.cliente_cedula,
              -- Si Siigo dice balance=0 → pagado (Siigo es fuente de verdad)
              -- Si Siigo dice balance>0 → solo actualizar si era pendiente (no revertir pagos manuales)
              estado_pago = CASE
                WHEN EXCLUDED.balance = 0 THEN 'pagado'
                ELSE crm_facturas.estado_pago
              END,
              -- Solo poner medio_pago de Siigo si no está ya definido manualmente
              medio_pago = COALESCE(crm_facturas.medio_pago, EXCLUDED.medio_pago),
              sync_at = NOW()
        """, final_rows)
        conn.commit()

    cur.execute("SELECT COUNT(*) FROM crm_facturas")
    inv_total = cur.fetchone()[0]

    # ── Notas de Crédito: marcar facturas con NC como pagadas ──────────
    # Las NC cancelan la deuda — no deben aparecer en conciliación pendiente
    nc_marcadas = 0
    try:
        job["step"] = "Procesando Notas de Crédito..."
        from siigo import fetch_credit_notes
        ncs = fetch_credit_notes()
        nc_invoice_ids = list({
            nc["invoice"]["id"]
            for nc in ncs
            if nc.get("invoice") and nc["invoice"].get("id")
        })
        if nc_invoice_ids:
            cur.execute("""
                UPDATE crm_facturas
                SET balance = 0, estado_pago = 'pagado', tiene_nc = TRUE
                WHERE siigo_invoice_id = ANY(%s)
            """, (nc_invoice_ids,))
            nc_marcadas = cur.rowcount
            conn.commit()
    except Exception as e:
        job["step"] = f"NC: error no crítico ({e})"

    cur.execute("""
        INSERT INTO crm_sync_log (tipo, registros, nuevos, actualizados, detalle)
        VALUES ('siigo_full', %s, %s, %s, %s)
    """, (len(cli_rows) + len(final_rows), len(cli_rows), len(final_rows),
          f"Clientes: {len(cli_rows)} | Facturas: {len(final_rows)} | NC marcadas: {nc_marcadas}"))
    conn.commit()
    cur.close(); conn.close()

    job["running"] = False
    job["ok"] = True
    job["msg"] = f"{len(cli_rows)} clientes, {len(final_rows)} facturas, {nc_marcadas} NC aplicadas."
    job["step"] = "Completado"
    job["result"] = {
        "ok": True,
        "clientes": {"total": len(cli_rows)},
        "facturas": {"total": len(final_rows)},
        "nc_marcadas": nc_marcadas
    }


@router.post("/sync/nc-siigo")
def sync_nc_siigo():
    """Aplica todas las NC de Siigo: marca facturas con NC como pagadas."""
    from siigo import fetch_credit_notes
    ncs = fetch_credit_notes()
    nc_invoice_ids = list({
        nc["invoice"]["id"]
        for nc in ncs
        if nc.get("invoice") and nc["invoice"].get("id")
    })
    if not nc_invoice_ids:
        return {"ok": True, "nc_total": 0, "marcadas": 0}
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        UPDATE crm_facturas
        SET balance = 0, estado_pago = 'pagado', tiene_nc = TRUE
        WHERE siigo_invoice_id = ANY(%s)
    """, (nc_invoice_ids,))
    marcadas = cur.rowcount
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True, "nc_total": len(nc_invoice_ids), "marcadas": marcadas}


@router.post("/sync/siigo")
def sync_siigo(background_tasks: BackgroundTasks):
    """Inicia sync de Siigo en background y retorna inmediatamente."""
    job = _sync_jobs["siigo"]
    if job["running"]:
        return {"started": False, "running": True, "msg": "Sync ya en curso"}
    background_tasks.add_task(_run_sync_siigo)
    return {"started": True, "running": True, "msg": "Sync iniciado"}


@router.get("/sync/siigo/status")
def sync_siigo_status():
    """Retorna el estado actual del sync de Siigo."""
    import time
    job = _sync_jobs["siigo"]
    # Auto-reset si lleva más de 30 min corriendo (Siigo caído o proceso muerto)
    started = job.get("started_at", 0)
    if job["running"] and started and (time.time() - started) > 1800:
        _sync_jobs["siigo"] = {"running": False, "ok": False, "msg": "Timeout: sync superó 30 min, reinicia manualmente.", "step": "Timeout", "result": None, "started_at": 0}
    result = dict(_sync_jobs["siigo"])
    # Limpiar result tras entregarlo (evita acumular en memoria)
    if result.get("result") and not result.get("running"):
        _sync_jobs["siigo"]["result"] = None
    return result


@router.post("/sync/siigo/reset")
def sync_siigo_reset():
    """Resetea el estado del job si quedó bloqueado."""
    _sync_jobs["siigo"] = {"running": False, "ok": None, "msg": "", "step": "", "result": None, "started_at": 0}
    return {"ok": True}


@router.post("/sync/excel")
async def sync_excel(file: UploadFile = File(...)):
    """Importa BASE DATOS y CLIENTES SIN COMPRA del Excel para enriquecer contactos."""
    import openpyxl

    def _cs(v):
        if v is None: return None
        s = str(v).strip(); return s or None

    def _cp(v):
        if v is None: return None
        s = re.sub(r"\.0$", "", str(v).strip())
        s = re.sub(r"[^0-9+]", "", s)
        return s or None

    def _cc(v):
        if v is None: return None
        s = re.sub(r"\.0$", "", str(v).strip())
        s = re.sub(r"[^0-9a-zA-Z-]", "", s)
        return s or None

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    conn = get_conn()
    cur = conn.cursor()
    ins = upd = 0

    # ── BASE DATOS ──
    for sheet_name in [' BASE DATOS', 'BASE DATOS']:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                nombre = _cs(row[0])
                if not nombre: continue
                direccion = _cs(row[1])
                apto      = _cs(row[2])
                zona      = _cs(row[4])
                telefono  = _cp(row[5])
                cedula    = _cc(row[6])
                email     = _cs(row[7])

                # Try match by cedula first
                if cedula:
                    cur.execute("SELECT id FROM crm_clientes WHERE cedula = %s LIMIT 1", (cedula,))
                    ex = cur.fetchone()
                    if ex:
                        cur.execute("""
                            UPDATE crm_clientes SET
                              direccion=COALESCE(crm_clientes.direccion,%s),
                              telefono=COALESCE(crm_clientes.telefono,%s),
                              email=COALESCE(crm_clientes.email,%s),
                              actualizado_en=NOW()
                            WHERE id=%s
                        """, (direccion, telefono, email, ex[0]))
                        upd += 1; continue

                # Try match by name
                cur.execute("SELECT id FROM crm_clientes WHERE nombre ILIKE %s LIMIT 1", (nombre,))
                ex = cur.fetchone()
                if ex:
                    cur.execute("""
                        UPDATE crm_clientes SET
                          direccion=COALESCE(crm_clientes.direccion,%s),
                          telefono=COALESCE(crm_clientes.telefono,%s),
                          cedula=COALESCE(crm_clientes.cedula,%s),
                          email=COALESCE(crm_clientes.email,%s),
                          actualizado_en=NOW()
                        WHERE id=%s
                    """, (direccion, telefono, cedula, email, ex[0]))
                    upd += 1; continue

                # New client from Excel
                cur.execute("""
                    INSERT INTO crm_clientes (nombre, direccion, telefono, cedula, email)
                    VALUES (%s,%s,%s,%s,%s)
                """, (nombre, direccion, telefono, cedula, email))
                ins += 1
            conn.commit()
            break

    # ── CLIENTES SIN COMPRA ──
    for sheet_name in ['CLIENTES SIN COMPRA ', 'CLIENTES SIN COMPRA']:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                nombre = _cs(row[0])
                if not nombre: continue
                cur.execute("SELECT id FROM crm_clientes WHERE nombre ILIKE %s LIMIT 1", (nombre,))
                if cur.fetchone(): continue
                direccion = _cs(row[1])
                apto      = _cs(row[2])
                zona      = _cs(row[4]) if len(row) > 4 else None
                telefono  = _cp(row[5]) if len(row) > 5 else None
                cur.execute("""
                    INSERT INTO crm_clientes (nombre, direccion, telefono)
                    VALUES (%s,%s,%s)
                """, (nombre, direccion, telefono))
                ins += 1
            conn.commit()
            break

    cur.execute("""
        INSERT INTO crm_sync_log (tipo, registros, nuevos, actualizados, detalle)
        VALUES ('excel', %s, %s, %s, %s)
    """, (ins + upd, ins, upd, "Importación Excel BASE DATOS + CLIENTES SIN COMPRA"))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True, "nuevos": ins, "actualizados": upd}


@router.post("/sync/bancos")
def sync_bancos():
    """Sincroniza movimientos bancarios desde Google Sheets."""
    try:
        service = _get_sheets_service()
    except Exception as e:
        raise HTTPException(500, f"Error conectando Google Sheets: {str(e)}")

    conn = get_conn()
    cur = conn.cursor()

    # Asegurar índice único por posición en sheet para upsert
    cur.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_mov_sheet_pos
        ON movimientos_bancarios (sheet_tab, sheet_row)
    """)
    conn.commit()

    ins = upd = skip = 0
    tabs_resumen = []

    for tab_name, banco in SHEET_TABS:
        rows = _fetch_sheet_tab(service, tab_name)
        tab_ins = tab_upd = tab_skip = 0
        if not rows or len(rows) < 2:
            tabs_resumen.append({"tab": tab_name, "banco": banco, "filas": 0, "error": "Sin datos o pestaña no encontrada"})
            continue

        is_bdb = banco == "BDB"
        is_avalpay = banco == "AVALPAY"

        for row_idx, row in enumerate(rows[1:], start=2):
            row = list(row) + [""] * 15

            if is_avalpay:
                # AVALPAY PSE: id|date|franchise|card_num|currency_code|amount|tax|invoice_num|reference_alt
                # amount está en centavos, date = "DD/MM/YYYY HH:MM"
                fecha_raw = row[1]   # col B: date
                amount_raw = row[5]  # col F: amount (centavos)
                invoice_num   = str(row[7]).strip() if row[7] else ""
                reference_alt = str(row[8]).strip() if row[8] else ""
                shopper_name  = str(row[14]).strip() if len(row) > 14 and row[14] else ""
                shopper_email = str(row[15]).strip() if len(row) > 15 and row[15] else ""
                # Descripción: número de orden Shopify
                if invoice_num and invoice_num.lower() not in ("shopify payment", ""):
                    desc_raw = invoice_num
                elif reference_alt and reference_alt.lower() != "shopify payment":
                    desc_raw = reference_alt
                else:
                    desc_raw = "AVALPAY PSE"
                # Comprador: nombre real del comprador + email
                if shopper_name:
                    buyer_id = shopper_name + (f" <{shopper_email}>" if shopper_email else "")
                elif shopper_email:
                    buyer_id = shopper_email
                elif reference_alt and reference_alt.lower() != "shopify payment":
                    buyer_id = reference_alt
                else:
                    buyer_id = None
                try:
                    valor = float(str(amount_raw).replace(",", "").replace(" ", "")) / 100
                except:
                    valor = None
                estado = None; rc_sheet = None; cli_sheet = buyer_id
            elif is_bdb:
                # BDB: formato colombiano, columnas: Fecha|Transaccion|Debitos|Creditos|Estado|RC|Cliente
                fecha_raw = row[0]
                desc_raw  = row[1]
                debito    = _clean_valor_sheet(row[2], us_format=False)
                credito   = _clean_valor_sheet(row[3], us_format=False)
                estado    = str(row[4]).strip() if row[4] else None
                rc_sheet  = str(row[5]).strip() if row[5] else None
                cli_sheet = str(row[6]).strip() if row[6] else None
                valor = credito if credito and credito > 0 else (-(debito) if debito and debito > 0 else None)
            else:
                # Bancolombia: formato americano (coma=miles, punto=decimal), fecha YYYYMMDD
                # Columnas: Fecha|Concepto|Valor Banco|Estado|RC|Cliente
                fecha_raw = row[0]
                desc_raw  = row[1]
                valor_raw = _clean_valor_sheet(row[2], us_format=True)
                estado    = str(row[3]).strip() if row[3] else None
                rc_sheet  = str(row[4]).strip() if row[4] else None
                cli_sheet = str(row[5]).strip() if row[5] else None
                valor = valor_raw

            fecha = _parse_date_sheet(fecha_raw)
            desc  = str(desc_raw).strip() if desc_raw else None

            if not fecha or valor is None:
                skip += 1; tab_skip += 1
                continue

            if valor <= 0:
                skip += 1; tab_skip += 1
                continue

            # Auto-detect conciliado por estado del sheet
            conciliado_sheet = bool(estado and (
                "CONCIL" in estado.upper() or
                "MEDIO DE PAGO" in estado.upper() or
                "QUEDO CON MEDIO" in estado.upper()
            ))

            # Upsert por (sheet_tab, sheet_row) — siempre refleja el estado actual del sheet
            # No pisar conciliado=TRUE si ya fue conciliado manualmente en la app
            cur.execute("""
                INSERT INTO movimientos_bancarios
                  (banco, fecha, descripcion, valor, estado, rc_sheet, cliente_sheet,
                   conciliado, sheet_row, sheet_tab)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (sheet_tab, sheet_row) DO UPDATE SET
                  fecha        = EXCLUDED.fecha,
                  descripcion  = EXCLUDED.descripcion,
                  valor        = EXCLUDED.valor,
                  estado       = EXCLUDED.estado,
                  rc_sheet     = EXCLUDED.rc_sheet,
                  cliente_sheet= EXCLUDED.cliente_sheet,
                  conciliado   = CASE
                    WHEN movimientos_bancarios.conciliado = TRUE THEN TRUE
                    ELSE EXCLUDED.conciliado
                  END
            """, (banco, fecha, desc, valor, estado, rc_sheet or None, cli_sheet or None,
                  conciliado_sheet, row_idx, tab_name))

            if cur.statusmessage == "INSERT 0 1":
                ins += 1; tab_ins += 1
            else:
                upd += 1; tab_upd += 1

        tabs_resumen.append({"tab": tab_name, "banco": banco,
                              "filas_leidas": len(rows) - 1,
                              "nuevos": tab_ins, "actualizados": tab_upd, "ignorados": tab_skip})

    # Marcar como conciliados los registros con estado que indica conciliación en el Sheet
    cur.execute("""
        UPDATE movimientos_bancarios
        SET conciliado = TRUE
        WHERE conciliado = FALSE
          AND (
            estado ILIKE '%CONCIL%' OR
            estado ILIKE '%MEDIO DE PAGO%' OR
            estado ILIKE '%QUEDO CON MEDIO%'
          )
    """)
    marcados = cur.rowcount

    conn.commit()
    cur.execute("""
        INSERT INTO crm_sync_log (tipo, registros, nuevos, actualizados, detalle)
        VALUES ('bancos', %s, %s, %s, %s)
    """, (ins + upd + marcados, ins, upd + marcados,
          f"Movimientos bancarios: {ins} nuevos, {upd} actualizados, {skip} ignorados, {marcados} marcados conciliados por estado"))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True, "nuevos": ins, "actualizados": upd, "ignorados": skip,
            "marcados_conciliados": marcados, "por_pestaña": tabs_resumen}


# ═══════════════════════════════════════════════════════════════
# ENDPOINTS — FACTURAS (Activos / Históricos)
# ═══════════════════════════════════════════════════════════════

@router.get("/activos")
def get_activos(q: Optional[str] = None):
    """Clientes con al menos una factura en los últimos 6 meses, con sus facturas."""
    corte = (date.today() - timedelta(days=180)).isoformat()
    return _get_clientes_con_facturas(corte_desde=corte, q=q)


@router.get("/historicos")
def get_historicos(q: Optional[str] = None):
    """Clientes con facturas más antiguas de 6 meses (sin facturas recientes)."""
    corte = (date.today() - timedelta(days=180)).isoformat()
    return _get_clientes_con_facturas(corte_hasta=corte, q=q)


def _get_clientes_con_facturas(corte_desde: str = None, corte_hasta: str = None, q: str = None) -> list:
    conn = get_conn()
    cur = conn.cursor()

    # Get client IDs that match date criteria
    if corte_desde:
        cur.execute("""
            SELECT DISTINCT cliente_id FROM crm_facturas
            WHERE fecha >= %s AND cliente_id IS NOT NULL
        """, (corte_desde,))
    elif corte_hasta:
        cur.execute("""
            SELECT DISTINCT cliente_id FROM crm_facturas f
            WHERE cliente_id IS NOT NULL
            AND NOT EXISTS (
                SELECT 1 FROM crm_facturas f2
                WHERE f2.cliente_id = f.cliente_id AND f2.fecha >= %s
            )
        """, (corte_hasta,))
    else:
        cur.execute("SELECT DISTINCT cliente_id FROM crm_facturas WHERE cliente_id IS NOT NULL")

    ids = [r[0] for r in cur.fetchall()]
    if not ids:
        cur.close(); conn.close()
        return []

    # Apply name filter
    q_filter = ""
    params = [list(ids)]
    if q:
        q_filter = "AND c.nombre ILIKE %s"
        params.append(f"%{q}%")

    cur.execute(f"""
        SELECT c.id, c.siigo_id, c.cedula, c.nombre, c.telefono, c.email,
               c.direccion, c.ciudad, c.origen_canal, c.notas,
               COUNT(f.id) AS num_facturas,
               SUM(f.total) AS total_ventas,
               SUM(f.balance) AS total_pendiente,
               MAX(f.fecha) AS ultima_factura
        FROM crm_clientes c
        JOIN crm_facturas f ON f.cliente_id = c.id
        WHERE c.id = ANY(%s) {q_filter}
        GROUP BY c.id
        ORDER BY ultima_factura DESC
    """, params)

    clients_rows = cur.fetchall()
    if not clients_rows:
        cur.close(); conn.close()
        return []

    client_ids = [row[0] for row in clients_rows]
    cur.execute("""
        SELECT id, siigo_invoice_id, numero, prefix, fecha, total, balance,
               estado_pago, medio_pago, cuenta_debito, origen_canal,
               movimiento_id, rc_siigo_id, rc_numero, rc_modo_prueba, cliente_id
        FROM crm_facturas
        WHERE cliente_id = ANY(%s)
        ORDER BY cliente_id, fecha DESC
    """, (client_ids,))
    facturas_by_cliente = {}
    for f in cur.fetchall():
        facturas_by_cliente.setdefault(f[15], []).append({
            "id": f[0], "siigo_invoice_id": f[1],
            "numero": f[2], "prefix": f[3],
            "factura": f"{f[3]}-{f[2]}" if f[3] else str(f[2]),
            "fecha": f[4].isoformat() if f[4] else None,
            "total": float(f[5] or 0), "balance": float(f[6] or 0),
            "estado_pago": f[7], "medio_pago": f[8],
            "cuenta_debito": f[9], "origen_canal": f[10],
            "movimiento_id": f[11], "rc_siigo_id": f[12],
            "rc_numero": f[13], "rc_modo_prueba": f[14],
        })

    clients = []
    for row in clients_rows:
        clients.append({
            "id": row[0], "siigo_id": row[1], "cedula": row[2], "nombre": row[3],
            "telefono": row[4], "email": row[5], "direccion": row[6], "ciudad": row[7],
            "origen_canal": row[8], "notas": row[9],
            "num_facturas": row[10], "total_ventas": float(row[11] or 0),
            "total_pendiente": float(row[12] or 0),
            "ultima_factura": row[13].isoformat() if row[13] else None,
            "facturas": facturas_by_cliente.get(row[0], []),
        })

    cur.close(); conn.close()
    return clients


@router.get("/clientes/{cliente_id}")
def get_cliente(cliente_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, siigo_id, cedula, nombre, telefono, email, direccion, ciudad, origen_canal, notas
        FROM crm_clientes WHERE id = %s
    """, (cliente_id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(404, "Cliente no encontrado")

    cur.execute("""
        SELECT id, siigo_invoice_id, numero, prefix, fecha, total, balance,
               estado_pago, medio_pago, cuenta_debito, origen_canal,
               movimiento_id, rc_siigo_id, rc_numero, rc_modo_prueba
        FROM crm_facturas WHERE cliente_id = %s ORDER BY fecha DESC
    """, (cliente_id,))
    facturas = []
    for f in cur.fetchall():
        facturas.append({
            "id": f[0], "siigo_invoice_id": f[1],
            "numero": f[2], "prefix": f[3],
            "factura": f"{f[3]}-{f[2]}" if f[3] else str(f[2]),
            "fecha": f[4].isoformat() if f[4] else None,
            "total": float(f[5] or 0), "balance": float(f[6] or 0),
            "estado_pago": f[7], "medio_pago": f[8],
            "cuenta_debito": f[9], "origen_canal": f[10],
            "movimiento_id": f[11], "rc_siigo_id": f[12],
            "rc_numero": f[13], "rc_modo_prueba": f[14],
        })
    cur.close(); conn.close()
    return {
        "id": row[0], "siigo_id": row[1], "cedula": row[2], "nombre": row[3],
        "telefono": row[4], "email": row[5], "direccion": row[6], "ciudad": row[7],
        "origen_canal": row[8], "notas": row[9],
        "facturas": facturas,
    }


class FacturaUpdate(BaseModel):
    estado_pago: Optional[str] = None
    medio_pago: Optional[str] = None
    cuenta_debito: Optional[str] = None
    origen_canal: Optional[str] = None

@router.put("/facturas/{factura_id}")
def update_factura(factura_id: int, data: FacturaUpdate):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id FROM crm_facturas WHERE id = %s", (factura_id,))
    if not cur.fetchone():
        cur.close(); conn.close()
        raise HTTPException(404, "Factura no encontrada")

    sets = []
    vals = []
    if data.estado_pago is not None:
        sets.append("estado_pago=%s"); vals.append(data.estado_pago)
    if data.medio_pago is not None:
        sets.append("medio_pago=%s"); vals.append(data.medio_pago)
    if data.cuenta_debito is not None:
        sets.append("cuenta_debito=%s"); vals.append(data.cuenta_debito)
    if data.origen_canal is not None:
        sets.append("origen_canal=%s"); vals.append(data.origen_canal)

    if sets:
        vals.append(factura_id)
        cur.execute(f"UPDATE crm_facturas SET {', '.join(sets)} WHERE id=%s", vals)
        conn.commit()

    cur.close(); conn.close()
    return {"ok": True}


# ═══════════════════════════════════════════════════════════════
# ENDPOINTS — RECIBO DE CAJA (RC)
# ═══════════════════════════════════════════════════════════════

# Debit account codes for RC in Siigo
CUENTAS_RC = {
    "Efectivo":          "11050501",
    "Banco de Bogota":   "11100501",
    "BDB":               "11100501",
    "Bancolombia":       "11200502",
    "Link":              "11100501",  # Link de pago cae en BDB
}

# Descripción de cuenta para el débito del RC (tal como aparece en Siigo)
_DESC_CUENTA = {
    "11050501": "Caja",
    "11100501": "Banco de Bogotá",
    "11200502": "Bancolombia Cta. de Ahorros",
}

def _cuenta_for_medio(medio: str) -> str:
    if not medio:
        return "11050501"
    m = medio.strip()
    for k, v in CUENTAS_RC.items():
        if k.upper() in m.upper():
            return v
    return "11050501"


def _crear_rc_en_siigo(factura: dict, modo_prueba: bool) -> dict:
    """
    Crea un Recibo de Caja en Siigo para la factura dada.
    Si modo_prueba=True, simula y retorna respuesta falsa.
    Retorna dict con {ok, rc_id, rc_numero, simulado}
    """
    if modo_prueba:
        return {
            "ok": True,
            "simulado": True,
            "rc_id": f"PRUEBA-{factura['id']}",
            "rc_numero": f"RC-PRUEBA-{factura['id']}",
            "mensaje": "Modo prueba: RC no creado en Siigo"
        }

    # ── Validaciones ──
    if not factura.get("siigo_invoice_id"):
        return {"ok": False, "error": "Factura sin ID de Siigo — sincroniza primero con Siigo"}

    # Datos del cliente + valor del movimiento bancario vinculado
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT c.siigo_id, c.cedula, c.nombre
        FROM crm_clientes c
        JOIN crm_facturas f ON f.cliente_id = c.id
        WHERE f.id = %s
    """, (factura["id"],))
    cli = cur.fetchone()
    mov_valor = None
    if factura.get("movimiento_id"):
        cur.execute("SELECT valor FROM movimientos_bancarios WHERE id = %s", (factura["movimiento_id"],))
        mov_row = cur.fetchone()
        if mov_row:
            mov_valor = round(float(mov_row[0]), 2)
    cur.close(); conn.close()

    if not cli:
        return {"ok": False, "error": "Cliente no encontrado para la factura"}

    siigo_cust_id = cli[0] or ""
    cedula = cli[1] or ""
    if not siigo_cust_id and not cedula:
        return {"ok": False, "error": "Cliente sin ID de Siigo ni cédula — no se puede crear RC"}

    # ── Consultar datos exactos del cliente en Siigo ──
    branch_office = 0
    if siigo_cust_id:
        try:
            cust_resp = requests.get(
                f"{SIIGO_BASE}/customers/{siigo_cust_id}",
                headers=siigo_headers(), timeout=15
            )
            if cust_resp.status_code == 200:
                cust_data = cust_resp.json()
                cedula = str(cust_data.get("identification", cedula) or cedula)
                branch_office = cust_data.get("branch_office", 0) or 0
        except Exception:
            pass  # usar datos locales si falla la consulta

    balance_val = float(factura.get("balance") or 0)
    total_val   = float(factura.get("total")   or 0)
    if balance_val <= 0 and total_val <= 0:
        return {"ok": False, "error": "Factura sin monto válido (balance y total son 0 o nulos)"}
    monto = round(balance_val if balance_val > 0 else total_val, 2)

    # Monto que llegó al banco (pesos redondos). Si no hay movimiento vinculado, igual al balance.
    monto_banco = mov_valor if mov_valor is not None else monto
    # Diferencia de centavos entre lo que dice la factura y lo que entró al banco
    ajuste = round(monto - monto_banco, 2)

    cuenta = factura.get("cuenta_debito") or _cuenta_for_medio(factura.get("medio_pago", ""))
    prefix  = factura.get("prefix", "") or ""
    numero  = factura.get("numero", "") or ""

    # ── Consultar factura en Siigo para obtener prefijo contable real ──
    # El campo 'prefix' en BD es el prefijo electrónico (FE), pero Siigo usa
    # el prefijo contable (ej: FV-2) para vincular el RC a la factura.
    siigo_inv_id = factura.get("siigo_invoice_id")
    due_prefix = prefix
    due_consecutive = int(numero) if str(numero).isdigit() else 0
    if siigo_inv_id:
        try:
            inv_resp = requests.get(
                f"{SIIGO_BASE}/invoices/{siigo_inv_id}",
                headers=siigo_headers(), timeout=15
            )
            if inv_resp.status_code == 200:
                inv_data = inv_resp.json()
                inv_name = inv_data.get("name", "") or ""
                inv_number = inv_data.get("number", numero)
                # name = "FV-2-4881" → prefix contable = "FV-2", consecutive = 4881
                num_str = str(inv_number)
                if inv_name.endswith(f"-{num_str}"):
                    due_prefix = inv_name[:-(len(num_str) + 1)]
                    due_consecutive = int(inv_number)
        except Exception:
            pass  # usar datos locales si falla

    # Usar prefijo contable (due_prefix) para todas las referencias visibles en el RC
    factura_ref = f"{due_prefix}-{due_consecutive}" if due_prefix else str(numero)

    # ── Payload según documentación Siigo API (Detailed voucher) ──
    payload = {
        "document": {"id": 3619},
        "type": "Detailed",
        "date": date.today().isoformat(),
        "customer": {
            "identification": cedula,
            "branch_office": branch_office
        },
        "observations": f"Generado por Daily Sistema de Gestión",
        "items": [
            {
                "account": {"code": cuenta, "movement": "Debit"},
                "description": _DESC_CUENTA.get(cuenta, "Caja"),
                "value": monto_banco
            },
            {
                "account": {"code": "13050501", "movement": "Credit"},
                "description": "Abono",
                "value": monto,
                "due": {
                    "prefix": due_prefix,
                    "consecutive": due_consecutive,
                    "quote": 1,
                    "date": date.today().isoformat()
                }
            },
            *([{
                "account": {"code": "53958101", "movement": "Debit"},
                "description": "Ajuste al peso",
                "value": ajuste
            }] if ajuste > 0 else [{
                "account": {"code": "42958101", "movement": "Credit"},
                "description": "Ajuste al peso",
                "value": abs(ajuste)
            }] if ajuste < 0 else [])
        ]
    }

    try:
        resp = requests.post(
            f"{SIIGO_BASE}/vouchers",
            headers=siigo_headers(),
            json=payload,
            timeout=30
        )
        if resp.status_code >= 400:
            err_body = resp.text[:500]
            return {"ok": False, "error": f"Siigo HTTP {resp.status_code}: {err_body}"}
        data = resp.json()
        rc_id = str(data.get("id", ""))
        rc_name = str(data.get("name", "") or "")
        prefix_rc = data.get("prefix", "")
        number_rc = data.get("number", "")
        # Prefer 'name' (e.g. "RC-1-1234"), fallback to prefix-number, then id
        rc_numero = rc_name if rc_name else (f"{prefix_rc}-{number_rc}" if number_rc else rc_id)
        return {"ok": True, "simulado": False, "rc_id": rc_id, "rc_numero": rc_numero}
    except requests.exceptions.RequestException as e:
        return {"ok": False, "error": f"Error de red: {str(e)}"}
    except Exception as e:
        return {"ok": False, "error": f"Error inesperado: {str(e)}"}


@router.post("/facturas/{factura_id}/generar-rc")
def generar_rc(factura_id: int):
    """Genera Recibo de Caja para una factura específica."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, siigo_invoice_id, numero, prefix, total, balance,
               estado_pago, medio_pago, cuenta_debito, rc_siigo_id, movimiento_id
        FROM crm_facturas WHERE id = %s
    """, (factura_id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(404, "Factura no encontrada")

    if row[9]:  # rc_siigo_id already set
        cur.close(); conn.close()
        return {"ok": False, "mensaje": "Ya tiene Recibo de Caja generado"}

    if row[5] is not None and float(row[5]) <= 0:
        cur.close(); conn.close()
        return {"ok": False, "mensaje": "Factura ya registrada como pagada en Siigo (balance=0) — no se genera RC duplicado"}

    factura = {
        "id": row[0], "siigo_invoice_id": row[1], "numero": row[2], "prefix": row[3],
        "total": row[4], "balance": row[5], "estado_pago": row[6],
        "medio_pago": row[7], "cuenta_debito": row[8], "movimiento_id": row[10]
    }
    modo = _modo_prueba()
    result = _crear_rc_en_siigo(factura, modo)

    if result["ok"]:
        cur.execute("""
            UPDATE crm_facturas
            SET rc_siigo_id=%s, rc_numero=%s, rc_modo_prueba=%s, estado_pago='pagado'
            WHERE id=%s
        """, (result["rc_id"], result["rc_numero"], result["simulado"], factura_id))
        conn.commit()

        # Escribir RC en Google Sheets si la factura tiene movimiento vinculado
        cur.execute("SELECT movimiento_id FROM crm_facturas WHERE id=%s", (factura_id,))
        mov_row = cur.fetchone()
        if mov_row and mov_row[0]:
            try:
                actualizar_movimiento_sheet(mov_row[0], rc_numero=result["rc_numero"])
            except Exception:
                pass  # no bloquear la respuesta si falla el Sheet

    cur.close(); conn.close()
    return result


@router.post("/facturas/{factura_id}/reset-rc")
def reset_rc(factura_id: int):
    """Limpia el RC de una factura (para re-generar después de anular en Siigo)."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        UPDATE crm_facturas
        SET rc_siigo_id = NULL, rc_numero = NULL, rc_modo_prueba = TRUE
        WHERE id = %s
    """, (factura_id,))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True, "mensaje": "RC limpiado. Puedes generar uno nuevo."}


@router.post("/generar-rc-masivo")
def generar_rc_masivo():
    """Genera RC para todas las facturas marcadas como pagado sin RC aún."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, siigo_invoice_id, numero, prefix, total, balance,
               estado_pago, medio_pago, cuenta_debito, rc_siigo_id, movimiento_id
        FROM crm_facturas
        WHERE estado_pago = 'pagado' AND rc_siigo_id IS NULL
    """)
    facturas = cur.fetchall()
    cur.close(); conn.close()

    modo = _modo_prueba()
    ok_count = 0
    errors = []

    for row in facturas:
        factura = {
            "id": row[0], "siigo_invoice_id": row[1], "numero": row[2], "prefix": row[3],
            "total": row[4], "balance": row[5], "estado_pago": row[6],
            "medio_pago": row[7], "cuenta_debito": row[8], "movimiento_id": row[10]
        }
        result = _crear_rc_en_siigo(factura, modo)
        if result["ok"]:
            conn2 = get_conn()
            cur2 = conn2.cursor()
            cur2.execute("""
                UPDATE crm_facturas
                SET rc_siigo_id=%s, rc_numero=%s, rc_modo_prueba=%s, estado_pago='pagado'
                WHERE id=%s
            """, (result["rc_id"], result["rc_numero"], result["simulado"], row[0]))
            conn2.commit()
            cur2.close(); conn2.close()
            ok_count += 1
        else:
            errors.append({"factura_id": row[0], "error": result.get("error", "Error desconocido")})

    return {
        "ok": True,
        "procesadas": len(facturas),
        "exitosas": ok_count,
        "errores": errors,
        "modo_prueba": modo
    }


# ═══════════════════════════════════════════════════════════════
# SYNC RC DESDE SIIGO
# ═══════════════════════════════════════════════════════════════

def _run_sync_rc():
    """Background task: busca RCs en Siigo y los vincula a facturas."""
    job = _sync_jobs["rc"]
    job["running"] = True; job["ok"] = None; job["msg"] = ""; job["result"] = None
    try:
        _run_sync_rc_inner(job)
    except Exception as e:
        job["running"] = False; job["ok"] = False
        job["msg"] = f"Error inesperado: {str(e)}"


def _run_sync_rc_inner(job):
    from siigo import fetch_vouchers_paginated

    conn = get_conn()
    cur = conn.cursor()

    # Traer facturas sin RC registrado en nuestra BD (pagadas o pendientes con balance=0)
    cur.execute("""
        SELECT f.id, f.numero, f.prefix, f.fecha
        FROM crm_facturas f
        WHERE f.rc_siigo_id IS NULL
          AND (f.estado_pago = 'pagado' OR f.balance = 0)
        ORDER BY f.fecha DESC
    """)
    facturas_sin_rc = cur.fetchall()

    if not facturas_sin_rc:
        cur.close(); conn.close()
        return {"ok": True, "encontrados": 0, "msg": "Todas las facturas ya tienen RC"}

    # Índice por número de factura: numero -> {id, prefix, fecha}
    fac_por_numero = {}
    for row in facturas_sin_rc:
        fid, numero, prefix, fecha = row
        fac_por_numero[int(numero)] = {"id": fid, "prefix": prefix, "fecha": fecha}

    job["step"] = "Descargando RCs de Siigo..."

    # Paginar todos los RCs de Siigo y extraer referencia exacta de factura
    # Cada RC tiene un item de crédito (13050501) con campo "due" que contiene
    # el prefix y consecutive (número) de la factura original
    rc_por_factura_numero = {}  # numero_factura -> {id, numero_rc, fecha}
    page = 1
    while True:
        data = fetch_vouchers_paginated(page=page, page_size=100)
        results = data.get("results", [])
        if not results:
            break
        for v in results:
            doc = v.get("document", {}) or {}
            if doc.get("id") != 3619:
                continue
            v_id        = str(v.get("id", ""))
            v_num       = v.get("number")
            v_name      = str(v.get("name", "") or "")
            v_fecha_str = (v.get("date") or "")[:10]
            if not v_id:
                continue
            try:
                v_fecha = datetime.strptime(v_fecha_str, "%Y-%m-%d").date() if v_fecha_str else None
            except ValueError:
                v_fecha = None
            rc_num_str = v_name if (v_name and v_name != str(v_num)) else str(v_num)

            # Buscar el item de crédito con "due" → referencia exacta a la factura
            items = v.get("items", []) or []
            for item in items:
                acc = item.get("account", {}) or {}
                due = item.get("due") or {}
                if acc.get("movement") == "Credit" and due.get("consecutive"):
                    inv_numero = int(due["consecutive"])
                    # Si ya tenemos un RC para ese número, no sobreescribir (el más reciente gana)
                    if inv_numero not in rc_por_factura_numero:
                        rc_por_factura_numero[inv_numero] = {
                            "id": v_id, "numero": rc_num_str, "fecha": v_fecha
                        }
                    break

        total_results = data.get("pagination", {}).get("total_results", 0)
        if page * 100 >= total_results:
            break
        page += 1

    # Cruzar: para cada factura sin RC, buscar si Siigo tiene un RC con ese número de factura
    encontrados = ambiguos = 0
    for inv_numero, fac in fac_por_numero.items():
        rc = rc_por_factura_numero.get(inv_numero)
        if not rc:
            continue
        cur.execute("""
            UPDATE crm_facturas
            SET rc_siigo_id = %s, rc_numero = %s, rc_modo_prueba = FALSE
            WHERE id = %s AND rc_siigo_id IS NULL
        """, (rc["id"], rc["numero"], fac["id"]))
        if cur.rowcount:
            encontrados += 1

    conn.commit()
    cur.close(); conn.close()
    result = {"ok": True, "encontrados": encontrados,
              "revisadas": len(facturas_sin_rc), "rcs_en_siigo": len(rc_por_factura_numero)}
    job["running"] = False; job["ok"] = True
    job["msg"] = f"{encontrados} RC vinculados de {len(rc_por_factura_numero)} encontrados en Siigo"
    job["step"] = "Completado"; job["result"] = result


@router.post("/sync/rc-siigo")
def sync_rc_siigo(background_tasks: BackgroundTasks):
    job = _sync_jobs["rc"]
    if job["running"]:
        return {"started": False, "running": True, "msg": "Verificación ya en curso"}
    background_tasks.add_task(_run_sync_rc)
    return {"started": True, "running": True, "msg": "Verificación iniciada"}


@router.get("/sync/rc-siigo/status")
def sync_rc_status():
    job = _sync_jobs["rc"]
    return {"running": job["running"], "ok": job["ok"], "msg": job["msg"],
            "step": job["step"], "result": job["result"]}


@router.get("/sync/rc-siigo/reset")
@router.post("/sync/rc-siigo/reset")
def sync_rc_reset():
    """Fuerza el reset del job de sync RC (por si quedó trabado)."""
    job = _sync_jobs["rc"]
    job["running"] = False; job["ok"] = None; job["msg"] = "Resetado manualmente"; job["step"] = ""
    return {"ok": True, "msg": "Job reseteado. Ya puedes volver a correr la verificación."}


# ═══════════════════════════════════════════════════════════════
# ENDPOINTS — CONCILIACIÓN BANCARIA
# ═══════════════════════════════════════════════════════════════

@router.get("/conciliacion")
def get_conciliacion(banco: Optional[str] = None, solo_pendientes: bool = True):
    """Retorna movimientos bancarios no conciliados + facturas pendientes para matching."""
    conn = get_conn()
    cur = conn.cursor()

    # Movimientos sin conciliar (créditos/ingresos)
    banco_filter = "AND banco = %s" if banco else ""
    pendiente_filter = "AND conciliado = FALSE" if solo_pendientes else ""
    params = []
    if banco:
        params.append(banco)

    cur.execute(f"""
        SELECT id, banco, fecha, descripcion, valor, estado,
               rc_sheet, cliente_sheet, conciliado, factura_id, sheet_tab
        FROM movimientos_bancarios
        WHERE valor > 0 {banco_filter} {pendiente_filter}
        ORDER BY fecha DESC
        LIMIT 500
    """, params)

    movimientos = []
    for r in cur.fetchall():
        movimientos.append({
            "id": r[0], "banco": r[1],
            "fecha": r[2].isoformat() if r[2] else None,
            "descripcion": r[3], "valor": float(r[4]),
            "estado": r[5], "rc_sheet": r[6], "cliente_sheet": r[7],
            "conciliado": r[8], "factura_id": r[9], "sheet_tab": r[10]
        })

    # Facturas pendientes de pago — con medio_esperado desde ventas_daily
    cur.execute("""
        SELECT f.id, f.siigo_invoice_id, f.numero, f.prefix, f.fecha,
               f.total, f.balance, f.estado_pago, f.medio_pago,
               COALESCE(NULLIF(TRIM(c.nombre), ''), NULLIF(f.cliente_nombre, '(Sin nombre)'), f.cliente_nombre) AS cliente_nombre,
               COALESCE(c.cedula, f.cliente_cedula) AS cliente_cedula,
               vd.medio_pago AS vd_medio_pago
        FROM crm_facturas f
        LEFT JOIN crm_clientes c ON c.id = f.cliente_id
        LEFT JOIN ventas_daily vd ON vd.factura_id = f.id
        WHERE f.estado_pago = 'pendiente' AND f.balance > 0
        ORDER BY f.fecha DESC
        LIMIT 500
    """)
    facturas_pendientes = []
    for r in cur.fetchall():
        medio_esperado = r[11]  # medio_pago de ventas_daily
        bancos = _MEDIO_A_BANCOS.get(medio_esperado, None) if medio_esperado else None
        facturas_pendientes.append({
            "id": r[0], "siigo_invoice_id": r[1], "numero": r[2], "prefix": r[3],
            "factura": f"{r[3]}-{r[2]}" if r[3] else str(r[2]),
            "fecha": r[4].isoformat() if r[4] else None,
            "total": float(r[5]), "balance": float(r[6]),
            "estado_pago": r[7], "medio_pago": r[8],
            "cliente_nombre": r[9], "cliente_cedula": r[10],
            "medio_esperado": medio_esperado,
            "bancos_filtro": bancos,
            "banco_sugerido": bancos[0] if bancos else None
        })

    cur.close(); conn.close()
    return {"movimientos": movimientos, "facturas_pendientes": facturas_pendientes}


class ConciliarIn(BaseModel):
    movimiento_id: int
    factura_id: int
    medio_pago: Optional[str] = None
    cuenta_debito: Optional[str] = None

@router.post("/conciliar")
def conciliar(data: ConciliarIn):
    """Vincula un movimiento bancario con una factura y la marca como pagada."""
    conn = get_conn()
    cur = conn.cursor()

    # Verify both exist
    cur.execute("SELECT id, valor FROM movimientos_bancarios WHERE id = %s", (data.movimiento_id,))
    mov = cur.fetchone()
    if not mov:
        cur.close(); conn.close()
        raise HTTPException(404, "Movimiento no encontrado")

    cur.execute("SELECT id FROM crm_facturas WHERE id = %s", (data.factura_id,))
    if not cur.fetchone():
        cur.close(); conn.close()
        raise HTTPException(404, "Factura no encontrada")

    # Auto-detect banco and set cuenta_debito if not provided
    cur.execute("SELECT banco FROM movimientos_bancarios WHERE id = %s", (data.movimiento_id,))
    banco = cur.fetchone()[0]
    cuenta = data.cuenta_debito
    if not cuenta:
        if banco == "BDB":
            cuenta = "11100501"
        elif banco == "BANCOLOMBIA":
            cuenta = "11200502"
        else:
            cuenta = "11050501"

    medio = data.medio_pago or (banco if banco else "Efectivo")

    cur.execute("""
        UPDATE movimientos_bancarios
        SET conciliado=TRUE, factura_id=%s
        WHERE id=%s
    """, (data.factura_id, data.movimiento_id))

    cur.execute("""
        UPDATE crm_facturas
        SET estado_pago='pagado', movimiento_id=%s,
            medio_pago=COALESCE(medio_pago,%s),
            cuenta_debito=COALESCE(cuenta_debito,%s)
        WHERE id=%s
    """, (data.movimiento_id, medio, cuenta, data.factura_id))

    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}


@router.get("/sugerencias/{factura_id}")
def get_sugerencias(factura_id: int):
    """Retorna movimientos sugeridos para una factura, rankeados por proximidad de monto y fecha."""
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, siigo_invoice_id, balance, total, fecha, cliente_nombre
        FROM crm_facturas WHERE id = %s
    """, (factura_id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(404, "Factura no encontrada")

    monto_ref = float(row[2]) if row[2] else float(row[3])  # balance si existe, si no total
    fecha_ref = row[4]

    # Buscar medio_pago en ventas_daily para filtrar por banco esperado
    cur.execute("""
        SELECT medio_pago FROM ventas_daily WHERE factura_id = %s LIMIT 1
    """, (factura_id,))
    vd_row = cur.fetchone()
    medio_pago_vd = vd_row[0] if vd_row else None
    bancos_filtro = _MEDIO_A_BANCOS.get(medio_pago_vd, None) if medio_pago_vd else None
    # bancos_filtro = None → sin filtro; [] → sin movimiento esperado (efectivo/cruce)

    # Traer movimientos no conciliados: desde 15 días antes de la factura hasta hoy
    fecha_minima = (fecha_ref - timedelta(days=15)) if fecha_ref else None

    # Para LINK: traer AVALPAY primero (identificación) + BDB (contable)
    # Para otros: filtrar por bancos_filtro si existe
    if bancos_filtro is not None and len(bancos_filtro) == 0:
        # EFECTIVO / CRUCE / MARIA INES: no hay movimiento bancario
        cur.close(); conn.close()
        return {
            "factura_id": factura_id, "monto_ref": monto_ref, "sugerencias": [],
            "medio_esperado": medio_pago_vd, "bancos_filtro": bancos_filtro,
            "sin_movimiento": True,
            "msg": f"Medio de pago '{medio_pago_vd}' no genera movimiento bancario"
        }

    # Construir query con filtro de banco opcional
    banco_sql = ""
    params_mov: list = []
    if bancos_filtro:
        banco_sql = "AND banco = ANY(%s)"
        params_mov.append(bancos_filtro)

    if fecha_minima:
        params_mov_q = [fecha_minima] + params_mov if not bancos_filtro else [fecha_minima] + params_mov
        cur.execute(f"""
            SELECT id, banco, fecha, descripcion, valor, estado, rc_sheet, cliente_sheet, sheet_tab
            FROM movimientos_bancarios
            WHERE conciliado = FALSE AND valor > 0 AND fecha >= %s {banco_sql}
            ORDER BY fecha DESC
        """, [fecha_minima] + (params_mov if bancos_filtro else []))
    else:
        cur.execute(f"""
            SELECT id, banco, fecha, descripcion, valor, estado, rc_sheet, cliente_sheet, sheet_tab
            FROM movimientos_bancarios
            WHERE conciliado = FALSE AND valor > 0 {banco_sql}
            ORDER BY fecha DESC
            LIMIT 1000
        """, params_mov)
    movimientos = cur.fetchall()
    cur.close(); conn.close()

    scored = []
    for m in movimientos:
        mov_valor = float(m[4])
        mov_fecha = m[2]
        mov_banco = m[1]

        # Score de monto: 100% si coincide exacto, cae linealmente hasta 0% con diferencia >50%
        if monto_ref > 0:
            diff_pct = abs(mov_valor - monto_ref) / monto_ref
        else:
            diff_pct = 1.0
        score_monto = max(0.0, 1.0 - diff_pct * 2)  # 50% diff → score 0

        # Score de fecha: 100% mismo día, cae a 0% con >60 días de diferencia
        if fecha_ref and mov_fecha:
            dias = abs((mov_fecha - fecha_ref).days)
            score_fecha = max(0.0, 1.0 - dias / 60)
        else:
            score_fecha = 0.0

        score_total = round(score_monto * 0.75 + score_fecha * 0.25, 3)

        if score_total < 0.1:
            continue

        diff_abs = round(mov_valor - monto_ref, 0)
        # Para LINK: AVALPAY es solo identificación (no conciliar contablemente)
        es_avalpay_identificacion = (mov_banco == "AVALPAY" and medio_pago_vd in ("LINK", "AVALPAY"))
        scored.append({
            "id": m[0], "banco": mov_banco,
            "fecha": m[2].isoformat() if m[2] else None,
            "descripcion": m[3], "valor": mov_valor,
            "estado": m[5], "rc_sheet": m[6], "cliente_sheet": m[7], "sheet_tab": m[8],
            "score": score_total,
            "score_monto": round(score_monto, 3),
            "score_fecha": round(score_fecha, 3),
            "diff_monto": diff_abs,
            "diff_pct": round(diff_pct * 100, 2),
            "match_exacto": diff_pct < 0.001,
            "match_cercano": 0.001 <= diff_pct <= 0.01,
            "solo_identificacion": es_avalpay_identificacion
        })

    # Para LINK: ordenar AVALPAY (identificación) antes que BDB
    if medio_pago_vd in ("LINK", "AVALPAY"):
        scored.sort(key=lambda x: (not x.get("solo_identificacion", False), -x["score"]))
    else:
        scored.sort(key=lambda x: -x["score"])

    return {
        "factura_id": factura_id, "monto_ref": monto_ref, "sugerencias": scored[:10],
        "medio_esperado": medio_pago_vd, "bancos_filtro": bancos_filtro,
        "sin_movimiento": False
    }


@router.post("/auto-conciliar")
def auto_conciliar():
    """Empareja automáticamente facturas con movimientos de monto exacto (±1%) y fecha ±30 días."""
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, balance, total, fecha FROM crm_facturas
        WHERE estado_pago = 'pendiente' AND movimiento_id IS NULL
        FOR UPDATE SKIP LOCKED
    """)
    facturas = cur.fetchall()

    cur.execute("""
        SELECT id, valor, fecha, banco FROM movimientos_bancarios
        WHERE conciliado = FALSE AND valor > 0
        FOR UPDATE SKIP LOCKED
    """)
    movimientos = cur.fetchall()

    vinculados = 0
    usados = set()

    for fac in facturas:
        fid, balance, total, fecha_fac = fac
        monto = float(balance) if balance else float(total)

        mejor = None
        mejor_dias = 9999
        for mov in movimientos:
            mid, valor, fecha_mov, banco = mov
            if mid in usados:
                continue
            valor = float(valor)
            if monto == 0:
                continue
            diff_pct = abs(valor - monto) / monto
            if diff_pct > 0.01:  # solo ±1%
                continue
            dias = abs((fecha_mov - fecha_fac).days) if fecha_fac and fecha_mov else 9999
            if dias > 30:
                continue
            if dias < mejor_dias:
                mejor = mov
                mejor_dias = dias

        if mejor:
            mid, valor, fecha_mov, banco = mejor
            cuenta = "11100501" if banco == "BDB" else ("11200502" if banco == "BANCOLOMBIA" else "11050501")
            cur.execute("""
                UPDATE movimientos_bancarios SET conciliado=TRUE, factura_id=%s WHERE id=%s
            """, (fid, mid))
            cur.execute("""
                UPDATE crm_facturas SET estado_pago='pagado', movimiento_id=%s,
                  medio_pago=COALESCE(medio_pago,%s), cuenta_debito=COALESCE(cuenta_debito,%s)
                WHERE id=%s
            """, (mid, banco, cuenta, fid))
            usados.add(mid)
            vinculados += 1

    conn.commit()
    cur.close(); conn.close()
    return {"ok": True, "vinculados": vinculados}


# ═══════════════════════════════════════════════════════════════
# ENDPOINTS — SYNC LOG
# ═══════════════════════════════════════════════════════════════

@router.get("/debug/sheet-preview/{tab_name:path}")
def debug_sheet_preview(tab_name: str):
    """Muestra las primeras 5 filas de una pestaña + hex del nombre para detectar caracteres ocultos."""
    try:
        service = _get_sheets_service()
        # Obtener nombre exacto desde metadatos
        meta = service.spreadsheets().get(spreadsheetId=SHEET_ID).execute()
        sheets = {s["properties"]["title"]: s["properties"]["sheetId"]
                  for s in meta.get("sheets", [])}

        # Buscar coincidencia normalizando espacios
        exact_name = None
        for name in sheets:
            if name.strip().upper().replace("\xa0", " ") == tab_name.strip().upper().replace("\xa0", " "):
                exact_name = name
                break

        if not exact_name:
            return {"error": "Pestaña no encontrada", "disponibles": list(sheets.keys())}

        # Mostrar hex del nombre para detectar caracteres especiales
        name_hex = exact_name.encode("utf-8").hex()

        safe_name = exact_name.replace("'", "\\'")
        result = service.spreadsheets().values().get(
            spreadsheetId=SHEET_ID,
            range=f"'{safe_name}'!A1:M10"
        ).execute()
        rows = result.get("values", [])
        return {"tab_exacto": exact_name, "hex": name_hex, "filas": rows}
    except Exception as e:
        return {"error": str(e)}


@router.get("/debug/sheet-tabs")
def debug_sheet_tabs():
    """Lista las pestañas reales del Sheet y cuántas filas tiene cada una."""
    try:
        service = _get_sheets_service()
        meta = service.spreadsheets().get(spreadsheetId=SHEET_ID).execute()
        sheets = meta.get("sheets", [])
        tabs_info = []
        for s in sheets:
            props = s.get("properties", {})
            tab_name = props.get("title", "")
            grid = props.get("gridProperties", {})
            tabs_info.append({
                "nombre": tab_name,
                "filas": grid.get("rowCount", 0),
                "columnas": grid.get("columnCount", 0),
                "configurada": any(t[0] == tab_name for t in SHEET_TABS)
            })
        return {"tabs": tabs_info, "configuradas": [t[0] for t in SHEET_TABS]}
    except Exception as e:
        return {"error": str(e)}


@router.get("/debug/vouchers-siigo")
def debug_vouchers_siigo():
    """Muestra los primeros 3 vouchers de Siigo para ver su estructura real."""
    from siigo import fetch_vouchers_paginated
    try:
        data = fetch_vouchers_paginated(page=1, page_size=5)
        results = data.get("results", [])
        total = data.get("pagination", {}).get("total_results", 0)
        sample = []
        for v in results[:3]:
            sample.append({
                "id": v.get("id"),
                "document": v.get("document"),
                "number": v.get("number"),
                "name": v.get("name"),
                "date": v.get("date"),
                "total": v.get("total"),
                "customer": v.get("customer"),
                "items": v.get("items"),
                "payment": v.get("payment"),
                "keys": list(v.keys()),
            })
        return {"total_en_siigo": total, "muestra": sample}
    except Exception as e:
        return {"error": str(e)}


@router.get("/debug/credit-notes-siigo")
def debug_credit_notes_siigo():
    """Muestra las primeras 3 NC de Siigo para ver su estructura."""
    from siigo import fetch_credit_notes
    try:
        ncs = fetch_credit_notes()
        sample = ncs[:3]
        return {"total": len(ncs), "muestra": sample, "campos": list(sample[0].keys()) if sample else []}
    except Exception as e:
        return {"error": str(e)}


@router.get("/debug/factura-local/{factura_id}")
def debug_factura_local(factura_id: int):
    """Busca el siigo_invoice_id de una factura por su ID local y consulta Siigo."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id, siigo_invoice_id, numero, prefix, cliente_nombre, total, rc_siigo_id FROM crm_facturas WHERE id=%s", (factura_id,))
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row:
        return {"error": "Factura no encontrada"}
    fid, siigo_id, numero, prefix, nombre, total, rc_id = row
    result = {"id": fid, "siigo_invoice_id": siigo_id, "factura": f"{prefix}-{numero}", "cliente": nombre, "total": float(total or 0), "rc_siigo_id": rc_id}
    if siigo_id:
        try:
            resp = requests.get(f"{SIIGO_BASE}/invoices/{siigo_id}", headers=siigo_headers(), timeout=15)
            resp.raise_for_status()
            data = resp.json()
            result["siigo_balance"] = data.get("balance")
            result["siigo_payments"] = data.get("payments")
            result["siigo_keys"] = list(data.keys())
        except Exception as e:
            result["siigo_error"] = str(e)
    return result


@router.get("/debug/invoice-siigo/{siigo_id}")
def debug_invoice_siigo(siigo_id: str):
    """Trae una factura individual de Siigo por su ID para ver campos de pagos."""
    try:
        resp = requests.get(
            f"{SIIGO_BASE}/invoices/{siigo_id}",
            headers=siigo_headers(),
            timeout=15
        )
        resp.raise_for_status()
        data = resp.json()
        return {
            "keys": list(data.keys()),
            "document": data.get("document"),
            "prefix": data.get("prefix"),
            "number": data.get("number"),
            "name": data.get("name"),
            "balance": data.get("balance"),
            "total": data.get("total"),
            "payments": data.get("payments"),
            "items": data.get("items"),
            "observations": data.get("observations"),
        }
    except Exception as e:
        return {"error": str(e)}


@router.get("/debug/factura-siigo")
def debug_factura_siigo():
    """Muestra los primeros campos crudos de una factura reciente de Siigo para diagnóstico."""
    try:
        from siigo import fetch_invoices
        from datetime import date, timedelta
        hoy = date.today()
        facturas = fetch_invoices((hoy - timedelta(days=30)).isoformat(), hoy.isoformat())
        if not facturas:
            return {"msg": "Sin facturas en los últimos 30 días"}
        sample = facturas[0]
        # Solo devolver campos relevantes, no todo el objeto
        return {
            "keys": list(sample.keys()),
            "payments": sample.get("payments"),
            "balance": sample.get("balance"),
            "total": sample.get("total"),
        }
    except Exception as e:
        return {"error": str(e)}


@router.get("/debug/customer-siigo/{siigo_id}")
def debug_customer_siigo(siigo_id: str):
    """Consulta un cliente en Siigo por su UUID para ver datos reales."""
    try:
        resp = requests.get(f"{SIIGO_BASE}/customers/{siigo_id}", headers=siigo_headers(), timeout=15)
        if resp.status_code >= 400:
            return {"error": f"HTTP {resp.status_code}", "body": resp.text[:500]}
        data = resp.json()
        return {
            "id": data.get("id"),
            "identification": data.get("identification"),
            "check_digit": data.get("check_digit"),
            "name": data.get("name"),
            "id_type": data.get("id_type"),
            "person_type": data.get("person_type"),
            "branch_office": data.get("branch_office"),
            "active": data.get("active"),
        }
    except Exception as e:
        return {"error": str(e)}


@router.get("/sync/log")
def get_sync_log():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT tipo, fecha, registros, nuevos, actualizados, detalle
        FROM crm_sync_log ORDER BY fecha DESC LIMIT 50
    """)
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [
        {"tipo": r[0], "fecha": r[1].isoformat() if r[1] else None,
         "registros": r[2], "nuevos": r[3], "actualizados": r[4], "detalle": r[5]}
        for r in rows
    ]


# ═══════════════════════════════════════════════════════════════
# TAREA 1 — CONCILIACIÓN EXTENDIDA
# ═══════════════════════════════════════════════════════════════

@router.get("/conciliados")
def get_conciliados(q: Optional[str] = None, filtro_rc: Optional[str] = None):
    """Facturas ya pagadas con info del cliente y movimiento bancario vinculado."""
    conn = get_conn()
    cur = conn.cursor()

    q_filter = ""
    params: list = []
    if q:
        q_filter = "AND (c.nombre ILIKE %s OR CAST(f.numero AS TEXT) ILIKE %s)"
        params.extend([f"%{q}%", f"%{q}%"])
    if filtro_rc == "con_rc":
        q_filter += " AND f.rc_siigo_id IS NOT NULL"
    elif filtro_rc == "sin_rc":
        q_filter += " AND f.rc_siigo_id IS NULL"

    cur.execute(f"""
        SELECT f.id, f.siigo_invoice_id, f.numero, f.prefix, f.fecha,
               f.total, f.balance, f.estado_pago, f.medio_pago, f.cuenta_debito,
               f.movimiento_id, f.rc_siigo_id, f.rc_numero, f.rc_modo_prueba,
               c.nombre AS cliente_nombre, c.cedula AS cliente_cedula,
               m.banco, m.fecha AS mov_fecha, m.valor AS mov_valor, m.descripcion AS mov_desc
        FROM crm_facturas f
        LEFT JOIN crm_clientes c ON c.id = f.cliente_id
        LEFT JOIN movimientos_bancarios m ON m.id = f.movimiento_id
        WHERE f.estado_pago = 'pagado'
          AND COALESCE(f.tiene_nc, FALSE) = FALSE
          {q_filter}
        ORDER BY f.fecha DESC
        LIMIT 500
    """, params)

    cols = [
        "id", "siigo_invoice_id", "numero", "prefix", "fecha",
        "total", "balance", "estado_pago", "medio_pago", "cuenta_debito",
        "movimiento_id", "rc_siigo_id", "rc_numero", "rc_modo_prueba",
        "cliente_nombre", "cliente_cedula",
        "banco", "mov_fecha", "mov_valor", "mov_desc"
    ]
    result = []
    for row in cur.fetchall():
        d = dict(zip(cols, row))
        if d["fecha"]:
            d["fecha"] = d["fecha"].isoformat()
        if d["mov_fecha"]:
            d["mov_fecha"] = d["mov_fecha"].isoformat()
        if d["total"] is not None:
            d["total"] = float(d["total"])
        if d["balance"] is not None:
            d["balance"] = float(d["balance"])
        if d["mov_valor"] is not None:
            d["mov_valor"] = float(d["mov_valor"])
        # Campo combinado factura
        d["factura"] = f"{d['prefix']}-{d['numero']}" if d.get("prefix") else str(d.get("numero", ""))
        # Fecha de pago: usar mov_fecha si existe, si no fecha de la factura
        d["fecha_pago"] = d["mov_fecha"] or d["fecha"]
        # Registrado en: banco del movimiento, medio_pago manual, o "Siigo" si se pagó directo allá
        _banco_nombres = {"BDB": "Banco de Bogotá", "BANCOLOMBIA": "Bancolombia"}
        banco_raw = d.get("banco") or d.get("medio_pago") or "Siigo"
        d["banco_display"] = _banco_nombres.get(banco_raw, banco_raw)
        result.append(d)

    cur.close(); conn.close()
    return result


@router.post("/facturas/{factura_id}/efectivo")
def pagar_efectivo(factura_id: int):
    """Marca una factura como pagada en efectivo (sin movimiento bancario)."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id, estado_pago FROM crm_facturas WHERE id = %s", (factura_id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(404, "Factura no encontrada")
    if row[1] == "pagado":
        cur.close(); conn.close()
        raise HTTPException(400, "La factura ya está pagada")

    cur.execute("""
        UPDATE crm_facturas
        SET estado_pago='pagado', medio_pago='Efectivo', cuenta_debito='11050501'
        WHERE id=%s
    """, (factura_id,))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True, "factura_id": factura_id}


@router.post("/facturas/{factura_id}/revertir")
def revertir_factura(factura_id: int):
    """Revierte una conciliación: vuelve la factura a pendiente y desliga el movimiento bancario."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, estado_pago, movimiento_id, rc_siigo_id, rc_modo_prueba
        FROM crm_facturas WHERE id = %s
    """, (factura_id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(404, "Factura no encontrada")

    _, estado, mov_id, rc_id, rc_prueba = row

    if rc_id and not rc_prueba:
        cur.close(); conn.close()
        raise HTTPException(400, "No se puede revertir: ya tiene un RC real generado en Siigo. Anúlalo primero en Siigo.")

    # Obtener datos del movimiento para limpiar Sheets
    sheet_info = None
    if mov_id:
        cur.execute("""
            SELECT banco, sheet_tab, sheet_row
            FROM movimientos_bancarios WHERE id = %s
        """, (mov_id,))
        sheet_info = cur.fetchone()

        cur.execute("""
            UPDATE movimientos_bancarios
            SET conciliado = FALSE, factura_id = NULL
            WHERE id = %s
        """, (mov_id,))

    # Revertir factura a pendiente
    cur.execute("""
        UPDATE crm_facturas
        SET estado_pago = 'pendiente',
            medio_pago = NULL,
            cuenta_debito = NULL,
            movimiento_id = NULL,
            rc_siigo_id = NULL,
            rc_numero = NULL,
            rc_modo_prueba = TRUE
        WHERE id = %s
    """, (factura_id,))

    conn.commit()
    cur.close(); conn.close()

    # Limpiar celdas en Google Sheets (ESTADO, RC, CLIENTE)
    sheet_ok = False
    if sheet_info:
        banco, sheet_tab, sheet_row = sheet_info
        if sheet_tab and sheet_row:
            try:
                cols = _SHEET_COLS.get(banco, {"estado": 3, "rc": 4, "cliente": 5})
                service = _get_sheets_service()
                exact_tab = _get_exact_tab_name(service, sheet_tab)
                safe_tab  = exact_tab.replace("'", "\\'")
                # Borrar ESTADO, RC y CLIENTE
                for col_key in ("estado", "rc", "cliente"):
                    col_letter = _col_letter(cols[col_key])
                    service.spreadsheets().values().update(
                        spreadsheetId=SHEET_ID,
                        range=f"'{safe_tab}'!{col_letter}{sheet_row}",
                        valueInputOption="RAW",
                        body={"values": [[""]]}
                    ).execute()
                sheet_ok = True
            except Exception:
                pass  # no bloquear la reversión si falla el Sheet

    return {"ok": True, "sheet_limpiado": sheet_ok}


class EditarFacturaIn(BaseModel):
    medio_pago: Optional[str] = None
    cuenta_debito: Optional[str] = None
    origen_canal: Optional[str] = None

@router.put("/facturas/{factura_id}/editar")
def editar_factura_conciliada(factura_id: int, data: EditarFacturaIn):
    """Edita campos de una factura conciliada sin revertir el pago."""
    conn = get_conn()
    cur = conn.cursor()
    sets, vals = [], []
    if data.medio_pago is not None:
        sets.append("medio_pago=%s"); vals.append(data.medio_pago)
        # Auto-actualizar cuenta_debito si no viene explícita
        if data.cuenta_debito is None:
            sets.append("cuenta_debito=%s"); vals.append(_cuenta_for_medio(data.medio_pago))
    if data.cuenta_debito is not None:
        sets.append("cuenta_debito=%s"); vals.append(data.cuenta_debito)
    if data.origen_canal is not None:
        sets.append("origen_canal=%s"); vals.append(data.origen_canal)
    if not sets:
        cur.close(); conn.close()
        return {"ok": True}
    vals.append(factura_id)
    cur.execute(f"UPDATE crm_facturas SET {', '.join(sets)} WHERE id=%s", vals)
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}


class RcMasivoIn(BaseModel):
    factura_ids: list

@router.post("/rc-masivo-seleccion")
def rc_masivo_seleccion(data: RcMasivoIn):
    """Genera RC en Siigo para las facturas indicadas (pagadas y sin RC)."""
    if not data.factura_ids:
        return {"ok": True, "exitosas": 0, "errores": []}

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, siigo_invoice_id, numero, prefix, total, balance,
               estado_pago, medio_pago, cuenta_debito, rc_siigo_id, movimiento_id
        FROM crm_facturas
        WHERE id = ANY(%s) AND estado_pago = 'pagado' AND rc_siigo_id IS NULL
    """, (list(data.factura_ids),))
    facturas = cur.fetchall()
    cur.close(); conn.close()

    modo = _modo_prueba()
    ok_count = 0
    errors = []

    for row in facturas:
        factura = {
            "id": row[0], "siigo_invoice_id": row[1], "numero": row[2], "prefix": row[3],
            "total": row[4], "balance": row[5], "estado_pago": row[6],
            "medio_pago": row[7], "cuenta_debito": row[8], "movimiento_id": row[10]
        }
        result = _crear_rc_en_siigo(factura, modo)
        if result["ok"]:
            conn2 = get_conn()
            cur2 = conn2.cursor()
            cur2.execute("""
                UPDATE crm_facturas
                SET rc_siigo_id=%s, rc_numero=%s, rc_modo_prueba=%s, estado_pago='pagado'
                WHERE id=%s
            """, (result["rc_id"], result["rc_numero"], result["simulado"], row[0]))
            conn2.commit()
            cur2.close(); conn2.close()
            ok_count += 1
        else:
            errors.append({"factura_id": row[0], "error": result.get("error", "Error desconocido")})

    return {"ok": True, "exitosas": ok_count, "errores": errors}


# ═══════════════════════════════════════════════════════════════
# TAREA 3 — ENDPOINTS CRM (dashboard, clientes, seguimientos, prospectos)
# ═══════════════════════════════════════════════════════════════

@router.get("/dashboard-crm")
def dashboard_crm():
    """Métricas generales del CRM."""
    conn = get_conn()
    cur = conn.cursor()
    hoy = date.today()
    hace_30 = (hoy - timedelta(days=30)).isoformat()
    hace_90 = (hoy - timedelta(days=90)).isoformat()

    # Clasificación por última compra (usa columna ultima_compra de crm_clientes si existe,
    # sino la calcula desde crm_facturas)
    cur.execute("""
        SELECT
          COUNT(*) FILTER (WHERE uc >= %s) AS activos,
          COUNT(*) FILTER (WHERE uc < %s AND uc >= %s) AS en_riesgo,
          COUNT(*) FILTER (WHERE uc < %s OR uc IS NULL) AS inactivos
        FROM (
          SELECT c.id,
                 COALESCE(c.ultima_compra, MAX(f.fecha)) AS uc
          FROM crm_clientes c
          LEFT JOIN crm_facturas f ON f.cliente_id = c.id
          GROUP BY c.id, c.ultima_compra
        ) sub
    """, (hace_90, hace_90, hace_30, hace_90))
    row = cur.fetchone()
    total_activos, total_riesgo, total_inactivos = (row[0] or 0), (row[1] or 0), (row[2] or 0)

    # Ticket promedio últimos 90 días
    cur.execute("""
        SELECT AVG(total) FROM crm_facturas
        WHERE fecha >= %s AND total > 0
    """, (hace_90,))
    ticket_row = cur.fetchone()
    ticket_promedio = float(ticket_row[0] or 0)

    # Top 10 clientes por total_compras
    cur.execute("""
        SELECT c.nombre,
               COALESCE(c.total_compras, SUM(f.total)) AS total,
               COALESCE(c.num_facturas, COUNT(f.id)) AS num_facturas,
               COALESCE(c.ultima_compra, MAX(f.fecha)) AS ultima_compra,
               c.segmento
        FROM crm_clientes c
        LEFT JOIN crm_facturas f ON f.cliente_id = c.id
        GROUP BY c.id
        ORDER BY total DESC NULLS LAST
        LIMIT 10
    """)
    top_clientes = [
        {
            "nombre": r[0],
            "total": float(r[1] or 0),
            "total_compras": float(r[1] or 0),
            "num_facturas": r[2] or 0,
            "ultima_compra": r[3].isoformat() if r[3] else None,
            "segmento": r[4]
        }
        for r in cur.fetchall()
    ]

    # Por segmento
    cur.execute("""
        SELECT segmento, COUNT(*) AS count,
               COALESCE(SUM(total_compras), 0) AS total_ventas
        FROM crm_clientes
        GROUP BY segmento
        ORDER BY count DESC
    """)
    por_segmento = [
        {"segmento": r[0], "count": r[1], "total_ventas": float(r[2] or 0)}
        for r in cur.fetchall()
    ]

    # Por canal
    cur.execute("""
        SELECT canal_adquisicion, COUNT(*) AS count
        FROM crm_clientes
        GROUP BY canal_adquisicion
        ORDER BY count DESC
    """)
    por_canal = [{"canal": r[0], "count": r[1]} for r in cur.fetchall()]

    # Clientes en riesgo (sin compra 30-90 días)
    cur.execute("""
        SELECT c.id, c.nombre, c.telefono,
               COALESCE(c.ultima_compra, MAX(f.fecha)) AS uc,
               COALESCE(c.total_compras, SUM(f.total)) AS total
        FROM crm_clientes c
        LEFT JOIN crm_facturas f ON f.cliente_id = c.id
        GROUP BY c.id
        HAVING COALESCE(c.ultima_compra, MAX(f.fecha)) < %s
           AND COALESCE(c.ultima_compra, MAX(f.fecha)) >= %s
        ORDER BY uc ASC
        LIMIT 20
    """, (hace_30, hace_90))
    clientes_en_riesgo = []
    for r in cur.fetchall():
        uc = r[3]
        dias = (hoy - uc).days if uc else None
        clientes_en_riesgo.append({
            "id": r[0], "nombre": r[1], "telefono": r[2],
            "ultima_compra": uc.isoformat() if uc else None,
            "dias_sin_compra": dias,
            "total_historico": float(r[4] or 0),
            "total_compras": float(r[4] or 0),
        })

    cur.close(); conn.close()
    return {
        "total_clientes_activos": total_activos,
        "total_clientes_riesgo": total_riesgo,
        "total_clientes_inactivos": total_inactivos,
        "ticket_promedio": round(ticket_promedio, 2),
        "top_clientes": top_clientes,
        "por_segmento": por_segmento,
        "por_canal": por_canal,
        "clientes_en_riesgo": clientes_en_riesgo,
    }


@router.get("/clientes-crm")
def get_clientes_crm(
    q: Optional[str] = None,
    segmento: Optional[str] = None,
    estado: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
    limit: Optional[int] = None,
    offset: Optional[int] = None,
):
    """Lista de clientes CRM con filtros y paginación."""
    conn = get_conn()
    cur = conn.cursor()

    conditions = ["1=1"]
    params: list = []
    if q:
        conditions.append("(c.nombre ILIKE %s OR c.cedula ILIKE %s OR c.telefono ILIKE %s)")
        params.extend([f"%{q}%", f"%{q}%", f"%{q}%"])
    if segmento:
        conditions.append("c.segmento = %s")
        params.append(segmento)
    if estado:
        conditions.append("c.estado_cliente = %s")
        params.append(estado)

    where = " AND ".join(conditions)
    # Soportar limit/offset directo (frontend) o page/page_size
    if limit is not None:
        _limit = limit
        _offset = offset if offset is not None else 0
    else:
        _limit = page_size
        _offset = (page - 1) * page_size

    cur.execute(f"SELECT COUNT(*) FROM crm_clientes c WHERE {where}", params)
    total_count = cur.fetchone()[0]

    cur.execute(f"""
        SELECT c.id, c.siigo_id, c.cedula, c.nombre, c.telefono, c.email,
               c.ciudad, c.segmento, c.canal_adquisicion, c.estado_cliente,
               c.ultima_compra, c.total_compras, c.num_facturas,
               c.responsable, c.cupo_credito, c.dias_credito, c.notas_crm
        FROM crm_clientes c
        WHERE {where}
        ORDER BY c.nombre ASC
        LIMIT %s OFFSET %s
    """, params + [_limit, _offset])

    cols = [
        "id", "siigo_id", "cedula", "nombre", "telefono", "email",
        "ciudad", "segmento", "canal_adquisicion", "estado_cliente",
        "ultima_compra", "total_compras", "num_facturas",
        "responsable", "cupo_credito", "dias_credito", "notas_crm"
    ]
    clientes = []
    for row in cur.fetchall():
        d = dict(zip(cols, row))
        if d["ultima_compra"]:
            d["ultima_compra"] = d["ultima_compra"].isoformat()
        if d["total_compras"] is not None:
            d["total_compras"] = float(d["total_compras"])
        if d["cupo_credito"] is not None:
            d["cupo_credito"] = float(d["cupo_credito"])
        # Aliases para compatibilidad frontend
        d["estado"] = d.get("estado_cliente")
        d["origen_canal"] = d.get("canal_adquisicion")
        d["total_ventas"] = d.get("total_compras")
        clientes.append(d)

    cur.close(); conn.close()
    return {
        "items": clientes,
        "total": total_count,
        "clientes": clientes,
        "total_count": total_count,
        "page": page,
        "page_size": page_size,
    }


@router.get("/clientes-crm/{cliente_id}")
def get_cliente_crm(cliente_id: int):
    """Ficha completa del cliente CRM."""
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, siigo_id, cedula, nombre, telefono, email, direccion, ciudad,
               segmento, canal_adquisicion, estado_cliente, ultima_compra,
               total_compras, num_facturas, responsable, cupo_credito, dias_credito,
               notas_crm, origen_canal, notas, activo
        FROM crm_clientes WHERE id = %s
    """, (cliente_id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(404, "Cliente no encontrado")

    cols = [
        "id", "siigo_id", "cedula", "nombre", "telefono", "email", "direccion", "ciudad",
        "segmento", "canal_adquisicion", "estado_cliente", "ultima_compra",
        "total_compras", "num_facturas", "responsable", "cupo_credito", "dias_credito",
        "notas_crm", "origen_canal", "notas", "activo"
    ]
    cliente = dict(zip(cols, row))
    if cliente["ultima_compra"]:
        cliente["ultima_compra"] = cliente["ultima_compra"].isoformat()
    if cliente["total_compras"] is not None:
        cliente["total_compras"] = float(cliente["total_compras"])
    if cliente["cupo_credito"] is not None:
        cliente["cupo_credito"] = float(cliente["cupo_credito"])

    # Últimas 20 facturas
    cur.execute("""
        SELECT id, siigo_invoice_id, numero, prefix, fecha, total, balance,
               estado_pago, medio_pago, rc_numero
        FROM crm_facturas WHERE cliente_id = %s
        ORDER BY fecha DESC LIMIT 20
    """, (cliente_id,))
    facturas = []
    for f in cur.fetchall():
        facturas.append({
            "id": f[0], "siigo_invoice_id": f[1], "numero": f[2], "prefix": f[3],
            "factura": f"{f[3]}-{f[2]}" if f[3] else str(f[2]),
            "fecha": f[4].isoformat() if f[4] else None,
            "total": float(f[5] or 0), "balance": float(f[6] or 0),
            "estado_pago": f[7], "medio_pago": f[8], "rc_numero": f[9],
        })

    # Últimos 10 seguimientos
    cur.execute("""
        SELECT id, tipo, descripcion, resultado, fecha, responsable,
               proxima_accion, proxima_fecha
        FROM crm_seguimientos WHERE cliente_id = %s
        ORDER BY fecha DESC LIMIT 10
    """, (cliente_id,))
    seguimientos = []
    for s in cur.fetchall():
        seguimientos.append({
            "id": s[0], "tipo": s[1], "descripcion": s[2], "resultado": s[3],
            "fecha": s[4].isoformat() if s[4] else None,
            "responsable": s[5], "proxima_accion": s[6],
            "proxima_fecha": s[7].isoformat() if s[7] else None,
        })

    # Resumen calculado
    cur.execute("""
        SELECT SUM(total), COUNT(*), MAX(fecha)
        FROM crm_facturas WHERE cliente_id = %s
    """, (cliente_id,))
    res = cur.fetchone()
    total_compras_calc = float(res[0] or 0)
    num_facturas_calc = res[1] or 0
    ultima_compra_calc = res[2]
    dias_sin_compra = (date.today() - ultima_compra_calc).days if ultima_compra_calc else None

    # Aliases para compatibilidad con frontend
    cliente["estado"] = cliente.get("estado_cliente")
    cliente["origen_canal"] = cliente.get("canal_adquisicion")

    cur.close(); conn.close()
    return {
        **cliente,
        "facturas": facturas,
        "seguimientos": seguimientos,
        "resumen": {
            "total_compras": total_compras_calc,
            "num_facturas": num_facturas_calc,
            "ultima_compra": ultima_compra_calc.isoformat() if ultima_compra_calc else None,
            "dias_sin_compra": dias_sin_compra,
        }
    }


class ClienteCrmUpdate(BaseModel):
    segmento: Optional[str] = None
    canal_adquisicion: Optional[str] = None
    estado_cliente: Optional[str] = None
    responsable: Optional[str] = None
    cupo_credito: Optional[float] = None
    dias_credito: Optional[int] = None
    notas_crm: Optional[str] = None

@router.put("/clientes-crm/{cliente_id}")
def update_cliente_crm(cliente_id: int, data: ClienteCrmUpdate):
    """Actualiza campos CRM de un cliente."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id FROM crm_clientes WHERE id = %s", (cliente_id,))
    if not cur.fetchone():
        cur.close(); conn.close()
        raise HTTPException(404, "Cliente no encontrado")

    sets = []
    vals = []
    for field in ("segmento", "canal_adquisicion", "estado_cliente", "responsable",
                  "cupo_credito", "dias_credito", "notas_crm"):
        v = getattr(data, field)
        if v is not None:
            sets.append(f"{field}=%s"); vals.append(v)

    if sets:
        sets.append("actualizado_en=NOW()")
        vals.append(cliente_id)
        cur.execute(f"UPDATE crm_clientes SET {', '.join(sets)} WHERE id=%s", vals)
        conn.commit()

    cur.close(); conn.close()
    return {"ok": True}


class SeguimientoIn(BaseModel):
    cliente_id: int
    tipo: str
    descripcion: str
    resultado: Optional[str] = None
    responsable: Optional[str] = None
    proxima_accion: Optional[str] = None
    proxima_fecha: Optional[str] = None

@router.post("/seguimientos")
def crear_seguimiento(data: SeguimientoIn):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id FROM crm_clientes WHERE id = %s", (data.cliente_id,))
    if not cur.fetchone():
        cur.close(); conn.close()
        raise HTTPException(404, "Cliente no encontrado")

    proxima_fecha = None
    if data.proxima_fecha:
        try:
            proxima_fecha = datetime.strptime(data.proxima_fecha, "%Y-%m-%d").date()
        except ValueError:
            pass

    cur.execute("""
        INSERT INTO crm_seguimientos
          (cliente_id, tipo, descripcion, resultado, responsable, proxima_accion, proxima_fecha)
        VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id
    """, (data.cliente_id, data.tipo, data.descripcion, data.resultado,
          data.responsable, data.proxima_accion, proxima_fecha))
    new_id = cur.fetchone()[0]
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True, "id": new_id}


@router.get("/prospectos")
def get_prospectos(estado: Optional[str] = None):
    conn = get_conn()
    cur = conn.cursor()
    where = "WHERE estado = %s" if estado else ""
    params = [estado] if estado else []
    cur.execute(f"""
        SELECT id, nombre, empresa, telefono, email, ciudad, segmento, canal,
               estado, responsable, notas, fecha_contacto, fecha_seguimiento,
               valor_potencial, convertido_cliente_id, creado_en
        FROM crm_prospectos {where}
        ORDER BY creado_en DESC
    """, params)
    cols = [
        "id", "nombre", "empresa", "telefono", "email", "ciudad", "segmento", "canal",
        "estado", "responsable", "notas", "fecha_contacto", "fecha_seguimiento",
        "valor_potencial", "convertido_cliente_id", "creado_en"
    ]
    result = []
    for row in cur.fetchall():
        d = dict(zip(cols, row))
        for f in ("fecha_contacto", "fecha_seguimiento"):
            if d[f]:
                d[f] = d[f].isoformat()
        if d["creado_en"]:
            d["creado_en"] = d["creado_en"].isoformat()
        if d["valor_potencial"] is not None:
            d["valor_potencial"] = float(d["valor_potencial"])
        result.append(d)
    cur.close(); conn.close()
    return result


class ProspectoIn(BaseModel):
    nombre: str
    empresa: Optional[str] = None
    telefono: Optional[str] = None
    email: Optional[str] = None
    direccion: Optional[str] = None
    ciudad: Optional[str] = None
    segmento: Optional[str] = None
    canal: Optional[str] = None
    estado: Optional[str] = "contactado"
    responsable: Optional[str] = None
    notas: Optional[str] = None
    fecha_contacto: Optional[str] = None
    fecha_seguimiento: Optional[str] = None
    valor_potencial: Optional[float] = 0

@router.get("/prospectos/{prospecto_id}")
def get_prospecto(prospecto_id: int):
    """Retorna un prospecto por ID."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, nombre, empresa, telefono, email, ciudad, direccion, segmento, canal,
               estado, responsable, notas, fecha_contacto, fecha_seguimiento,
               valor_potencial, convertido_cliente_id, creado_en
        FROM crm_prospectos WHERE id = %s
    """, (prospecto_id,))
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row:
        raise HTTPException(404, "Prospecto no encontrado")
    cols = [
        "id", "nombre", "empresa", "telefono", "email", "ciudad", "direccion", "segmento", "canal",
        "estado", "responsable", "notas", "fecha_contacto", "fecha_seguimiento",
        "valor_potencial", "convertido_cliente_id", "creado_en"
    ]
    d = dict(zip(cols, row))
    for f in ("fecha_contacto", "fecha_seguimiento"):
        if d[f]:
            d[f] = d[f].isoformat()
    if d["creado_en"]:
        d["creado_en"] = d["creado_en"].isoformat()
    if d["valor_potencial"] is not None:
        d["valor_potencial"] = float(d["valor_potencial"])
    return d


@router.post("/prospectos")
def crear_prospecto(data: ProspectoIn):
    conn = get_conn()
    cur = conn.cursor()

    def _parse_d(s):
        if not s:
            return None
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except ValueError:
            return None

    cur.execute("""
        INSERT INTO crm_prospectos
          (nombre, empresa, telefono, email, direccion, ciudad, segmento, canal,
           estado, responsable, notas, fecha_contacto, fecha_seguimiento, valor_potencial)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id
    """, (
        data.nombre, data.empresa, data.telefono, data.email, data.direccion,
        data.ciudad, data.segmento, data.canal, data.estado or "contactado",
        data.responsable, data.notas,
        _parse_d(data.fecha_contacto) or date.today(),
        _parse_d(data.fecha_seguimiento), data.valor_potencial or 0
    ))
    new_id = cur.fetchone()[0]
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True, "id": new_id}


class ProspectoUpdate(BaseModel):
    nombre: Optional[str] = None
    empresa: Optional[str] = None
    telefono: Optional[str] = None
    email: Optional[str] = None
    direccion: Optional[str] = None
    ciudad: Optional[str] = None
    segmento: Optional[str] = None
    canal: Optional[str] = None
    estado: Optional[str] = None
    responsable: Optional[str] = None
    notas: Optional[str] = None
    fecha_contacto: Optional[str] = None
    fecha_seguimiento: Optional[str] = None
    valor_potencial: Optional[float] = None

@router.put("/prospectos/{prospecto_id}")
def update_prospecto(prospecto_id: int, data: ProspectoUpdate):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, nombre, empresa, telefono, email, direccion, ciudad, segmento
        FROM crm_prospectos WHERE id = %s
    """, (prospecto_id,))
    pro = cur.fetchone()
    if not pro:
        cur.close(); conn.close()
        raise HTTPException(404, "Prospecto no encontrado")

    def _parse_d(s):
        if not s:
            return None
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except ValueError:
            return None

    sets = []
    vals = []
    for field in ("nombre", "empresa", "telefono", "email", "direccion", "ciudad",
                  "segmento", "canal", "estado", "responsable", "notas", "valor_potencial"):
        v = getattr(data, field)
        if v is not None:
            sets.append(f"{field}=%s"); vals.append(v)
    for field in ("fecha_contacto", "fecha_seguimiento"):
        v = getattr(data, field)
        if v is not None:
            parsed = _parse_d(v)
            sets.append(f"{field}=%s"); vals.append(parsed)
    if sets:
        sets.append("actualizado_en=NOW()")
        vals.append(prospecto_id)
        cur.execute(f"UPDATE crm_prospectos SET {', '.join(sets)} WHERE id=%s", vals)

    # Si se convierte, crear cliente en crm_clientes
    if data.estado == "convertido":
        cur.execute("SELECT nombre, empresa, telefono, email, direccion, ciudad, segmento FROM crm_prospectos WHERE id=%s", (prospecto_id,))
        p = cur.fetchone()
        if p:
            nombre_cli = p[0]
            cur.execute("""
                INSERT INTO crm_clientes (nombre, telefono, email, direccion, ciudad, segmento)
                VALUES (%s,%s,%s,%s,%s,%s) RETURNING id
            """, (nombre_cli, p[2], p[3], p[4], p[5], p[6]))
            cli_id = cur.fetchone()[0]
            cur.execute("""
                UPDATE crm_prospectos SET convertido_cliente_id=%s WHERE id=%s
            """, (cli_id, prospecto_id))

    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}


@router.get("/clientes-crm/{cliente_id}/actualizar-stats")
def actualizar_stats_cliente(cliente_id: int):
    """Recalcula ultima_compra, total_compras y num_facturas desde crm_facturas."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id FROM crm_clientes WHERE id = %s", (cliente_id,))
    if not cur.fetchone():
        cur.close(); conn.close()
        raise HTTPException(404, "Cliente no encontrado")

    cur.execute("""
        SELECT MAX(fecha), SUM(total), COUNT(*)
        FROM crm_facturas WHERE cliente_id = %s
    """, (cliente_id,))
    row = cur.fetchone()
    ultima_compra = row[0]
    total_compras = float(row[1] or 0)
    num_facturas = row[2] or 0

    cur.execute("""
        UPDATE crm_clientes
        SET ultima_compra=%s, total_compras=%s, num_facturas=%s, actualizado_en=NOW()
        WHERE id=%s
    """, (ultima_compra, total_compras, num_facturas, cliente_id))
    conn.commit()
    cur.close(); conn.close()
    return {
        "ok": True,
        "cliente_id": cliente_id,
        "ultima_compra": ultima_compra.isoformat() if ultima_compra else None,
        "total_compras": total_compras,
        "num_facturas": num_facturas,
    }
